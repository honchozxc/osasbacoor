import csv
import json
import logging
import os
import re
import traceback
import uuid
from smtplib import SMTPException
import pandas as pd
from venv import logger
from django import forms
from django.db import IntegrityError
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.models import Permission
from django.contrib.auth.views import PasswordChangeView
from django.core.exceptions import ValidationError
from django.core.files.storage import FileSystemStorage
from django.core.mail import send_mail
from django.core.paginator import PageNotAnInteger, Paginator, EmptyPage
from django.db import transaction
from django.db.models import Q, Count, F
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.dateparse import parse_date, parse_time
from django.utils.decorators import method_decorator
from django.utils.html import strip_tags
from django.utils.timesince import timesince
from django.views import View
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.views.decorators.http import require_POST, require_GET
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin, PermissionRequiredMixin
from django.urls import reverse_lazy, reverse
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl import load_workbook
from django.conf import settings
from reportlab.lib.utils import ImageReader

from . import models
from .models import CustomUser, UserActivityLog, Downloadable, AnnouncementImage, Announcement, AboutPageContent, \
    ComplaintImage, ComplaintDocument, Complaint, \
    FooterContent, StudentDisciplineContent, Scholarship, ScholarshipApplication, ScholarshipPageContent, \
    HomePageContent, StudentAdmission, AdmissionPageContent, NSTPStudentInfo, NSTPFile, \
    NSTPPageContent, Course, ClinicPageContent, OJTCompany, \
    OJTPageContent, Organization, Certificate, SDSPageContent, AccomplishmentRecord, \
    SupportingFile, Hearing

from .forms import CustomUserCreationForm, CustomAuthenticationForm, CustomUserUpdateForm, DownloadableForm, \
    CustomPasswordChangeForm, AccountInfoForm, UserProfileForm, AnnouncementForm, AnnouncementImageFormSet, \
    AboutPageForm, ComplaintForm, FooterContentForm, \
    StudentDisciplineForm, RegistrationForm, ScholarshipForm, ScholarshipApplicationForm, \
    ScholarshipApplicationEditForm, ScholarshipPageForm, HomePageForm, \
    StudentAdmissionForm, AdmissionPageForm, NSTPStudentInfoForm, NSTPFileForm, NSTPPageForm, CourseForm, \
    ClinicPageForm, OJTCompanyForm, OJTPageForm, \
    OrganizationCreateForm, OrganizationEditForm, SDSPageForm, AccomplishmentRecordForm
from django.contrib.auth import login, logout, update_session_auth_hash
from django.contrib import messages
from django.http import JsonResponse, FileResponse, Http404, HttpResponse, BadHeaderError, HttpResponseRedirect
from django.template.loader import render_to_string
from django.http import JsonResponse
from datetime import datetime, timedelta, date
from django.utils import timezone
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO

from .utils import generate_certificate_png


class LoginView(View):
    template_name = 'osas/login.html'
    form_class = CustomAuthenticationForm

    def get(self, request):
        if request.user.is_authenticated:
            return redirect('dashboard')
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = self.form_class(data=request.POST)

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            if form.is_valid():
                user = form.get_user()
                login(request, user)
                UserActivityLog.objects.create(user=user, activity="Logged in")
                return JsonResponse({
                    'success': True,
                    'redirect_url': reverse('dashboard')
                })
            else:
                errors = {}
                error_message = None

                if 'inactive' in form.errors.get('__all__', []):
                    error_message = self.form_class.error_messages['inactive']
                elif 'archived' in form.errors.get('__all__', []):
                    error_message = self.form_class.error_messages['archived']
                else:
                    non_field_errors = form.errors.get('__all__', [])
                    if non_field_errors:
                        error_message = non_field_errors[0]

                for field, error_list in form.errors.items():
                    if field != '__all__':
                        errors[field] = error_list[0]

                return JsonResponse({
                    'success': False,
                    'error': error_message or 'Invalid credentials',
                    'field_errors': errors
                }, status=400)

        if form.is_valid():
            user = form.get_user()
            login(request, user)
            UserActivityLog.objects.create(user=user, activity="Logged in")
            return redirect('dashboard')
        return render(request, self.template_name, {'form': form})


class RegistrationView(CreateView):
    model = CustomUser
    form_class = CustomUserCreationForm
    template_name = 'osas/register.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        footer_content = FooterContent.objects.first()
        context['footer_content'] = footer_content or FooterContent.objects.create()
        context['courses'] = Course.objects.all()

        if self.request.method == 'POST' and 'form_submitted' in self.request.session:
            context['form_submitted'] = True
            del self.request.session['form_submitted']
        return context

    def form_valid(self, form):
        # Handle file uploads
        if 'id_photo' in self.request.FILES:
            form.instance.id_photo = self.request.FILES['id_photo']
        if 'cor_photo' in self.request.FILES:
            form.instance.cor_photo = self.request.FILES['cor_photo']

        user = form.save(commit=False)
        user.is_active = True
        user.is_verified = False
        user.save()

        messages.success(
            self.request,
            "Registration successful! Your account is pending verification."
        )

        # Get display values
        user_type = form.cleaned_data['user_type']
        user_type_display = dict(CustomUser.USER_TYPE_CHOICES).get(user_type, '')
        position_display = dict(CustomUser.OSAS_POSITION_CHOICES).get(form.cleaned_data.get('osas_position'),
                                                                      'Not specified')

        # Handle course display properly
        course_value = form.cleaned_data.get('course')
        if course_value:
            if hasattr(course_value, 'name'):
                course_display = course_value.name
            else:
                course_display = str(course_value)
        else:
            course_display = ''

        # Format birth date for display
        birth_date = form.cleaned_data.get('birth_date')
        if birth_date:
            birth_date_display = birth_date.strftime('%B %d, %Y')
        else:
            birth_date_display = ''

        registration_data = {
            'first_name': form.cleaned_data['first_name'],
            'last_name': form.cleaned_data['last_name'],
            'user_type': str(user_type),
            'user_type_display': user_type_display,
            'email': form.cleaned_data['email'],
            'phone_number': form.cleaned_data['phone_number'],
            'username': form.cleaned_data.get('username', ''),
            'birth_date': birth_date_display,
            'address': form.cleaned_data.get('address', ''),
        }

        # Add user type specific fields
        if user_type == 14:  # Student
            registration_data.update({
                'student_number': form.cleaned_data.get('student_number', ''),
                'course': course_display,
                'year_level': form.cleaned_data.get('year_level', ''),
                'section': form.cleaned_data.get('section', ''),
            })
        else:  # OSAS staff (user_type 1-13)
            registration_data.update({
                'department': form.cleaned_data.get('department', ''),
                'position_display': position_display,
            })

        self.request.session['registration_data'] = registration_data
        self.request.session['form_submitted'] = True
        return redirect(self.request.path)

    def form_invalid(self, form):
        messages.error(self.request, "Registration failed. Please correct the errors below.")
        for field, errors in form.errors.items():
            for error in errors:
                field_label = form.fields[field].label if field in form.fields else field
                messages.error(self.request, f"{field_label}: {error}")
        return super().form_invalid(form)


def clear_registration_session(request):
    # Clear the registration data from session
    if 'registration_data' in request.session:
        del request.session['registration_data']
    if 'form_submitted' in request.session:
        del request.session['form_submitted']
    return redirect('home')


class LogoutView(View):
    def get(self, request):
        if request.user.is_authenticated:
            UserActivityLog.objects.create(user=request.user, activity="Logged out")
            logout(request)
        return redirect('home')


# -------------------------------------- Core(Landing Page, About, etc) Sections ---------------------------------------
class HomePageView(TemplateView):
    template_name = 'core/home.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        home_content = HomePageContent.objects.first()
        if not home_content:
            home_content = HomePageContent.objects.create()
        context['home_content'] = home_content

        now = timezone.now()
        twenty_four_hours_ago = now - timezone.timedelta(hours=24)

        # Get student count
        try:
            context['student_count'] = CustomUser.objects.filter(
                user_type=14,
                is_archived=False,
            ).count()
        except Exception as e:
            print(f"Error getting student count: {e}")
            context['student_count'] = 0

        try:
            about_content = AboutPageContent.objects.first()
            if about_content:
                context['courses_count'] = about_content.courses.count()
            else:
                context['courses_count'] = 0
        except Exception as e:
            print(f"Error getting courses count: {e}")
            context['courses_count'] = 0

        context['upcoming_events'] = Announcement.objects.filter(
            category='EVENT',
            event_date__gte=now,
            is_published=True,
            is_archived=False
        ).order_by('event_date')[:3]

        context['latest_announcements'] = Announcement.objects.filter(
            is_published=True,
            is_archived=False
        ).exclude(
            category='EVENT'
        ).order_by('-created_at')[:6]

        context['emergency_announcements'] = Announcement.objects.filter(
            category='EMERGENCY',
            is_published=True,
            is_archived=False,
            created_at__gte=twenty_four_hours_ago
        ).order_by('-created_at')

        # Class suspensions
        context['suspension_notices'] = Announcement.objects.filter(
            category='SUSPENSION',
            suspension_date__gte=now.date(),
            is_published=True,
            is_archived=False
        ).order_by('suspension_date')

        # Enrollment periods
        context['enrollment_periods'] = Announcement.objects.filter(
            category='ENROLLMENT',
            enrollment_end__gte=now,
            is_published=True,
            is_archived=False
        ).order_by('enrollment_start')

        # Retrieving Footer Information
        footer_content = FooterContent.objects.first()
        if not footer_content:
            footer_content = FooterContent.objects.create()
        context['footer_content'] = footer_content

        return context


class AllAnnouncementView(TemplateView):
    template_name = 'core/all_announcements.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get all published announcements (for counts and initial page)
        announcements_queryset = Announcement.objects.filter(
            is_published=True,
            is_archived=False
        ).order_by('-publish_date').select_related('author')

        # Paginate announcements
        page = self.request.GET.get('page', 1)
        paginator = Paginator(announcements_queryset, 9)

        try:
            announcements = paginator.page(page)
        except PageNotAnInteger:
            announcements = paginator.page(1)
        except EmptyPage:
            announcements = paginator.page(paginator.num_pages)

        # Calculate category counts
        category_counts = {
            'all': announcements_queryset.count(),
            'basic': announcements_queryset.filter(category='BASIC').count(),
            'enrollment': announcements_queryset.filter(category='ENROLLMENT').count(),
            'event': announcements_queryset.filter(category='EVENT').count(),
            'suspension': announcements_queryset.filter(category='SUSPENSION').count(),
            'emergency': announcements_queryset.filter(category='EMERGENCY').count(),
            'scholarship': announcements_queryset.filter(category='SCHOLARSHIP').count(),
        }

        context['announcements'] = announcements
        context['announcements_count'] = announcements_queryset.count()
        context['category_counts'] = category_counts

        # Prepare JSON data for JavaScript (using PAGINATED data only)
        announcements_data = []
        for ann in announcements:
            first_image = ann.get_first_image()
            announcements_data.append({
                'id': ann.id,
                'title': ann.title,
                'content': ann.content,
                'category': ann.category,
                'category_display': ann.get_category_display(),
                'publish_date': ann.publish_date.isoformat(),
                'author_full_name': ann.author.get_full_name(),
                'author_unit': ann.author.user_type,
                'author_unit_display': ann.author.get_user_type_display(),
                'link': ann.link,
                'first_image': {
                    'url': first_image.image.url if first_image else '',
                    'caption': first_image.caption if first_image else ''
                },
                'courses_display': ann.get_unique_courses_display(),
                'enrollment_start': ann.enrollment_start.isoformat() if ann.enrollment_start else None,
                'enrollment_end': ann.enrollment_end.isoformat() if ann.enrollment_end else None,
                'event_date': ann.event_date.isoformat() if ann.event_date else None,
                'location': ann.location,
                'suspension_date': ann.suspension_date.isoformat() if ann.suspension_date else None,
                'until_suspension_date': ann.until_suspension_date.isoformat() if ann.until_suspension_date else None,
                'contact_info': ann.contact_info
            })

        context['announcements_json'] = json.dumps(announcements_data)
        context['now'] = timezone.now()
        context['all_categories'] = dict(Announcement.CATEGORY_CHOICES)
        context['user_units'] = dict(CustomUser.USER_TYPE_CHOICES)

        # Retrieving Footer Information
        footer_content = FooterContent.objects.first()
        if not footer_content:
            footer_content = FooterContent.objects.create()
        context['footer_content'] = footer_content

        return context


def all_announcements_api(request):
    announcements = Announcement.objects.filter(
        is_published=True,
        is_archived=False
    ).order_by('-publish_date').select_related('author')

    data = []
    for ann in announcements:
        first_image = ann.get_first_image()
        data.append({
            'id': ann.id,
            'title': ann.title,
            'content': ann.content,
            'category': ann.category,
            'category_display': ann.get_category_display(),
            'publish_date': ann.publish_date.isoformat(),
            'author_full_name': ann.author.get_full_name(),
            'author_unit': ann.author.user_type,
            'author_unit_display': ann.author.get_user_type_display(),
            'link': ann.link,
            'first_image': {
                'url': first_image.image.url if first_image else '',
                'caption': first_image.caption if first_image else ''
            },
            'courses_display': ann.get_unique_courses_display(),
            'enrollment_start': ann.enrollment_start.isoformat() if ann.enrollment_start else None,
            'enrollment_end': ann.enrollment_end.isoformat() if ann.enrollment_end else None,
            'event_date': ann.event_date.isoformat() if ann.event_date else None,
            'location': ann.location,
            'suspension_date': ann.suspension_date.isoformat() if ann.suspension_date else None,
            'until_suspension_date': ann.until_suspension_date.isoformat() if ann.until_suspension_date else None,
            'contact_info': ann.contact_info
        })

    return JsonResponse(data, safe=False)


class HomeAnnouncementDetailView(DetailView):
    model = Announcement
    template_name = 'core/announcement_detail.html'
    context_object_name = 'announcement'

    def get_queryset(self):
        return super().get_queryset().prefetch_related('images').select_related('author')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        announcement = context['announcement']

        context['related_announcements'] = Announcement.objects.filter(
            category=announcement.category,
            is_published=True,
            is_archived=False
        ).exclude(
            id=announcement.id
        ).order_by('-created_at')[:3].prefetch_related('images')

        context['recent_announcements'] = Announcement.objects.filter(
            is_published=True,
            is_archived=False
        ).exclude(
            id=announcement.id
        ).order_by('-created_at')[:5].prefetch_related('images')

        # Retrieving Footer Information
        footer_content = FooterContent.objects.first()
        if not footer_content:
            footer_content = FooterContent.objects.create()
        context['footer_content'] = footer_content

        return context


class HomePageEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = HomePageContent
    form_class = HomePageForm
    template_name = 'osas/modals/home_edit.html'
    success_url = reverse_lazy('dashboard')

    def test_func(self):
        return self.request.user.is_superuser or self.request.user.user_type == 1  # Only OSAS admins

    def get_object(self):
        # Get or create the home page content
        obj, created = HomePageContent.objects.get_or_create(pk=1)
        return obj

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)


class AboutPageView(TemplateView):
    template_name = 'core/about.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        try:
            # Get or create about page content
            about_content, created = AboutPageContent.objects.get_or_create(
                defaults={
                    'title': "OSAS Bacoor City Campus",
                    'tagline': "Shaping Future Leaders with Character and Excellence",
                    'about_text': "The Office of Student Affairs and Services (OSAS) at Cavite State University Bacoor City Campus is committed to providing comprehensive support services that enhance student development, welfare, and overall campus experience. We strive to create a nurturing environment that fosters academic excellence, personal growth, and social responsibility.",
                    'mission': "Cavite State University shall provide excellent, equitable, and relevant educational opportunities in the arts, sciences, and technology through quality instruction and responsive research and development activities. It shall produce professional, skilled, and morally upright individuals for global competitiveness.",
                    'vision': "The premier university in historic Cavite recognized for excellence in the development of globally competitive and morally upright individuals.",
                    'goals': "To look after the educational, vocational, and personal development needs of students through various services and programs that complement the academic curriculum and prepare students for responsible citizenship and productive careers.",
                    'objectives': [
                        "Provide counseling and testing services to support student mental health and academic success",
                        "Offer comprehensive information, training, placement, and follow-up services for career development",
                        "Administer financial assistance programs for needy and deserving students",
                        "Provide diverse avenues to hone students' talents and interests through extracurricular activities",
                        "Develop skills and potentials of student writers and campus journalists",
                        "Ensure student welfare through responsive services and support systems",
                        "Foster leadership skills and social responsibility among students",
                        "Promote cultural awareness and artistic expression within the campus community"
                    ]
                }
            )

            # Create default courses if about content was just created
            if created:
                default_courses = [
                    {"name": "Bachelor of Secondary Education", "subtext": "Major in English, Mathematics, Science"},
                    {"name": "BS Business Management", "subtext": "Entrepreneurship and Marketing Track"},
                    {"name": "BS Computer Science", "subtext": "Software Engineering and Data Science"},
                    {"name": "BS Criminology", "subtext": "Law Enforcement Administration"},
                    {"name": "BS Hospitality Management", "subtext": "formerly BS Hotel and Restaurant Management"},
                    {"name": "BS Information Technology", "subtext": "Network Administration and Web Development"},
                    {"name": "BS Psychology", "subtext": "Clinical and Industrial Psychology"}
                ]

                for course_data in default_courses:
                    course, _ = Course.objects.get_or_create(
                        name=course_data['name'],
                        defaults={'subtext': course_data['subtext']}
                    )
                    about_content.courses.add(course)

            # Get footer content
            footer_content, _ = FooterContent.objects.get_or_create()
            context.update({
                'about_content': about_content,
                'courses': about_content.courses.all().order_by('name'),
                'footer_content': footer_content,
            })

        except Exception as e:
            # Log the error and provide fallback data
            if settings.DEBUG:
                print(f"Error in AboutPageView: {e}")

            # Fallback context data
            context.update({
                'about_content': None,
                'courses': Course.objects.all().order_by('name')[:7],
                'footer_content': None,
                'error_message': 'Unable to load about page data.'
            })

        return context


class AboutPageEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = AboutPageContent
    form_class = AboutPageForm
    template_name = 'osas/modals/about_edit.html'
    success_url = reverse_lazy('dashboard')

    def test_func(self):
        return self.request.user.is_superuser or self.request.user.user_type == 1

    def get_object(self):
        return AboutPageContent.objects.first()

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()

        if form.is_valid():
            about_page = form.save(commit=False)
            about_page.updated_by = request.user
            about_page.save()

            deleted_courses = request.POST.getlist('deleted_courses', [])
            for course_id in deleted_courses:
                try:
                    course = Course.objects.get(id=course_id)
                    course.delete()
                except Course.DoesNotExist:
                    pass

            existing_courses = request.POST.getlist('existing_courses', [])
            for course_id in existing_courses:
                if course_id not in deleted_courses:
                    try:
                        course = Course.objects.get(id=course_id)
                        course.name = request.POST.get(f'course_name_{course_id}')
                        course.subtext = request.POST.get(f'course_subtext_{course_id}')

                        # Handle logo deletion
                        if f'course_logo_clear_{course_id}' in request.POST:
                            if course.logo:
                                course.logo.delete()
                                course.logo = None

                        # Handle new logo upload
                        logo_file = request.FILES.get(f'course_logo_{course_id}')
                        if logo_file:
                            if course.logo:
                                course.logo.delete()
                            course.logo = logo_file

                        course.save()
                    except Course.DoesNotExist:
                        pass

            # Process new courses
            new_courses = request.POST.getlist('new_courses', [])
            for course_counter in new_courses:
                name = request.POST.get(f'new_course_name_{course_counter}')
                subtext = request.POST.get(f'new_course_subtext_{course_counter}')
                logo_file = request.FILES.get(f'new_course_logo_{course_counter}')

                if name:
                    course = Course.objects.create(
                        name=name,
                        subtext=subtext,
                        logo=logo_file
                    )
                    self.object.courses.add(course)

            return redirect(self.success_url)

        return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['objectives_json'] = json.dumps(self.object.objectives)
        context['courses'] = self.object.courses.all()
        return context


class TemplatePageView(TemplateView):
    template_name = 'core/downloadables.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get filter parameters from request
        search_term = self.request.GET.get('search', '').strip()
        category_filter = self.request.GET.get('category', 'all')
        date_filter = self.request.GET.get('date', 'latest')

        # Base queryset
        downloadables = Downloadable.objects.filter(
            is_active=True,
            is_archived=False
        )

        if search_term:
            query = Q()

            query |= Q(title__icontains=search_term)
            query |= Q(description__icontains=search_term)

            if any(char.isdigit() for char in search_term):
                words = search_term.split()
                for word in words:
                    if word.isdigit():
                        query |= Q(title__icontains=word)
                        query |= Q(description__icontains=word)

            if search_term.isdigit():
                query |= Q(title__icontains=search_term)
                query |= Q(description__icontains=search_term)

            downloadables = downloadables.filter(query)

        # Apply category filter
        if category_filter != 'all':
            downloadables = downloadables.filter(category=category_filter)

        # Apply date filter and ordering
        if date_filter == 'latest':
            downloadables = downloadables.order_by('-created_at')
        elif date_filter == 'oldest':
            downloadables = downloadables.order_by('created_at')
        elif date_filter == 'this_week':
            start_of_week = timezone.now() - timedelta(days=timezone.now().weekday())
            downloadables = downloadables.filter(created_at__gte=start_of_week)
        elif date_filter == 'this_month':
            start_of_month = timezone.now().replace(day=1)
            downloadables = downloadables.filter(created_at__gte=start_of_month)
        elif date_filter == 'this_year':
            start_of_year = timezone.now().replace(month=1, day=1)
            downloadables = downloadables.filter(created_at__gte=start_of_year)

        # Get total counts before pagination
        total_count = downloadables.count()

        # Get counts for each category from the filtered queryset
        osas_forms_count = downloadables.filter(category='osas_forms').count()
        society_forms_count = downloadables.filter(category='society_forms').count()
        ojt_forms_count = downloadables.filter(category='ojt_forms').count()
        guidelines_count = downloadables.filter(category='guidelines').count()
        manuals_count = downloadables.filter(category='manuals').count()
        others_count = downloadables.filter(category='others').count()

        # Get page number from request
        page_number = self.request.GET.get('page', 1)

        # Create paginator with 10 items per page
        paginator = Paginator(downloadables, 10)

        try:
            page_obj = paginator.get_page(page_number)
        except (PageNotAnInteger, EmptyPage):
            page_obj = paginator.get_page(1)

        context.update({
            'downloadables': page_obj,
            'total_count': total_count,
            'osas_forms_count': osas_forms_count,
            'society_forms_count': society_forms_count,
            'ojt_forms_count': ojt_forms_count,
            'guidelines_count': guidelines_count,
            'manuals_count': manuals_count,
            'others_count': others_count,
            'current_search': search_term,
            'current_category': category_filter,
            'current_date_filter': date_filter,
            'paginator': paginator,
            'page_obj': page_obj,
        })

        # Retrieving Footer Information
        footer_content = FooterContent.objects.first()
        if not footer_content:
            footer_content = FooterContent.objects.create()
        context['footer_content'] = footer_content

        return context


class StudentDisciplinePageView(TemplateView):
    template_name = 'core/student_discipline_unit.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        discipline_content = StudentDisciplineContent.objects.first()

        if not discipline_content:
            # Create default content if none exists
            discipline_content = StudentDisciplineContent.objects.create(
                principles_items=[
                    "Fair and impartial treatment for all students",
                    "Educational rather than punitive approach",
                    "Respect for individual rights and dignity",
                    "Transparency in procedures"
                ],
                approach_items=[
                    "Initial assessment of reported incidents",
                    "Fact-finding and evidence gathering",
                    "Meeting with involved parties",
                    "Determination of appropriate resolution"
                ],
                complaint_types=[
                    {
                        "title": "Academic Misconduct",
                        "icon": "fa-graduation-cap",
                        "items": [
                            "Cheating on exams",
                            "Plagiarism",
                            "Unauthorized collaboration"
                        ]
                    },
                    {
                        "title": "Behavioral Issues",
                        "icon": "fa-users",
                        "items": [
                            "Disruptive behavior",
                            "Harassment",
                            "Bullying"
                        ]
                    }
                ],
                how_to_file_steps=[
                    {
                        "title": "Step 1: Gather Information",
                        "icon": "fa-clipboard-list",
                        "description": "Collect all relevant details about the incident"
                    },
                    {
                        "title": "Step 2: Submit Report",
                        "icon": "fa-file-upload",
                        "description": "Complete the online reporting form"
                    }
                ],
                process_notes=[
                    {
                        "title": "Confidentiality",
                        "icon": "fa-lock",
                        "description": "All reports are handled with strict confidentiality"
                    },
                    {
                        "title": "Timeline",
                        "icon": "fa-clock",
                        "description": "Cases are typically resolved within 30 days"
                    }
                ]
            )

        context['discipline_content'] = discipline_content

        # Retrieving Footer Information
        footer_content = FooterContent.objects.first()
        if not footer_content:
            footer_content = FooterContent.objects.create()
        context['footer_content'] = footer_content

        return context


class StudentDisciplineEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = StudentDisciplineContent
    form_class = StudentDisciplineForm
    template_name = 'osas/modals/student_discipline_unit.html'
    success_url = reverse_lazy('dashboard')

    def test_func(self):
        return self.request.user.is_superuser or self.request.user.user_type == 1 or self.request.user.user_type == 11

    def get_object(self):
        return StudentDisciplineContent.objects.first()

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['principles_items_json'] = json.dumps(self.object.principles_items)
        context['approach_items_json'] = json.dumps(self.object.approach_items)
        context['complaint_types_json'] = json.dumps(self.object.complaint_types)
        context['how_to_file_steps_json'] = json.dumps(self.object.how_to_file_steps)
        context['process_notes_json'] = json.dumps(self.object.process_notes)
        return context


class OJTView(TemplateView):
    template_name = 'core/ojt.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get or create OJT page content
        ojt_content = OJTPageContent.objects.first()
        if not ojt_content:
            ojt_content = OJTPageContent.objects.create()

        top_companies = OJTCompany.objects.filter(
            is_archived=False,
            status='active'
        ).order_by('-created_at')[:4]

        # You can also get some statistics if needed
        total_companies = OJTCompany.objects.filter(is_archived=False).count()
        active_companies = OJTCompany.objects.filter(is_archived=False, status='active').count()
        inactive_companies = OJTCompany.objects.filter(is_archived=False, status='inactive').count()

        context.update({
            'ojt_content': ojt_content,
            'top_companies': top_companies,
            'total_companies': total_companies,
            'active_companies': active_companies,
            'inactive_companies': inactive_companies,
            'footer_content': FooterContent.objects.first() or FooterContent.objects.create()
        })

        return context


class OJTPageEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = OJTPageContent
    form_class = OJTPageForm
    template_name = 'osas/modals/ojt_edit.html'
    success_url = reverse_lazy('dashboard')

    def test_func(self):
        return self.request.user.is_superuser or self.request.user.user_type == 1 or self.request.user.user_type == 13

    def get_object(self):
        return OJTPageContent.objects.first()

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['overview_cards_json'] = json.dumps(self.object.overview_cards)
        context['services_cards_json'] = json.dumps(self.object.services_cards)
        context['process_steps_json'] = json.dumps(self.object.process_steps)
        context['faq_content_json'] = json.dumps(self.object.faq_content)
        return context


class ScholarshipPageView(TemplateView):
    template_name = 'core/scholarship.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Add user information to context
        context['user'] = self.request.user

        # Get or create scholarship page content
        scholarship_content = ScholarshipPageContent.objects.first()
        if not scholarship_content:
            scholarship_content = ScholarshipPageContent.objects.create(
                faq_content=[
                    {
                        "question": "How do I apply for a scholarship?",
                        "answer": "To apply for a scholarship, follow these steps:\n1. Check the available scholarships and their requirements\n2. Download the corresponding application form\n3. Complete the form and gather all required documents\n4. Submit your application to the Scholarship Office\n5. Wait for the evaluation results"
                    },
                    {
                        "question": "When is the scholarship application deadline?",
                        "answer": "Deadlines vary depending on the scholarship program. Each scholarship listing includes its specific deadline information. For university-managed scholarships, applications are typically accepted at the beginning of each semester."
                    }
                ]
            )

        # Get scholarships data
        public_scholarships = Scholarship.objects.filter(
            is_active=True,
            is_archived=False,
            scholarship_type='public'
        ).order_by('-created_at')[:2]

        private_scholarships = Scholarship.objects.filter(
            is_active=True,
            is_archived=False,
            scholarship_type='private'
        ).order_by('-created_at')[:2]

        total_scholarships = Scholarship.objects.filter(
            is_active=True,
            is_archived=False
        ).count()

        context.update({
            'scholarship_content': scholarship_content,
            'public_scholarships': public_scholarships,
            'private_scholarships': private_scholarships,
            'total_scholarships': total_scholarships,
            'footer_content': FooterContent.objects.first() or FooterContent.objects.create()
        })

        return context


class ScholarshipPageEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = ScholarshipPageContent
    form_class = ScholarshipPageForm
    template_name = 'osas/modals/scholarship_edit.html'
    success_url = reverse_lazy('dashboard')

    def test_func(self):
        return self.request.user.is_superuser or self.request.user.user_type == 1 or self.request.user.user_type == 5

    def get_object(self):
        return ScholarshipPageContent.objects.first()

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['faq_json'] = json.dumps(self.object.faq_content)
        return context


class AdmissionLandingView(TemplateView):
    template_name = 'core/admission.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user'] = self.request.user

        # Get admission content with defaults
        admission_content = AdmissionPageContent.objects.first()
        if not admission_content:
            admission_content = AdmissionPageContent.objects.create(
                hero_title="Admission Portal",
                hero_tagline="Start your journey with us today",
                cta_text="Ready to begin your academic journey?",
                cta_description="Join our community of scholars and innovators",
                requirements={
                    'shs_graduate': {
                        'title': 'Senior High School Graduate Requirements',
                        'items': [
                            'Original Grade 12 Report Card (Form 138)',
                            'Certificate of Good Moral Character',
                            '2x2 ID Picture (white background)',
                            'PSA Birth Certificate'
                        ],
                        'note': 'All documents must be submitted in original copies'
                    },
                    'current_grade12': {
                        'title': 'Current Grade 12 Student Requirements',
                        'items': [
                            'Certificate of Registration as Grade 12 Student',
                            'Most Recent Report Card',
                            'Certificate of Good Moral Character'
                        ],
                        'note': 'Final Report Card must be submitted upon graduation'
                    },
                    'transferee': {
                        'title': 'Transferee Requirements',
                        'items': [
                            'Original Transcript of Records',
                            'Honorable Dismissal',
                            'Certificate of Good Moral Character',
                            '2x2 ID Picture (white background)'
                        ],
                        'note': 'Transfer credentials must be from CHED-recognized institutions'
                    }
                }
            )

        # Get courses from AboutPage or create default
        about_page = AboutPageContent.objects.first()
        if not about_page:
            about_page = AboutPageContent.objects.create()
            default_courses = [
                {"name": "Bachelor of Secondary Education", "code": "BSEd"},
                {"name": "BS Business Management", "code": "BSBM"},
                {"name": "BS Computer Science", "code": "BSCS"},
                {"name": "BS Criminology", "code": "BSCrim"},
                {"name": "BS Hospitality Management", "code": "BSHM"},
                {"name": "BS Information Technology", "code": "BSIT"},
                {"name": "BS Psychology", "code": "BSPsych"}
            ]
            for course_data in default_courses:
                course, _ = Course.objects.get_or_create(
                    name=course_data['name'],
                    defaults={'code': course_data['code']}
                )
                about_page.courses.add(course)

        courses = about_page.courses.all()

        # Get school info
        home_page = HomePageContent.objects.first()
        school_name = home_page.title if home_page else "Cavite State University - Bacoor City Campus"
        tagline = home_page.tagline if home_page else "Office of Student Affairs and Services"

        # Get or create footer content
        footer_content, _ = FooterContent.objects.get_or_create()

        context.update({
            'admission_content': admission_content,
            'courses': courses,
            'school_name': school_name,
            'tagline': tagline,
            'footer_content': footer_content
        })

        return context


class AdmissionPageEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = AdmissionPageContent
    form_class = AdmissionPageForm
    template_name = 'osas/modals/admission_edit.html'
    success_url = reverse_lazy('dashboard')

    def test_func(self):
        return self.request.user.is_superuser or self.request.user.user_type == 1  or self.request.user.user_type == 12

    def get_object(self):
        # Get the first record or create a default one if none exists
        obj, created = AdmissionPageContent.objects.get_or_create(
            defaults={
                'hero_title': "ADMISSION S.Y. 2025-2026",
                'hero_tagline': "Shape your future with Cavite State University - Bacoor City Campus",
                'cta_text': "READY TO BEGIN YOUR JOURNEY?",
                'cta_description': "Take the first step towards your future at Cavite State University - Bacoor City Campus. Our admissions team is ready to assist you.",
                'requirements': {
                    "shs_graduate": {
                        "title": "Requirements for 1st Year (SHS Graduate)",
                        "items": [
                            "Accomplished Admission Portal registration (Online)",
                            "Completed Grade 12 Report Card",
                            "Certificate of non-issuance of Form 137"
                        ],
                        "note": "This certification shall provide that the applicant has never been enrolled in another university/college.<br><br><strong>\"ALIGNMENT OF STRANDS WITH THE RESPECTIVE PROGRAMS IS STRICTLY OBSERVED AND FOLLOWED S.Y 2025-2026\"</strong>"
                    },
                    "current_grade12": {
                        "title": "Requirements for 1st Year (Current G12 Students)",
                        "items": [
                            "Accomplished Admission Portal registration (Online)",
                            "Grade 11 Report Card",
                            "Certificate from the Principal or Adviser indicating current enrollment as Grade 12 student with STRAND indicated"
                        ],
                        "note": "The certificate must be originally signed. E-signature is NOT ALLOWED.<br><br><strong>\"ALIGNMENT OF STRANDS WITH THE RESPECTIVE PROGRAMS IS STRICTLY OBSERVED AND FOLLOWED S.Y 2025-2026\"</strong>"
                    },
                    "transferee": {
                        "title": "Requirements for Transferees",
                        "items": [
                            "Accomplished Admission Portal registration (Online)",
                            "Transcript of Grades or Certificate of Grades",
                            "Certificate of Good Moral Character",
                            "Honorable Dismissal",
                            "NBI or Police Clearance"
                        ],
                        "note": "<strong>\"With Retention Policy/Grades Evaluation Only\"</strong><br><br>For S.Y 2025-2026 1st SEMESTER"
                    }
                }
            }
        )
        return obj

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['requirements_json'] = json.dumps(self.object.requirements if self.object.requirements else {})
        return context


class NSTPLandingView(TemplateView):
    template_name = 'core/nstp.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user'] = self.request.user

        # Get or create NSTP content
        nstp_content = NSTPPageContent.objects.first()
        if not nstp_content:
            nstp_content = NSTPPageContent.objects.create(
                programs=[
                    {
                        "title": "Reserve Officers' Training Corps (ROTC)",
                        "description": "A program institutionalized under Sections 38 and 39 of Republic Act No. 7077...",
                        "icon": "fas fa-shield-alt"
                    },
                    {
                        "title": "Civic Welfare Training Service (CWTS)",
                        "description": "Activities contributory to the general welfare and the betterment of life...",
                        "icon": "fas fa-book-open"
                    }
                ],
                faqs=[
                    {
                        "question": "Who is required to take NSTP?",
                        "answer": "All incoming college students, male and female, enrolled in baccalaureate..."
                    },
                    {
                        "question": "How many units is NSTP?",
                        "answer": "The NSTP is a 3-unit course for two semesters..."
                    }
                ]
            )

        context['nstp_content'] = nstp_content

        footer_content = FooterContent.objects.first()
        if not footer_content:
            footer_content = FooterContent.objects.create()
        context['footer_content'] = footer_content

        return context


class NSTPPageEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = NSTPPageContent
    form_class = NSTPPageForm
    template_name = 'osas/modals/edit_nstp_page.html'
    success_url = reverse_lazy('dashboard')

    def test_func(self):
        return self.request.user.is_superuser or self.request.user.user_type == 1 or self.request.user.user_type == 2

    def get_object(self):
        obj, created = NSTPPageContent.objects.get_or_create(
            defaults={
                'programs': [
                    {
                        "title": "Reserve Officers' Training Corps (ROTC)",
                        "description": "A program institutionalized under Sections 38 and 39 of Republic Act No. 7077...",
                        "icon": "fas fa-shield-alt"
                    },
                    {
                        "title": "Civic Welfare Training Service (CWTS)",
                        "description": "Activities contributory to the general welfare and the betterment of life...",
                        "icon": "fas fa-book-open"
                    }
                ],
                'faqs': [
                    {
                        "question": "Who is required to take NSTP?",
                        "answer": "All incoming college students, male and female, enrolled in baccalaureate..."
                    },
                    {
                        "question": "How many units is NSTP?",
                        "answer": "The NSTP is a 3-unit course for two semesters..."
                    }
                ]
            }
        )
        return obj

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        response = super().form_valid(form)
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'redirect_url': str(self.success_url)
            })
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['programs_json'] = json.dumps(self.object.programs if self.object.programs else [])
        context['faqs_json'] = json.dumps(self.object.faqs if self.object.faqs else [])
        return context


class ClinicView(TemplateView):
    template_name = 'core/clinic.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get clinic content
        clinic_content = ClinicPageContent.objects.first()
        if not clinic_content:
            clinic_content = ClinicPageContent.objects.create()

        # Get footer content
        footer_content = FooterContent.objects.first()
        if not footer_content:
            footer_content = FooterContent.objects.create()

        context.update({
            'clinic_content': clinic_content,
            'footer_content': footer_content,
        })
        return context


class ClinicPageEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = ClinicPageContent
    form_class = ClinicPageForm
    template_name = 'osas/modals/clinic_edit.html'
    success_url = reverse_lazy('dashboard')

    def test_func(self):
        return self.request.user.is_superuser or self.request.user.user_type == 1 or self.request.user.user_type == 3

    def get_object(self):
        return ClinicPageContent.objects.first()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        clinic_content = self.get_object()

        context['clinic_content'] = {
            'services': clinic_content.services,
            'faqs': clinic_content.faqs,
            'gallery_images': clinic_content.gallery_images,
            'hero_title': clinic_content.hero_title,
            'hero_description': clinic_content.hero_description,
            'phone': clinic_content.phone,
            'email': clinic_content.email,
            'address': clinic_content.address,
        }
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()

        if form.is_valid():
            clinic_page = form.save(commit=False)
            clinic_page.updated_by = request.user

            services_data = []
            service_counters = request.POST.getlist('service_counters[]')

            for counter in service_counters:
                icon = request.POST.get(f'service_icon_{counter}', '⚕️')
                title = request.POST.get(f'service_title_{counter}', '')
                description = request.POST.get(f'service_description_{counter}', '')

                if title and description:
                    services_data.append({
                        'icon': icon,
                        'title': title,
                        'description': description
                    })

            clinic_page.services = services_data

            # Process FAQs
            faqs_data = []
            faq_counters = request.POST.getlist('faq_counters[]')
            for counter in faq_counters:
                question = request.POST.get(f'faq_question_{counter}', '')
                answer = request.POST.get(f'faq_answer_{counter}', '')

                if question and answer:
                    faqs_data.append({
                        'question': question,
                        'answer': answer
                    })

            clinic_page.faqs = faqs_data

            gallery_data = []
            gallery_counters = request.POST.getlist('gallery_counters[]')

            for counter in gallery_counters:
                image_file = request.FILES.get(f'gallery_image_{counter}')
                existing_image = request.POST.get(f'existing_gallery_image_{counter}', '')
                alt_text = request.POST.get(f'gallery_alt_{counter}', '')

                # Use new image if uploaded, otherwise keep existing
                image_path = existing_image
                if image_file and image_file.name:
                    # Simple file save
                    try:
                        import os
                        from django.conf import settings
                        import uuid

                        # Create directory
                        directory = os.path.join(settings.MEDIA_ROOT, 'clinic/gallery/')
                        os.makedirs(directory, exist_ok=True)

                        # Generate unique filename
                        file_extension = os.path.splitext(image_file.name)[1]
                        unique_filename = f"{uuid.uuid4().hex}{file_extension}"
                        file_path = os.path.join(directory, unique_filename)

                        # Save file
                        with open(file_path, 'wb+') as destination:
                            for chunk in image_file.chunks():
                                destination.write(chunk)

                        image_path = f"media/clinic/gallery/{unique_filename}"

                    except Exception as e:
                        print(f"Error saving image: {e}")
                        # If upload fails, keep existing image
                        image_path = existing_image

                if image_path and alt_text:
                    gallery_data.append({
                        'image': image_path,
                        'alt': alt_text
                    })

            clinic_page.gallery_images = gallery_data

            clinic_page.save()
            return redirect(self.success_url)

        return self.form_invalid(form)


class SDSOrganizationView(TemplateView):
    template_name = 'core/sdsorganization.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get SDS page content
        sds_content = SDSPageContent.objects.first()
        if not sds_content:
            sds_content = SDSPageContent.objects.create()

        # Get all active organizations
        active_organizations = Organization.objects.filter(
            _organization_status='active'
        ).select_related('user_account')

        student_org_count = active_organizations.filter(organization_type='student').count()
        sociocultural_org_count = active_organizations.filter(organization_type='sociocultural').count()
        total_members = sum(org.organization_member_count for org in active_organizations)
        total_active_certificates = Certificate.objects.filter(organization__in=active_organizations,is_active=True).count()
        organizations_list = list(active_organizations)
        organizations_list.sort(key=lambda x: x.organization_member_count, reverse=True)

        # Pagination
        page = self.request.GET.get('page', 1)
        paginator = Paginator(organizations_list, 6)

        try:
            organizations_page = paginator.page(page)
        except PageNotAnInteger:
            organizations_page = paginator.page(1)
        except EmptyPage:
            organizations_page = paginator.page(paginator.num_pages)

        # Get footer content
        footer_content = FooterContent.objects.first()
        if not footer_content:
            footer_content = FooterContent.objects.create()

        context.update({
            'sds_content': sds_content,
            'organizations': organizations_page,
            'paginator': paginator,
            'student_org_count': student_org_count,
            'sociocultural_org_count': sociocultural_org_count,
            'total_members': total_members,
            'total_active_certificates': total_active_certificates,
            'footer_content': footer_content,
        })
        return context


class SDSPageEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = SDSPageContent
    form_class = SDSPageForm
    template_name = 'osas/modals/sds_edit.html'
    success_url = reverse_lazy('dashboard')

    def test_func(self):
        return self.request.user.is_superuser or self.request.user.user_type == 1

    def get_object(self):
        return SDSPageContent.objects.first()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sds_content = self.get_object()

        context['sds_content'] = {
            'features': sds_content.features,
            'faqs': sds_content.faqs,
            'hero_title': sds_content.hero_title,
            'hero_subtitle': sds_content.hero_subtitle,
            'hero_badge_text': sds_content.hero_badge_text,
            'section_subtitle': sds_content.section_subtitle,
            'mission_title': sds_content.mission_title,
            'mission_content': sds_content.mission_content,
            'what_we_do_title': sds_content.what_we_do_title,
            'what_we_do_content': sds_content.what_we_do_content,
            'cta_title': sds_content.cta_title,
            'cta_content': sds_content.cta_content,
        }
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()

        if form.is_valid():
            sds_page = form.save(commit=False)
            sds_page.updated_by = request.user

            # Process Features
            features_data = []
            feature_counters = request.POST.getlist('feature_counters[]')
            for counter in feature_counters:
                feature_text = request.POST.get(f'feature_text_{counter}', '')
                if feature_text:
                    features_data.append({
                        'text': feature_text
                    })
            sds_page.features = features_data

            # Process FAQs
            faqs_data = []
            faq_counters = request.POST.getlist('faq_counters[]')
            for counter in faq_counters:
                question = request.POST.get(f'faq_question_{counter}', '')
                subtitle = request.POST.get(f'faq_subtitle_{counter}', '')
                answer = request.POST.get(f'faq_answer_{counter}', '')

                # Process tips
                tips_data = []
                tip_counters = request.POST.getlist(f'faq_{counter}_tip_counters[]')
                for tip_counter in tip_counters:
                    tip_text = request.POST.get(f'faq_{counter}_tip_{tip_counter}', '')
                    if tip_text:
                        tips_data.append(tip_text)

                # Process benefits
                benefits_data = []
                benefit_counters = request.POST.getlist(f'faq_{counter}_benefit_counters[]')
                for benefit_counter in benefit_counters:
                    benefit_icon = request.POST.get(f'faq_{counter}_benefit_icon_{benefit_counter}', '')
                    benefit_text = request.POST.get(f'faq_{counter}_benefit_text_{benefit_counter}', '')
                    if benefit_icon and benefit_text:
                        benefits_data.append({
                            'icon': benefit_icon,
                            'text': benefit_text
                        })

                if question and answer:
                    faqs_data.append({
                        'question': question,
                        'subtitle': subtitle,
                        'answer': answer,
                        'tips': tips_data,
                        'benefits': benefits_data
                    })

            sds_page.faqs = faqs_data
            sds_page.save()
            return redirect(self.success_url)

        return self.form_invalid(form)


class CoreOrganizationDetailView(TemplateView):
    template_name = 'core/organization_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        organization_id = self.kwargs.get('pk')

        # Get organization by ID
        organization = get_object_or_404(
            Organization.objects.select_related('user_account'),
            id=organization_id
        )

        # Get members from JSON field and paginate them
        all_members = organization.organization_members or []
        valid_members = [
            member for member in all_members
            if isinstance(member, dict) and member.get('first_name') and member.get('last_name')
        ]

        paginator = Paginator(valid_members, 6)

        page_number = self.request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        # Get footer content
        footer_content = FooterContent.objects.first()
        if not footer_content:
            footer_content = FooterContent.objects.create()

        context.update({
            'organization': organization,
            'footer_content': footer_content,
            'page_obj': page_obj,
            'members': page_obj.object_list,
            'paginator': paginator,
        })
        return context


class OSASOrganizationChartView(TemplateView):
    template_name = 'core/organization-chart.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        osas_users = CustomUser.objects.filter(
            osas_position__isnull=False
        ).exclude(
            osas_position=''
        ).select_related('course')

        context['osas_users'] = osas_users

        footer_content = FooterContent.objects.first()
        if not footer_content:
            footer_content = FooterContent.objects.create()
        context['footer_content'] = footer_content

        return context


class FooterView(TemplateView):
    template_name = 'core/footer.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        footer_content = FooterContent.objects.first()
        if not footer_content:
            footer_content = FooterContent.objects.create()
        context['footer_content'] = footer_content
        return context


class FooterEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = FooterContent
    form_class = FooterContentForm
    template_name = 'osas/modals/footer_edit.html'
    success_url = reverse_lazy('dashboard')

    def test_func(self):
        return self.request.user.is_superuser or self.request.user.user_type == 1  # Only OSAS admins

    def get_object(self):
        return FooterContent.objects.first()

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)


# ----------------------------------- Admin Dashboard & Functionalities Sections ---------------------------------------
class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'osas/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Check if user is an organization
        if self.request.user.is_organization:
            context['is_organization_user'] = True
            context['current_organization'] = self.request.user.organization
            context['current_user_type'] = 15  # Organization user type
        else:
            context['is_organization_user'] = False
            context['current_user_type'] = self.request.user.user_type
            context['is_superuser'] = self.request.user.is_superuser

        from .models import Complaint
        context['complaint_status_choices'] = Complaint.STATUS_CHOICES
        context['respondent_type_choices'] = Complaint.RESPONDENT_TYPE_CHOICES

        self.add_organization_statistics(context)
        context['organizations'] = self.get_filtered_organizations()

        # Add basic context data
        self.add_basic_context(context)

        # Add paginated data
        self.add_paginated_data(context)

        # Add archived data
        self.add_archived_data(context)

        # Add announcements data
        self.add_announcements_data(context)

        # Add permissions data
        self.add_permissions_data(context)

        # User Statistics
        self.add_user_statistics(context)

        # Downloadable Statistics
        self.add_downloadable_statistics(context)

        # Admission Statistics
        self.add_admission_statistics(context)

        # Complaint Statistics
        self.add_complaint_statistics(context)

        # Announcement Statistics
        self.add_announcement_statistics(context)

        # NSTP Enlistment Statistics
        self.add_nstp_statistics(context)

        # NSTP Files Statistics
        self.add_nstp_files_statistics(context)

        # Scholarship Statistics
        self.add_scholarship_statistics(context)
        
        # Scholarship Application Statistics
        self.add_scholarship_application_statistics(context)
        
        # OJT Company Statistics
        self.add_ojt_statistics(context)

        # Organization Certificate
        self.add_certificates_data(context)

        # Organization AR
        self.add_accomplishment_data(context)
        return context

    def add_basic_context(self, context):
        context['COURSE_CHOICES'] = Announcement.COURSE_CHOICES
        context['user_type'] = self.request.user.get_user_type_display()
        context['form'] = CustomUserCreationForm()

        # All Scholarships for the filter dropdown
        context['all_scholarships'] = Scholarship.objects.filter(is_archived=False).order_by('name')

    def add_paginated_data(self, context):
        # NSTP Files
        context['nstp_files'] = self.get_paginated_nstp_files()

        # NSTP Enlistments
        context['nstp_enlistments'] = self.get_paginated_nstp_enlistments()

        # Admissions
        context['admissions'] = self.get_paginated_admissions()

        # Scholarship Applications
        context['applications'] = self.get_paginated_scholarship_applications()

        # Scholarship
        context['scholarships'] = self.get_filtered_scholarships()

        # Complaints
        context['complaints'] = self.get_paginated_complaints(status='under_review')
        context['resolved_complaints'] = self.get_paginated_complaints(status='resolved')

        # OJT Companies
        context['ojt_companies'] = self.get_paginated_ojt_companies()

        # Downloadables, Announcements, Scholarships
        context['downloadables'] = self.get_filtered_downloadables()
        context['announcements'] = self.get_filtered_announcements()
        context['scholarships'] = self.get_filtered_scholarships()

    def add_archived_data(self, context):
        if self.request.user.is_superuser:
            self.add_superuser_archived_data(context)
        else:
            self.add_regular_user_archived_data(context)

        # Admission Specific archived data
        self.add_admission_archived_data(context)

        # NSTP Specific archived data
        self.add_nstp_archived_data(context)

        # Complaint Specific archived data
        self.add_complaint_archived_data(context)

        # OJT Specific archived data
        self.add_ojt_archived_data(context)

        # Organization Specific archived ata
        self.add_organizations_archived_data(context)

        # Accomplishment Reports Specific archived data
        self.add_accomplishment_archived_data(context)

    def add_superuser_archived_data(self, context):
        context['archived_users'] = CustomUser.objects.filter(
            is_archived=True
        ).order_by('-archived_at').select_related('archived_by')

        context['archived_announcements'] = Announcement.objects.filter(
            is_archived=True
        ).order_by('-archived_at').select_related('archived_by', 'author')

        context['archived_downloadables'] = Downloadable.objects.filter(
            is_archived=True
        ).order_by('-archived_at').select_related('archived_by', 'created_by')

        context['archived_scholarships'] = Scholarship.objects.filter(
            is_archived=True
        ).order_by('-archived_at').select_related('archived_by', 'created_by', 'application_form')

        context['archived_scholarship_applications'] = ScholarshipApplication.objects.filter(
            is_archived=True
        ).order_by('-archived_at').select_related('student', 'scholarship', 'archived_by')

        context['archived_ojt_companies'] = OJTCompany.objects.filter(
            is_archived=True
        ).order_by('-archived_at').select_related('archived_by')

    def add_regular_user_archived_data(self, context):
        context['archived_users'] = CustomUser.objects.none()

        context['archived_announcements'] = Announcement.objects.filter(
            is_archived=True,
            author=self.request.user
        ).order_by('-archived_at').select_related('archived_by', 'author')

        context['archived_downloadables'] = Downloadable.objects.filter(
            is_archived=True,
            created_by=self.request.user
        ).order_by('-archived_at').select_related('archived_by', 'created_by')

        context['archived_scholarships'] = Scholarship.objects.filter(
            is_archived=True
        ).filter(
            Q(created_by=self.request.user) |
            Q(archived_by=self.request.user)
        ).order_by('-archived_at').select_related('archived_by', 'created_by', 'application_form')

        context['archived_scholarship_applications'] = ScholarshipApplication.objects.filter(
            is_archived=True
        ).filter(
            Q(student=self.request.user) |
            Q(archived_by=self.request.user) |
            Q(scholarship__created_by=self.request.user)
        ).order_by('-archived_at').select_related('student', 'scholarship', 'archived_by')

        context['archived_ojt_companies'] = OJTCompany.objects.filter(
            is_archived=True,
            archived_by=self.request.user
        ).order_by('-archived_at').select_related('archived_by')

    def add_admission_archived_data(self, context):
        if self.request.user.is_superuser or self.request.user.user_type in [1, 12]:
            context['archived_admissions'] = StudentAdmission.objects.filter(
                is_archived=True
            ).order_by('-archived_at').select_related('archived_by')
        else:
            context['archived_admissions'] = StudentAdmission.objects.filter(
                Q(is_archived=True) & (
                        Q(user=self.request.user) |
                        Q(archived_by=self.request.user)
                )
            ).order_by('-archived_at').select_related('archived_by')

    def add_nstp_archived_data(self, context):
        # NSTP Enlistment Data
        if self.request.user.is_superuser or self.request.user.user_type in [1, 2]:
            context['archived_nstp_students'] = NSTPStudentInfo.objects.filter(
                is_archived=True
            ).order_by('-archived_at').select_related('user', 'archived_by')
        else:
            context['archived_nstp_students'] = NSTPStudentInfo.objects.filter(
                Q(is_archived=True) & (
                        Q(user=self.request.user) |
                        Q(archived_by=self.request.user)
                )
            ).order_by('-archived_at').select_related('archived_by')

        # NSTP Files Data
        if self.request.user.is_superuser or self.request.user.user_type in [1, 2]:
            context['archived_nstp_files'] = NSTPFile.objects.filter(
                is_archived=True
            ).order_by('-archived_at').select_related('archived_by', 'created_by')
        else:
            context['archived_nstp_files'] = NSTPFile.objects.none()

    def add_complaint_archived_data(self, context):
        if self.request.user.is_superuser or self.request.user.user_type in [1, 11]:
            context['archived_complaints'] = Complaint.objects.filter(
                is_archived=True
            ).order_by('-archived_at').select_related('archived_by', 'created_by')
        else:
            context['archived_complaints'] = Complaint.objects.filter(
                is_archived=True,
                created_by=self.request.user
            ).order_by('-archived_at').select_related('archived_by', 'created_by')

    def add_ojt_archived_data(self, context):
        if self.request.user.is_superuser:
            context['archived_ojt_companies'] = OJTCompany.objects.filter(
                is_archived=True
            ).order_by('-archived_at').select_related('archived_by')
        else:
            context['archived_ojt_companies'] = OJTCompany.objects.filter(
                is_archived=True,
                archived_by=self.request.user
            ).order_by('-archived_at').select_related('archived_by')

    def add_organizations_archived_data(self, context):
        if self.request.user.is_superuser or self.request.user.user_type in [1, 10]:
            context['archived_organizations'] = Organization.objects.filter(
                is_archived=True
            ).order_by('-archived_at').select_related('archived_by')
        elif self.request.user.user_type == 15:
            if hasattr(self.request.user, 'organization') and self.request.user.organization:
                context['archived_organizations'] = Organization.objects.filter(
                    id=self.request.user.organization.id,
                    is_archived=True
                ).order_by('-archived_at').select_related('archived_by')
            else:
                context['archived_organizations'] = Organization.objects.none()
        else:
            context['archived_organizations'] = Organization.objects.none()

    def add_accomplishment_archived_data(self, context):
        if self.request.user.is_superuser or self.request.user.user_type in [1, 10]:
            # User type 1 or 10 can see all archived accomplishment reports
            context['archived_accomplishment_reports'] = AccomplishmentRecord.objects.filter(
                is_archived=True
            ).order_by('-archived_at').select_related('organization', 'submitted_by', 'archived_by')
        elif self.request.user.user_type == 15:
            # User type 15 can only see their owned data
            if hasattr(self.request.user, 'organization') and self.request.user.organization:
                context['archived_accomplishment_reports'] = AccomplishmentRecord.objects.filter(
                    organization=self.request.user.organization,
                    is_archived=True
                ).order_by('-archived_at').select_related('organization', 'submitted_by', 'archived_by')
            else:
                context['archived_accomplishment_reports'] = AccomplishmentRecord.objects.none()
        else:
            context['archived_accomplishment_reports'] = AccomplishmentRecord.objects.none()
            
    def add_announcements_data(self, context):
        current_time = timezone.now()
        twenty_four_hours_ago = current_time - timedelta(hours=24)

        context['current_time'] = current_time
        context['twenty_four_hours_ago'] = twenty_four_hours_ago
        context['time_delta'] = timedelta(hours=24)

        # Get published announcements with pagination for home section
        published_announcements = Announcement.objects.filter(
            is_published=True,
            is_archived=False
        ).order_by('-created_at').select_related('author')

        # Get paginated announcements for home section
        page = self.request.GET.get('home_page', 1)
        paginator = Paginator(published_announcements, 10)  # 10 announcements per page

        try:
            home_announcements = paginator.page(page)
        except PageNotAnInteger:
            home_announcements = paginator.page(1)
        except EmptyPage:
            home_announcements = paginator.page(paginator.num_pages)

        context['published_announcements'] = home_announcements
        context['has_next_home_page'] = home_announcements.has_next()
        context['next_home_page'] = home_announcements.next_page_number() if home_announcements.has_next() else None

        # Get ALL regular announcements for home section (for other uses)
        context['regular_announcements'] = published_announcements.exclude(
            category__in=['EMERGENCY', 'EVENT']
        )

        # Get ALL emergency announcements for home section
        context['emergency_announcements'] = published_announcements.filter(
            category='EMERGENCY',
            created_at__gte=twenty_four_hours_ago
        )

        # Get ALL upcoming events for home section
        context['upcoming_events'] = published_announcements.filter(
            category='EVENT',
            event_date__gt=current_time
        )

    def get_home_announcements_ajax(self, request):
        page = request.GET.get('page', 1)

        published_announcements = Announcement.objects.filter(
            is_published=True,
            is_archived=False
        ).order_by('-created_at').select_related('author')

        paginator = Paginator(published_announcements, 10)

        try:
            announcements_page = paginator.page(page)
        except PageNotAnInteger:
            announcements_page = paginator.page(1)
        except EmptyPage:
            return JsonResponse({
                'success': False,
                'message': 'No more announcements'
            })

        # Prepare announcement data for JSON response
        announcements_data = []
        for announcement in announcements_page:
            announcement_data = {
                'id': announcement.id,
                'title': announcement.title,
                'content': announcement.content,
                'category': announcement.category,
                'category_display': announcement.get_category_display(),
                'author_name': announcement.author.get_full_name(),
                'created_at': announcement.created_at.isoformat(),
                'timesince': timesince(announcement.created_at),
                'is_published': announcement.is_published,
                'link': announcement.link or '',
                'has_images': announcement.images.exists(),
                'images': []
            }

            # Add images if any
            if announcement.images.exists():
                for image in announcement.images.all()[:1]:
                    announcement_data['images'].append({
                        'url': image.image.url,
                        'caption': image.caption or ''
                    })

            # Add category-specific data
            if announcement.category == 'ENROLLMENT' and announcement.courses:
                announcement_data['courses'] = list(announcement.courses)
                announcement_data[
                    'enrollment_start'] = announcement.enrollment_start.isoformat() if announcement.enrollment_start else None
                announcement_data[
                    'enrollment_end'] = announcement.enrollment_end.isoformat() if announcement.enrollment_end else None

            elif announcement.category == 'EVENT':
                announcement_data[
                    'event_date'] = announcement.event_date.isoformat() if announcement.event_date else None
                announcement_data['location'] = announcement.location or ''

            elif announcement.category == 'SUSPENSION':
                announcement_data[
                    'suspension_date'] = announcement.suspension_date.isoformat() if announcement.suspension_date else None
                announcement_data[
                    'until_suspension_date'] = announcement.until_suspension_date.isoformat() if announcement.until_suspension_date else None

            elif announcement.category == 'EMERGENCY':
                announcement_data['contact_info'] = announcement.contact_info or ''

            elif announcement.category == 'SCHOLARSHIP':
                announcement_data[
                    'application_start'] = announcement.application_start.isoformat() if announcement.application_start else None
                announcement_data[
                    'application_end'] = announcement.application_end.isoformat() if announcement.application_end else None
                announcement_data['requirements'] = announcement.requirements or ''
                announcement_data['benefits'] = announcement.benefits or ''
                announcement_data[
                    'scholarship_name'] = announcement.scholarship.name if announcement.scholarship else ''

            announcements_data.append(announcement_data)

        return JsonResponse({
            'success': True,
            'announcements': announcements_data,
            'has_next': announcements_page.has_next(),
            'next_page': announcements_page.next_page_number() if announcements_page.has_next() else None
        })

    def add_permissions_data(self, context):
        if self.request.user.is_superuser:
            permissions = Permission.objects.all().select_related('content_type')
            grouped_permissions = {}
            for perm in permissions:
                app_model = f"{perm.content_type.app_label} | {perm.content_type.model}"
                if app_model not in grouped_permissions:
                    grouped_permissions[app_model] = []
                grouped_permissions[app_model].append(perm)
            context['grouped_permissions'] = grouped_permissions

    def get_paginated_data(self, queryset, page_param, per_page=10):
        page = self.request.GET.get(page_param, 1)
        paginator = Paginator(queryset, per_page)

        try:
            return paginator.page(page)
        except PageNotAnInteger:
            return paginator.page(1)
        except EmptyPage:
            return paginator.page(paginator.num_pages)

    def get_paginated_nstp_files(self):
        if self.request.user.is_superuser or self.request.user.user_type in [1, 2]:
            nstp_files_queryset = NSTPFile.objects.filter(is_archived=False).order_by('-created_at')
        else:
            nstp_files_queryset = NSTPFile.objects.none()

        return self.get_paginated_data(nstp_files_queryset.order_by('-created_at'), 'nstp_page')

    def get_paginated_nstp_enlistments(self):
        if self.request.user.is_superuser or self.request.user.user_type in [1, 2]:
            nstp_queryset = NSTPStudentInfo.objects.filter(
                is_archived=False
            ).select_related('user').order_by('-created_at')
        elif self.request.user.user_type == 14:  # Student
            nstp_queryset = NSTPStudentInfo.objects.filter(
                user=self.request.user,
                is_archived=False
            ).select_related('user').order_by('-created_at')
        else:
            nstp_queryset = NSTPStudentInfo.objects.none()

        return self.get_paginated_data(nstp_queryset, 'nstp_page')

    def get_paginated_admissions(self):
        if self.request.user.is_superuser or self.request.user.user_type in [1, 12]:
            admissions_queryset = StudentAdmission.objects.filter(
                is_archived=False
            ).select_related('user').order_by('-created_at')
        elif self.request.user.user_type == 14:  # Student
            admissions_queryset = StudentAdmission.objects.filter(
                user=self.request.user,
                is_archived=False
            ).select_related('user').order_by('-created_at')
        else:
            admissions_queryset = StudentAdmission.objects.none()

        return self.get_paginated_data(admissions_queryset, 'admission_page')

    def get_paginated_scholarship_applications(self):
        if self.request.user.is_superuser or self.request.user.user_type in [1, 5]:
            applications_queryset = ScholarshipApplication.objects.filter(
                is_archived=False
            ).select_related('student', 'scholarship').order_by('-application_date')
        elif self.request.user.user_type == 14:  # Student
            applications_queryset = ScholarshipApplication.objects.filter(
                student=self.request.user,
                is_archived=False
            ).select_related('scholarship').order_by('-application_date')
        else:
            applications_queryset = ScholarshipApplication.objects.none()

        return self.get_paginated_data(applications_queryset, 'scholarship_app_page')

    def get_paginated_complaints(self, status='under_review'):
        if status == 'resolved':
            # For resolved complaints
            if self.request.user.is_superuser or self.request.user.user_type in [1, 11]:
                complaints_queryset = Complaint.objects.filter(
                    status='resolved',
                    is_archived=False
                ).order_by('-updated_at')
            else:
                complaints_queryset = Complaint.objects.filter(
                    status='resolved',
                    is_archived=False,
                    created_by=self.request.user
                ).order_by('-updated_at')
            page_param = 'resolved_page'
        else:
            # For active complaints (all non-resolved statuses)
            # Updated statuses to match new model
            active_statuses = ['under_review', '1st_hearing', '2nd_hearing', 'other_hearing', 'ongoing_hearing',
                               'canceled']
            if self.request.user.is_superuser or self.request.user.user_type in [1, 11]:
                complaints_queryset = Complaint.objects.filter(
                    status__in=active_statuses,
                    is_archived=False
                ).order_by('-created_at')
            else:
                complaints_queryset = Complaint.objects.filter(
                    status__in=active_statuses,
                    is_archived=False,
                    created_by=self.request.user
                ).order_by('-created_at')
            page_param = 'under_review_page'

        return self.get_paginated_data(complaints_queryset, page_param)

    def get_filtered_announcements(self):
        queryset = Announcement.objects.filter(is_archived=False).select_related('author')
        page = self.request.GET.get('announcement_page', 1)

        if self.request.user.user_type == 14:  # Student
            return None
        elif self.request.user.user_type == 1:  # Admin/superuser
            paginator = Paginator(queryset.order_by('-created_at'), 10)
        else:
            queryset = queryset.filter(author=self.request.user)
            paginator = Paginator(queryset.order_by('-created_at'), 10)

        try:
            return paginator.page(page)
        except PageNotAnInteger:
            return paginator.page(1)
        except EmptyPage:
            return paginator.page(paginator.num_pages)

    def get_filtered_downloadables(self):
        queryset = Downloadable.objects.filter(is_archived=False).select_related('created_by')
        request = self.request

        # Apply filters based on request parameters
        search_term = request.GET.get('search', '').lower()
        category_filter = request.GET.get('category', 'all')
        status_filter = request.GET.get('status', 'all')
        date_filter = request.GET.get('date', 'all')

        # Apply search filter
        if search_term:
            queryset = queryset.filter(
                Q(title__icontains=search_term) |
                Q(description__icontains=search_term)
            )

        # Apply category filter
        if category_filter != 'all':
            queryset = queryset.filter(category=category_filter)

        # Apply status filter
        if status_filter != 'all':
            queryset = queryset.filter(is_active=(status_filter == 'active'))

        # Apply date filters
        if date_filter != 'all':
            now = timezone.now()
            if date_filter == 'week':
                one_week_ago = now - timedelta(days=7)
                queryset = queryset.filter(created_at__gte=one_week_ago)
            elif date_filter == 'month':
                queryset = queryset.filter(created_at__month=now.month, created_at__year=now.year)
            elif date_filter == 'this_year':
                queryset = queryset.filter(created_at__year=now.year)
            elif date_filter == 'last_year':
                queryset = queryset.filter(created_at__year=now.year - 1)
            elif date_filter.isdigit():
                queryset = queryset.filter(created_at__year=int(date_filter))

        if request.GET.get('get_all_downloadables') == '1':
            return queryset.order_by('-created_at')

        # For regular requests, apply pagination
        paginator = Paginator(queryset.order_by('-created_at'), 9)
        page = request.GET.get('template_page', 1)

        try:
            return paginator.page(page)
        except PageNotAnInteger:
            return paginator.page(1)
        except EmptyPage:
            return paginator.page(paginator.num_pages)

    def add_downloadable_statistics(self, context):
        downloadables = Downloadable.objects.filter(is_archived=False)

        # Count by category
        context['society_forms_count'] = downloadables.filter(category='society_forms').count()
        context['osas_forms_count'] = downloadables.filter(category='osas_forms').count()
        context['ojt_forms_count'] = downloadables.filter(category='ojt_forms').count()
        context['scholarship_forms_count'] = downloadables.filter(category='scholarship_forms').count()
        context['guidelines_count'] = downloadables.filter(category='guidelines').count()
        context['manuals_count'] = downloadables.filter(category='manuals').count()
        context['others_count'] = downloadables.filter(category='others').count()

        # Total count
        context['total_downloadables'] = downloadables.count()

    def get_filtered_scholarships(self):
        queryset = Scholarship.objects.filter(is_archived=False)
        page = self.request.GET.get('scholarship_page', 1)

        print(f"DEBUG: Initial scholarships load - User type: {self.request.user.user_type}")

        if self.request.user.is_superuser or self.request.user.user_type == 1:
            paginator = Paginator(queryset.order_by('-created_at'), 10)
            print(f"DEBUG: Admin user - showing all non-archived scholarships")
        elif self.request.user.user_type == 14:
            today = timezone.now().date()
            queryset = queryset.filter(is_active=True)  # Students only see active scholarships
            paginator = Paginator(queryset.order_by('-created_at'), 10)
            print(f"DEBUG: Student user - showing active non-archived scholarships")
        else:
            queryset = queryset.filter(
                Q(created_by=self.request.user) | Q(is_active=True))
            paginator = Paginator(queryset.order_by('-created_at'), 10)
            print(f"DEBUG: Other user - showing created_by or active non-archived scholarships")

        print(f"DEBUG: Final queryset count: {queryset.count()}")

        try:
            return paginator.page(page)
        except PageNotAnInteger:
            return paginator.page(1)
        except EmptyPage:
            return paginator.page(paginator.num_pages)

    def get(self, request, *args, **kwargs):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            if 'get_filtered_complaints' in request.GET:
                return self.get_filtered_complaints_ajax(request)
            if 'get_filtered_users' in request.GET:
                return self.get_filtered_users(request)
            if 'get_filtered_admissions' in request.GET:
                return self.get_filtered_admissions(request)
            if 'get_filtered_announcements' in request.GET:
                return self.get_filtered_announcements_ajax(request)
            if 'get_filtered_nstp' in request.GET:
                return self.get_filtered_nstp_enlistments(request)
            if 'get_filtered_nstp_files' in request.GET:
                return self.get_filtered_nstp_files(request)
            if 'get_filtered_scholarships' in request.GET:
                return self.get_filtered_scholarships_ajax(request)
            if 'get_filtered_applications' in request.GET:
                return self.get_filtered_applications(request)
            if 'get_filtered_ojt_companies' in request.GET:
                return self.get_filtered_ojt_companies(request)
            if 'get_home_announcements' in request.GET:
                return self.get_home_announcements_ajax(request)
            if 'get_filtered_organizations' in request.GET:
                return self.get_filtered_organizations_ajax(request)
            if 'get_filtered_certificates' in request.GET:
                return self.get_filtered_certificates_ajax(request)
            if 'get_filtered_accomplishment_reports' in request.GET:
                return self.get_filtered_accomplishment_reports_ajax(request)
        context = self.get_context_data(**kwargs)
        return self.render_to_response(context)

    def get_paginated_data(self, queryset, page_param, per_page=10):
        page = self.request.GET.get(page_param, 1)
        paginator = Paginator(queryset, per_page)

        try:
            page = int(page)
        except (ValueError, TypeError):
            page = 1

        try:
            return paginator.page(page)
        except PageNotAnInteger:
            return paginator.page(1)
        except EmptyPage:
            return paginator.page(paginator.num_pages)

    def get_filtered_users(self, request):
        search_term = request.GET.get('search', '').lower()
        unit_filter = request.GET.get('unit', '')
        verified_filter = request.GET.get('verified', '')
        status_filter = request.GET.get('status', '')
        sort_column = request.GET.get('sort', 'date_joined')
        sort_direction = request.GET.get('direction', 'desc')
        page_number = request.GET.get('user_page', 1)
        per_page = 10

        # Start with base queryset
        users = CustomUser.objects.filter(is_archived=False)

        # Apply filters
        if search_term:
            users = users.filter(
                Q(first_name__icontains=search_term) |
                Q(last_name__icontains=search_term) |
                Q(username__icontains=search_term) |
                Q(id__icontains=search_term) |
                Q(organization_account__organization_name__icontains=search_term)
            )

        if unit_filter:
            users = users.filter(user_type=unit_filter)

        if verified_filter:
            verified = verified_filter == 'Verified'
            users = users.filter(is_verified=verified)

        if status_filter:
            active = status_filter == 'Active'
            users = users.filter(is_active=active)

        sort_mapping = {
            'id': 'id',
            'name': 'first_name',
            'username': 'username',
            'user_type': 'user_type',
            'is_verified': 'is_verified',
            'is_active': 'is_active',
            'date_joined': 'date_joined',
            'created_at': 'date_joined'
        }

        if sort_column not in sort_mapping:
            sort_column = 'date_joined'
            sort_direction = 'desc'

        sort_field = sort_mapping[sort_column]
        if sort_direction == 'desc':
            sort_field = f'-{sort_field}'

        if sort_column != 'id':
            users = users.order_by(sort_field, '-id')
        else:
            users = users.order_by(sort_field)

        page_obj = self.get_paginated_data(users, 'user_page', per_page)

        # Prepare data for JSON response
        user_data = []
        for user in page_obj.object_list:
            # Get the appropriate name based on user type
            if user.is_organization and hasattr(user, 'organization_account'):
                # For organizations, use organization name
                display_name = user.organization_account.organization_name
                first_name = display_name
                last_name = ""
            else:
                # For regular users, use first and last name
                first_name = user.first_name or ''
                last_name = user.last_name or ''
                display_name = f"{first_name} {last_name}".strip() or user.username

            user_data.append({
                'id': user.id,
                'first_name': first_name,
                'last_name': last_name,
                'display_name': display_name,
                'username': user.username,
                'user_type': user.get_user_type_display(),
                'is_verified': user.is_verified,
                'is_active': user.is_active,
                'date_joined': user.date_joined.isoformat() if hasattr(user,
                                                                       'date_joined') else user.created_at.isoformat(),
                'is_organization': user.is_organization,
                'can_edit': request.user.has_perm('osas.change_customuser'),
                'can_delete': request.user.has_perm('osas.delete_customuser'),
                'can_view': request.user.has_perm('osas.view_customuser'),
                'can_approve': not user.is_verified and request.user.has_perm('osas.change_customuser')
            })

        return JsonResponse({
            'users': user_data,
            'pagination': {
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'current_page': page_obj.number,
                'num_pages': page_obj.paginator.num_pages,
                'count': page_obj.paginator.count,
                'start_index': page_obj.start_index(),
                'end_index': page_obj.end_index(),
            }
        })

    def add_user_statistics(self, context):
        users = CustomUser.objects.filter(is_archived=False)

        context['total_users'] = users.count()
        context['verified_users'] = users.filter(is_verified=True).count()
        context['unverified_users'] = users.filter(is_verified=False).count()
        context['active_users'] = users.filter(is_active=True).count()
        context['inactive_users'] = users.filter(is_active=False).count()

    def get_filtered_admissions(self, request):
        # Get all filter/sort parameters from the request
        search_term = request.GET.get('search', '').lower()
        type_filter = request.GET.get('type', '')
        sort_column = request.GET.get('sort', 'created_at')
        sort_direction = request.GET.get('direction', 'desc')
        page_number = request.GET.get('admission_page', 1)
        per_page = 10

        # Start with base queryset
        if request.user.is_superuser or request.user.user_type in [1, 12]:
            admissions = StudentAdmission.objects.filter(is_archived=False)
        elif request.user.user_type == 14:  # Student
            admissions = StudentAdmission.objects.filter(
                user=request.user,
                is_archived=False
            )
        else:
            admissions = StudentAdmission.objects.none()

        # Apply filters
        if search_term:
            admissions = admissions.filter(
                Q(control_no__icontains=search_term) |
                Q(user__first_name__icontains=search_term) |
                Q(user__last_name__icontains=search_term) |
                Q(course__name__icontains=search_term) |
                Q(remarks__icontains=search_term)
            )

        if type_filter:
            admissions = admissions.filter(student_type=type_filter)

        # Apply sorting
        if sort_column:
            if sort_direction == 'desc':
                sort_column = f'-{sort_column}'
            admissions = admissions.order_by(sort_column)

        # Create paginator
        paginator = Paginator(admissions.select_related('user', 'course'), per_page)

        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        # Prepare data for JSON response
        admission_data = []
        for admission in page_obj.object_list:
            admission_data.append({
                'id': admission.id,
                'control_no': admission.control_no,
                'student_name': f"{admission.user.last_name}, {admission.user.first_name}" if admission.user else '(No user associated)',
                'student_type': admission.student_type,
                'student_type_display': admission.get_student_type_display(),
                'course': admission.course.name if admission.course else '',
                'status': admission.status,
                'status_display': admission.get_status_display(),
                'remarks': admission.remarks,
                'created_at': admission.created_at.isoformat(),
                'can_approve': request.user.is_superuser or request.user.user_type in [1, 12],
                'can_edit': not request.user.is_student or admission.status != 'done',
                'can_archive': not request.user.is_student or admission.status != 'done'
            })

        return JsonResponse({
            'admissions': admission_data,
            'pagination': {
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'current_page': page_obj.number,
                'num_pages': paginator.num_pages,
                'total_count': paginator.count,
                'start_index': page_obj.start_index(),
                'end_index': page_obj.end_index(),
            }
        })

    def add_admission_statistics(self, context):
        admissions = StudentAdmission.objects.filter(is_archived=False)

        # Count by student type
        context['current_grade12_count'] = admissions.filter(student_type='current_grade12').count()
        context['shs_graduate_count'] = admissions.filter(student_type='shs_graduate').count()
        context['transferee_count'] = admissions.filter(student_type='transferee').count()

        # Count by requirement status
        context['complete_requirements_count'] = admissions.filter(status='done').count()
        context['incomplete_requirements_count'] = admissions.filter(status='incomplete').count()

        # Total count
        context['total_admissions'] = admissions.count()

        # User-specific statistics (for non-superusers)
        if not (self.request.user.is_superuser or self.request.user.user_type == 1 or self.request.user.user_type == 12):
            user_admissions = admissions.filter(user=self.request.user)

            context['user_current_grade12_count'] = user_admissions.filter(student_type='current_grade12').count()
            context['user_shs_graduate_count'] = user_admissions.filter(student_type='shs_graduate').count()
            context['user_transferee_count'] = user_admissions.filter(student_type='transferee').count()
            context['user_complete_requirements_count'] = user_admissions.filter(status='done').count()
            context['user_incomplete_requirements_count'] = user_admissions.filter(status='incomplete').count()
            context['user_total_admissions'] = user_admissions.count()

    def get_filtered_announcements_ajax(self, request):
        search_term = request.GET.get('search', '').lower()
        category_filter = request.GET.get('category', 'all')
        status_filter = request.GET.get('status', 'all')
        date_filter = request.GET.get('date', 'all')
        sort_column = request.GET.get('sort', 'created_at')
        sort_direction = request.GET.get('direction', 'desc')
        page_number = request.GET.get('announcement_page', 1)

        per_page = int(request.GET.get('per_page', 10))

        # Start with base queryset
        if request.user.user_type == 14:  # Student
            announcements = Announcement.objects.none()
        elif request.user.user_type == 1:  # Admin/superuser
            announcements = Announcement.objects.filter(is_archived=False)
        else:
            announcements = Announcement.objects.filter(
                is_archived=False,
                author=request.user
            )

        # Apply search filter
        if search_term:
            announcements = announcements.filter(
                Q(title__icontains=search_term) |
                Q(content__icontains=search_term) |
                Q(author__first_name__icontains=search_term) |
                Q(author__last_name__icontains=search_term)
            )

        # Apply category filter
        if category_filter != 'all':
            category_mapping = {
                'Basic Announcements': 'BASIC',
                'Enrollment Announcements': 'ENROLLMENT',
                'Event Announcements': 'EVENT',
                'Class Suspension Announcements': 'CLASS_SUSPENSION',
                'Emergency Announcements': 'EMERGENCY'
            }

            if category_filter in category_mapping:
                announcements = announcements.filter(category=category_mapping[category_filter])

        # Apply status filter
        if status_filter != 'all':
            is_published = status_filter == 'published'
            announcements = announcements.filter(is_published=is_published)

        # Apply date filters
        if date_filter != 'all':
            now = timezone.now()
            if date_filter == 'week':
                one_week_ago = now - timedelta(days=7)
                announcements = announcements.filter(created_at__gte=one_week_ago)
            elif date_filter == 'month':
                announcements = announcements.filter(created_at__month=now.month, created_at__year=now.year)
            elif date_filter == 'this_year':
                announcements = announcements.filter(created_at__year=now.year)
            elif date_filter == 'last_year':
                announcements = announcements.filter(created_at__year=now.year - 1)
            elif date_filter.isdigit():
                announcements = announcements.filter(created_at__year=int(date_filter))

        # Apply sorting
        sort_mapping = {
            'id': 'id',
            'title': 'title',
            'category': 'category',
            'author': 'author__first_name',
            'created_at': 'created_at',
            'is_published': 'is_published'
        }

        if sort_column in sort_mapping:
            sort_field = sort_mapping[sort_column]
            if sort_direction == 'desc':
                sort_field = f'-{sort_field}'
            announcements = announcements.order_by(sort_field)

        # Create paginator
        paginator = Paginator(announcements.select_related('author'), per_page)

        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        # Prepare data for JSON response
        announcement_data = []
        for announcement in page_obj.object_list:
            announcement_data.append({
                'id': announcement.id,
                'title': announcement.title,
                'category': announcement.get_category_display(),
                'category_value': announcement.category,
                'author_name': announcement.author.get_full_name(),
                'created_at': announcement.created_at.isoformat(),
                'is_published': announcement.is_published,
                'can_view': request.user.has_perm('osas.view_announcement'),
                'can_edit': (request.user.has_perm('osas.change_announcement') and
                             (request.user == announcement.author or request.user.is_superuser)),
                'can_delete': (request.user.has_perm('osas.delete_announcement') and
                               (request.user == announcement.author or request.user.is_superuser))
            })

        return JsonResponse({
            'announcements': announcement_data,
            'pagination': {
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'current_page': page_obj.number,
                'num_pages': paginator.num_pages,
                'total_count': paginator.count,
                'start_index': page_obj.start_index(),
                'end_index': page_obj.end_index(),
            }
        })

    def add_announcement_statistics(self, context):
        if self.request.user.is_superuser:
            # Superusers can see all announcements
            announcements = Announcement.objects.filter(is_archived=False)
        else:
            # Regular users can only see their own announcements
            announcements = Announcement.objects.filter(
                author=self.request.user,
                is_archived=False
            )

        # Total count
        context['total_announcements'] = announcements.count()

        # Status counts
        context['published_count'] = announcements.filter(is_published=True).count()
        context['draft_count'] = announcements.filter(is_published=False).count()

        # Category counts
        context['basic_count'] = announcements.filter(category='BASIC').count()
        context['enrollment_count'] = announcements.filter(category='ENROLLMENT').count()
        context['event_count'] = announcements.filter(category='EVENT').count()
        context['class_suspension_count'] = announcements.filter(category='CLASS_SUSPENSION').count()
        context['emergency_count'] = announcements.filter(category='EMERGENCY').count()

    def get_filtered_complaints_ajax(self, request):
        status = request.GET.get('status', 'under_review')
        search_term = request.GET.get('search', '').lower()
        date_filter = request.GET.get('date', 'all')
        sort_column = request.GET.get('sort', 'created_at')
        sort_direction = request.GET.get('direction', 'desc')
        page_number = request.GET.get(f'{status}_page', 1)
        per_page = int(request.GET.get('per_page', 10))

        print(f"Filtering complaints - Status: {status}, Search: {search_term}, Date Filter: {date_filter}")

        # Start with base queryset
        if status == 'resolved':
            # For resolved complaints
            if request.user.is_superuser or request.user.user_type in [1, 11]:
                complaints = Complaint.objects.filter(
                    status='resolved',
                    is_archived=False
                )
            else:
                complaints = Complaint.objects.filter(
                    status='resolved',
                    is_archived=False,
                    created_by=request.user
                )
            page_param = 'resolved_page'
        else:
            # For active complaints (all non-resolved statuses)
            active_statuses = ['under_review', '1st_hearing', '2nd_hearing', 'other_hearing', 'ongoing_hearing',
                               'canceled']
            if request.user.is_superuser or request.user.user_type in [1, 11]:
                complaints = Complaint.objects.filter(
                    status__in=active_statuses,
                    is_archived=False
                )
            else:
                complaints = Complaint.objects.filter(
                    status__in=active_statuses,
                    is_archived=False,
                    created_by=request.user
                )
            page_param = 'under_review_page'

        # Apply search filter
        if search_term:
            complaints = complaints.filter(
                Q(reference_number__icontains=search_term) |
                Q(title__icontains=search_term) |
                Q(complainant_first_name__icontains=search_term) |
                Q(complainant_last_name__icontains=search_term)
            )

        # Apply date filters
        if date_filter != 'all':
            now = timezone.now()
            today = now.date()

            if date_filter == 'incident_day':
                complaints = complaints.filter(incident_date=today)
            elif date_filter == 'resolved_day':
                complaints = complaints.filter(updated_at__date=today)
            elif date_filter == 'incident_week' or date_filter == 'resolved_week':
                one_week_ago = now - timedelta(days=7)
                if date_filter == 'incident_week':
                    complaints = complaints.filter(incident_date__gte=one_week_ago.date())
                else:
                    complaints = complaints.filter(updated_at__gte=one_week_ago)
            elif date_filter == 'incident_month' or date_filter == 'resolved_month':
                if date_filter == 'incident_month':
                    complaints = complaints.filter(incident_date__month=now.month, incident_date__year=now.year)
                else:
                    complaints = complaints.filter(updated_at__month=now.month, updated_at__year=now.year)
            elif date_filter == 'this_year':
                if status == 'resolved':
                    complaints = complaints.filter(updated_at__year=now.year)
                else:
                    complaints = complaints.filter(incident_date__year=now.year)
            elif date_filter == 'last_year':
                if status == 'resolved':
                    complaints = complaints.filter(updated_at__year=now.year - 1)
                else:
                    complaints = complaints.filter(incident_date__year=now.year - 1)
            elif date_filter.isdigit():
                year = int(date_filter)
                if status == 'resolved':
                    complaints = complaints.filter(updated_at__year=year)
                else:
                    complaints = complaints.filter(incident_date__year=year)
            elif date_filter == 'oldest_resolved' or date_filter == 'incident_oldest':
                sort_column = 'updated_at' if status == 'resolved' else 'incident_date'
                sort_direction = 'asc'
            elif date_filter == 'newest_resolved' or date_filter == 'incident_newest':
                sort_column = 'updated_at' if status == 'resolved' else 'incident_date'
                sort_direction = 'desc'

        # Apply sorting
        sort_mapping = {
            'reference_number': 'reference_number',
            'title': 'title',
            'complainant': 'complainant_first_name',
            'incident_date': 'incident_date',
            'created_at': 'created_at',
            'updated_at': 'updated_at',
            'next_hearing_date': 'next_hearing_date'
        }

        if sort_column in sort_mapping:
            sort_field = sort_mapping[sort_column]
            if sort_direction == 'desc':
                sort_field = f'-{sort_field}'
            complaints = complaints.order_by(sort_field)
        else:
            # Default sorting
            if status == 'resolved':
                complaints = complaints.order_by('-updated_at')
            else:
                complaints = complaints.order_by('-created_at')

        # Create paginator
        paginator = Paginator(complaints, per_page)

        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        # Prepare data for JSON response
        complaint_data = []
        for complaint in page_obj.object_list:
            # Get the latest hearing for this complaint
            latest_hearing = complaint.hearings.order_by('-schedule_date', '-schedule_time').first()

            # Determine status display - Now using the complaint's status directly
            status_display = complaint.get_status_display()

            # Get next hearing info
            next_hearing = complaint.next_hearing
            next_hearing_display = None
            if next_hearing:
                next_hearing_display = f"{next_hearing.schedule_date.strftime('%b %d, %Y')} {next_hearing.schedule_time.strftime('%I:%M %p')}"

            # Determine permission flags
            is_owner = complaint.created_by == request.user
            is_admin_user = request.user.is_superuser or request.user.user_type in [1, 11]

            # View permission: everyone can view their own complaints or admin can view all
            can_view = is_owner or is_admin_user

            # Edit permission: owners can edit their complaints, admins can edit all
            can_edit = is_owner or is_admin_user

            # Delete/Archive permission: owners can archive their complaints, admins can archive all
            can_delete = is_owner or is_admin_user

            # Resolve permission: only admins can resolve complaints (for user_type 14, this should be false)
            can_resolve = is_admin_user and status != 'resolved'

            complaint_data.append({
                'id': complaint.id,
                'reference_number': complaint.reference_number,
                'title': complaint.title,
                'complainant_first_name': complaint.complainant_first_name,
                'complainant_last_name': complaint.complainant_last_name,
                'incident_date': complaint.incident_date.isoformat() if complaint.incident_date else None,
                'created_at': complaint.created_at.isoformat(),
                'updated_at': complaint.updated_at.isoformat(),
                'status': complaint.status,
                'status_display': status_display,
                'next_hearing_display': next_hearing_display,
                'hearing_count': complaint.hearing_count if hasattr(complaint, 'hearing_count') else 0,
                'can_view': can_view,
                'can_edit': can_edit,
                'can_delete': can_delete,
                'can_resolve': can_resolve
            })

        return JsonResponse({
            'complaints': complaint_data,
            'pagination': {
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'current_page': page_obj.number,
                'num_pages': paginator.num_pages,
                'total_count': paginator.count,
                'start_index': page_obj.start_index(),
                'end_index': page_obj.end_index(),
            }
        })

    def add_complaint_statistics(self, context):
        if self.request.user.is_superuser or self.request.user.user_type in [1, 11]:
            # Superusers and complaint managers can see all complaints
            total_complaints = Complaint.objects.filter(is_archived=False).count()

            # Count all non-resolved statuses as "Active"
            active_statuses = ['under_review', '1st_hearing', '2nd_hearing', 'other_hearing', 'ongoing_hearing',
                               'canceled']
            under_review_count = Complaint.objects.filter(
                status__in=active_statuses,
                is_archived=False
            ).count()

            resolved_count = Complaint.objects.filter(
                status='resolved',
                is_archived=False
            ).count()

            # Count 1st Hearing complaints - Now based on status, not hearing_type
            first_hearing_count = Complaint.objects.filter(
                status='1st_hearing',
                is_archived=False
            ).count()

            # Count 2nd Hearing complaints - Now based on status, not hearing_type
            second_hearing_count = Complaint.objects.filter(
                status='2nd_hearing',
                is_archived=False
            ).count()

            # Count Other Hearing complaints - Now based on status, not hearing_type
            other_hearing_count = Complaint.objects.filter(
                status='other_hearing',
                is_archived=False
            ).count()

            # Count ongoing hearings
            ongoing_hearing_count = Complaint.objects.filter(
                status='ongoing_hearing',
                is_archived=False
            ).count()

            # Count canceled complaints
            canceled_count = Complaint.objects.filter(
                status='canceled',
                is_archived=False
            ).count()

            # Count under review (not hearing scheduled)
            pure_under_review_count = Complaint.objects.filter(
                status='under_review',
                is_archived=False
            ).count()

        else:
            # Regular users can only see their own complaints
            total_complaints = Complaint.objects.filter(
                created_by=self.request.user,
                is_archived=False
            ).count()

            active_statuses = ['under_review', '1st_hearing', '2nd_hearing', 'other_hearing', 'ongoing_hearing',
                               'canceled']
            under_review_count = Complaint.objects.filter(
                created_by=self.request.user,
                status__in=active_statuses,
                is_archived=False
            ).count()

            resolved_count = Complaint.objects.filter(
                created_by=self.request.user,
                status='resolved',
                is_archived=False
            ).count()

            # Count 1st Hearing complaints for regular users - Now based on status
            first_hearing_count = Complaint.objects.filter(
                created_by=self.request.user,
                status='1st_hearing',
                is_archived=False
            ).count()

            # Count 2nd Hearing complaints for regular users - Now based on status
            second_hearing_count = Complaint.objects.filter(
                created_by=self.request.user,
                status='2nd_hearing',
                is_archived=False
            ).count()

            # Count Other Hearing complaints for regular users - Now based on status
            other_hearing_count = Complaint.objects.filter(
                created_by=self.request.user,
                status='other_hearing',
                is_archived=False
            ).count()

            # Count ongoing hearings for regular users
            ongoing_hearing_count = Complaint.objects.filter(
                created_by=self.request.user,
                status='ongoing_hearing',
                is_archived=False
            ).count()

            # Count canceled complaints for regular users
            canceled_count = Complaint.objects.filter(
                created_by=self.request.user,
                status='canceled',
                is_archived=False
            ).count()

            # Count under review (not hearing scheduled) for regular users
            pure_under_review_count = Complaint.objects.filter(
                created_by=self.request.user,
                status='under_review',
                is_archived=False
            ).count()

        # Add all counts to context
        context['total_complaints'] = total_complaints
        context['under_review_count'] = under_review_count
        context['resolved_count'] = resolved_count
        context['first_hearing_count'] = first_hearing_count
        context['second_hearing_count'] = second_hearing_count
        context['other_hearing_count'] = other_hearing_count
        context['ongoing_hearing_count'] = ongoing_hearing_count
        context['canceled_count'] = canceled_count
        context['pure_under_review_count'] = pure_under_review_count

    def get_filtered_nstp_enlistments(self, request):
        search_term = request.GET.get('search', '').lower()
        semester_filter = request.GET.get('semester', '')
        sort_column = request.GET.get('sort', 'created_at')
        sort_direction = request.GET.get('direction', 'desc')
        page_number = request.GET.get('page', 1)
        per_page = 10

        # Start with base queryset
        if request.user.is_superuser or request.user.user_type in [1, 2]:
            nstp_queryset = NSTPStudentInfo.objects.filter(
                is_archived=False
            ).select_related('user').order_by('-created_at')
        elif request.user.user_type == 14:  # Student
            nstp_queryset = NSTPStudentInfo.objects.filter(
                user=request.user,
                is_archived=False
            ).select_related('user').order_by('-created_at')
        else:
            nstp_queryset = NSTPStudentInfo.objects.none()

        # Apply search filter
        if search_term:
            nstp_queryset = nstp_queryset.filter(
                Q(student_number__icontains=search_term) |
                Q(first_name__icontains=search_term) |
                Q(last_name__icontains=search_term) |
                Q(program__icontains=search_term)
            )

        # Apply semester filter
        if semester_filter:
            nstp_queryset = nstp_queryset.filter(semester=semester_filter)

        # Apply sorting
        sort_mapping = {
            'student_number': 'student_number',
            'last_name': 'last_name',
            'program': 'program',
            'semester': 'semester',
            'approval_status': 'approval_status',
            'created_at': 'created_at'
        }

        if sort_column in sort_mapping:
            sort_field = sort_mapping[sort_column]
            if sort_direction == 'desc':
                sort_field = f'-{sort_field}'
            nstp_queryset = nstp_queryset.order_by(sort_field)

        # Create paginator
        paginator = Paginator(nstp_queryset, per_page)

        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        # Prepare data for JSON response
        nstp_data = []
        for enlistment in page_obj.object_list:
            nstp_data.append({
                'id': enlistment.id,
                'student_number': enlistment.student_number,
                'first_name': enlistment.first_name,
                'last_name': enlistment.last_name,
                'program': enlistment.program,
                'semester': enlistment.semester,
                'approval_status': enlistment.approval_status,
                'approval_status_display': enlistment.get_approval_status_display(),
                'created_at': enlistment.created_at.strftime('%Y-%m-%d'),
                'can_approve': request.user.is_superuser or request.user.user_type in [1, 2],
                'can_edit': (enlistment.approval_status == 'pending' or
                             request.user.is_superuser or
                             request.user.user_type in [1, 2]),
                'can_archive': (enlistment.approval_status == 'pending' or
                                request.user.is_superuser or
                                request.user.user_type in [1, 2])
            })

        return JsonResponse({
            'nstp_enlistments': nstp_data,
            'pagination': {
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'current_page': page_obj.number,
                'num_pages': paginator.num_pages,
                'total_count': paginator.count,
                'start_index': page_obj.start_index(),
                'end_index': page_obj.end_index(),
            }
        })

    def add_nstp_statistics(self, context):
        if self.request.user.is_superuser or self.request.user.user_type in [1, 2]:
            # Admins can see all NSTP enlistments
            nstp_enlistments = NSTPStudentInfo.objects.filter(is_archived=False)
        elif self.request.user.user_type == 14:  # Student
            # Students can only see their own enlistments
            nstp_enlistments = NSTPStudentInfo.objects.filter(
                user=self.request.user,
                is_archived=False
            )
        else:
            nstp_enlistments = NSTPStudentInfo.objects.none()

        # Total count
        context['total_nstp_enlistments'] = nstp_enlistments.count()

        # Semester counts
        context['first_sem_count'] = nstp_enlistments.filter(semester='1st Sem').count()
        context['second_sem_count'] = nstp_enlistments.filter(semester='2nd Sem').count()

        # Status counts
        context['pending_nstp_count'] = nstp_enlistments.filter(approval_status='pending').count()
        context['approved_nstp_count'] = nstp_enlistments.filter(approval_status='approved').count()
        context['rejected_nstp_count'] = nstp_enlistments.filter(approval_status='rejected').count()

    def get_filtered_nstp_files(self, request):
        search_term = request.GET.get('search', '').lower()
        category_filter = request.GET.get('category', 'all')
        semester_filter = request.GET.get('semester', 'all')
        sort_column = request.GET.get('sort', 'created_at')
        sort_direction = request.GET.get('direction', 'desc')
        page_number = request.GET.get('nstp_page', 1)
        per_page = 9

        # Start with base queryset
        if request.user.is_superuser or request.user.user_type in [1, 2]:
            nstp_files = NSTPFile.objects.filter(is_archived=False)
        else:
            nstp_files = NSTPFile.objects.none()

        # Apply search filter
        if search_term:
            nstp_files = nstp_files.filter(
                Q(title__icontains=search_term) |
                Q(description__icontains=search_term)
            )

        # Apply category filter
        if category_filter != 'all':
            nstp_files = nstp_files.filter(category=category_filter)

        # Apply semester filter
        if semester_filter != 'all':
            nstp_files = nstp_files.filter(semester=semester_filter)

        # Apply sorting
        sort_mapping = {
            'title': 'title',
            'category': 'category',
            'semester': 'semester',
            'year': 'school_year',
            'date': 'created_at',
            'size': 'file__size'
        }

        if sort_column in sort_mapping:
            sort_field = sort_mapping[sort_column]
            if sort_direction == 'desc':
                sort_field = f'-{sort_field}'
            nstp_files = nstp_files.order_by(sort_field)
        else:
            # Default sorting
            nstp_files = nstp_files.order_by('-created_at')

        # Create paginator
        paginator = Paginator(nstp_files.select_related('created_by'), per_page)

        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        # Prepare data for JSON response
        nstp_data = []
        for file in page_obj.object_list:
            # Determine file icon based on extension
            file_extension = file.file.name.split('.')[-1].lower() if file.file.name else ''
            file_icon = 'bx bxs-file'
            if file_extension == 'pdf':
                file_icon = 'bx bxs-file-pdf'
            elif file_extension in ['doc', 'docx']:
                file_icon = 'bx bxs-file-doc'
            elif file_extension in ['xls', 'xlsx']:
                file_icon = 'bx bxs-file-xls'

            nstp_data.append({
                'id': file.id,
                'title': file.title,
                'description': file.description or "No description provided",
                'category': file.category,
                'category_display': file.get_category_display(),
                'semester': file.semester,
                'semester_display': file.get_semester_display(),
                'school_year': file.school_year,
                'created_at': file.created_at.isoformat(),
                'file_url': file.file.url,
                'file_size': file.file.size,
                'file_icon': file_icon,
                'can_view': request.user.has_perm('osas.view_nstpfile'),
                'can_edit': request.user.has_perm('osas.change_nstpfile'),
                'can_delete': request.user.has_perm('osas.delete_nstpfile')
            })

        return JsonResponse({
            'nstp_files': nstp_data,
            'pagination': {
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'current_page': page_obj.number,
                'num_pages': paginator.num_pages,
                'total_count': paginator.count,
                'start_index': page_obj.start_index(),
                'end_index': page_obj.end_index(),
            }
        })

    def add_nstp_files_statistics(self, context):
        # Get NSTP files based on user permissions
        if self.request.user.is_superuser or self.request.user.user_type in [1, 2]:
            nstp_files = NSTPFile.objects.filter(is_archived=False)
        else:
            nstp_files = NSTPFile.objects.none()

        # Total count
        context['total_nstp_files'] = nstp_files.count()

        # Semester counts
        context['first_sem_count'] = nstp_files.filter(semester='1st_semester').count()
        context['second_sem_count'] = nstp_files.filter(semester='2nd_semester').count()

        # Category counts
        context['accomplishment_reports_count'] = nstp_files.filter(category='accomplishment_reports').count()
        context['communication_letters_count'] = nstp_files.filter(category='communication_letters').count()
        context['financial_plan_count'] = nstp_files.filter(category='financial_plan').count()
        context['letters_count'] = nstp_files.filter(category='letters').count()
        context['moa_count'] = nstp_files.filter(category='moa').count()
        context['nstp_files_count'] = nstp_files.filter(category='nstp_files').count()
        context['recommendation_count'] = nstp_files.filter(category='recommendation').count()
        context['schedule_count'] = nstp_files.filter(category='schedule').count()

    def get_filtered_scholarships_ajax(self, request):
        search_term = request.GET.get('search', '').lower()
        scholarship_type_filter = request.GET.get('scholarship_type', 'all')
        status_filter = request.GET.get('status', 'all')
        sort_column = request.GET.get('sort', 'created_at')
        sort_direction = request.GET.get('direction', 'desc')
        page_number = request.GET.get('scholarship_page', 1)
        per_page = int(request.GET.get('per_page', 10))

        if request.user.is_superuser or request.user.user_type == 1:
            scholarships = Scholarship.objects.filter(is_archived=False)
        elif request.user.user_type == 14:  # Student
            today = timezone.now().date()
            scholarships = Scholarship.objects.filter(is_active=True, is_archived=False)
        else:
            scholarships = Scholarship.objects.filter(
                Q(created_by=request.user) | Q(is_active=True),
                is_archived=False
            )

        # Apply search filter
        if search_term:
            scholarships = scholarships.filter(
                Q(name__icontains=search_term) |
                Q(description__icontains=search_term)
            )

        # Apply scholarship type filter
        if scholarship_type_filter != 'all':
            scholarships = scholarships.filter(scholarship_type=scholarship_type_filter)

        # Apply status filter
        if status_filter != 'all':
            is_active = status_filter == 'active'
            scholarships = scholarships.filter(is_active=is_active)

        sort_mapping = {
            'name': 'name',
            'type': 'scholarship_type',
            'created_at': 'created_at',
            'is_active': 'is_active'
        }

        if sort_column in sort_mapping:
            sort_field = sort_mapping[sort_column]
            if sort_direction == 'desc':
                sort_field = f'-{sort_field}'
            scholarships = scholarships.order_by(sort_field)
        else:
            # Default sorting
            scholarships = scholarships.order_by('-created_at')

        # Create paginator
        paginator = Paginator(scholarships.select_related('created_by', 'application_form'), per_page)

        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        scholarship_data = []
        for scholarship in page_obj.object_list:
            scholarship_data.append({
                'id': scholarship.id,
                'name': scholarship.name,
                'type': scholarship.scholarship_type,
                'type_display': scholarship.get_scholarship_type_display(),
                'is_active': scholarship.is_active,
                'created_at': scholarship.created_at.isoformat(),
                'can_view': request.user.has_perm('osas.view_scholarship'),
                'can_edit': request.user.has_perm('osas.change_scholarship'),
                'can_delete': request.user.has_perm('osas.delete_scholarship')
            })

        return JsonResponse({
            'scholarships': scholarship_data,
            'pagination': {
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'current_page': page_obj.number,
                'num_pages': paginator.num_pages,
                'total_count': paginator.count,
                'start_index': page_obj.start_index(),
                'end_index': page_obj.end_index(),
            }
        })

    def add_scholarship_statistics(self, context):
        # Get all non-archived scholarships
        scholarships = Scholarship.objects.filter(is_archived=False)

        # Count by type
        context['total_scholarships'] = scholarships.count()
        context['public_scholarships_count'] = scholarships.filter(scholarship_type='public').count()
        context['private_scholarships_count'] = scholarships.filter(scholarship_type='private').count()

        # Count by status
        context['active_scholarships_count'] = scholarships.filter(is_active=True).count()
        context['inactive_scholarships_count'] = scholarships.filter(is_active=False).count()

    def get_filtered_applications(self, request):
        search_term = request.GET.get('search', '').lower()
        scholarship_filter = request.GET.get('scholarship', 'all')
        status_filter = request.GET.get('status', 'all')
        date_filter = request.GET.get('date', 'all')
        sort_column = request.GET.get('sort', 'application_date')
        sort_direction = request.GET.get('direction', 'desc')
        page_number = request.GET.get('scholarship_app_page', 1)
        per_page = int(request.GET.get('per_page', 10))

        # Start with base queryset
        if request.user.is_superuser or request.user.user_type in [1, 5]:
            applications = ScholarshipApplication.objects.filter(
                is_archived=False
            ).select_related('student', 'scholarship')
        elif request.user.user_type == 14:  # Student
            applications = ScholarshipApplication.objects.filter(
                student=request.user,
                is_archived=False
            ).select_related('scholarship')
        else:
            applications = ScholarshipApplication.objects.none()

        # Apply search filter
        if search_term:
            applications = applications.filter(
                Q(student__first_name__icontains=search_term) |
                Q(student__last_name__icontains=search_term) |
                Q(scholarship__name__icontains=search_term)
            )

        # Apply scholarship filter
        if scholarship_filter != 'all':
            applications = applications.filter(scholarship_id=scholarship_filter)

        # Apply sorting
        sort_mapping = {
            'student': 'student__last_name',
            'scholarship': 'scholarship__name',
            'date': 'application_date',
            'status': 'status'
        }

        if sort_column in sort_mapping:
            sort_field = sort_mapping[sort_column]
            if sort_direction == 'desc':
                sort_field = f'-{sort_field}'
            applications = applications.order_by(sort_field)
        else:
            # Default sorting
            applications = applications.order_by('-application_date')

        # Create paginator
        paginator = Paginator(applications, per_page)

        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        # Prepare data for JSON response
        application_data = []
        for application in page_obj.object_list:
            application_data.append({
                'id': application.id,
                'student_name': application.student.get_full_name() if application.student else 'Unknown',
                'scholarship_name': application.scholarship.name if application.scholarship else 'Unknown',
                'application_date': application.application_date.isoformat(),
                'status': application.status,
                'status_display': application.get_status_display(),
                'application_form_url': application.application_form.url if application.application_form else None,
                'cog_url': application.cog.url if application.cog else None,
                'cor_url': application.cor.url if application.cor else None,
                'can_approve': request.user.is_superuser or request.user.user_type in [1, 5],
                'can_edit': (application.status == 'pending' or
                             request.user.is_superuser or
                             request.user.user_type in [1, 5]),
                'can_archive': (application.status == 'pending' or
                                request.user.is_superuser or
                                request.user.user_type in [1, 5])
            })

        return JsonResponse({
            'applications': application_data,
            'pagination': {
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'current_page': page_obj.number,
                'num_pages': paginator.num_pages,
                'total_count': paginator.count,
                'start_index': page_obj.start_index(),
                'end_index': page_obj.end_index(),
            }
        })

    def add_scholarship_application_statistics(self, context):
        if self.request.user.is_superuser or self.request.user.user_type in [1, 5]:
            # Admin/Staff can see all applications
            applications = ScholarshipApplication.objects.filter(is_archived=False)
            context['scholarship_total_applications'] = applications.count()
            context['scholarship_approved_applications'] = applications.filter(status='approved').count()
            context['scholarship_pending_applications'] = applications.filter(status='pending').count()
            context['scholarship_rejected_applications'] = applications.filter(status='rejected').count()

        elif self.request.user.user_type == 14:  # Student
            # Students can only see their own applications
            applications = ScholarshipApplication.objects.filter(
                student=self.request.user,
                is_archived=False
            )
            context['total_applications'] = applications.count()
            context['approved_applications'] = applications.filter(status='approved').count()
            context['pending_applications'] = applications.filter(status='pending').count()
            context['rejected_applications'] = applications.filter(status='rejected').count()

        else:
            # Other user types - set all to 0
            context['total_applications'] = 0
            context['approved_applications'] = 0
            context['pending_applications'] = 0
            context['rejected_applications'] = 0

    def get_paginated_ojt_companies(self):
        companies_queryset = OJTCompany.objects.filter(is_archived=False)

        # For students (user_type 14), hide inactive companies
        if self.request.user.user_type == 14:  # Student
            companies_queryset = companies_queryset.filter(status='active')

        return self.get_paginated_data(companies_queryset.order_by('-created_at'), 'company_page')

    def get_filtered_ojt_companies(self, request):
        search_term = request.GET.get('search', '').lower()
        status_filter = request.GET.get('status_filter', '')
        sort_order = request.GET.get('sort_order', 'created_desc')
        page_number = request.GET.get('page', 1)
        per_page = 10

        companies = OJTCompany.objects.all()

        # Apply status filter
        if status_filter == 'archived':
            companies = companies.filter(is_archived=True)
        elif status_filter == 'active':
            companies = companies.filter(is_archived=False, status='active')
        elif status_filter == 'inactive':
            companies = companies.filter(is_archived=False, status='inactive')
        else:
            # Default: show all non-archived companies
            companies = companies.filter(is_archived=False)

        # Apply search filter
        if search_term:
            companies = companies.filter(
                Q(name__icontains=search_term) |
                Q(address__icontains=search_term) |
                Q(contact_number__icontains=search_term) |
                Q(description__icontains=search_term) |
                Q(email__icontains=search_term)
            )

        # Apply sorting
        sort_mapping = {
            'name_asc': 'name',
            'name_desc': '-name',
            'created_asc': 'created_at',
            'created_desc': '-created_at',
            'updated_asc': 'updated_at',
            'updated_desc': '-updated_at',
            'status_asc': 'status',
            'status_desc': '-status'
        }

        if sort_order in sort_mapping:
            companies = companies.order_by(sort_mapping[sort_order])
        else:
            # Default to newest first
            companies = companies.order_by('-created_at')

        # Create paginator
        paginator = Paginator(companies, per_page)

        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        # Prepare data for JSON response
        company_data = []
        for company in page_obj.object_list:
            # Check user permissions for actions
            user_type = request.user.user_type
            is_superuser = request.user.is_superuser

            # User type 1 (Super Admin) and 13 (Job Placement) can edit/archive/change status
            can_edit = is_superuser or user_type in [1, 13]
            can_archive = is_superuser or user_type in [1, 13]
            can_change_status = is_superuser or user_type in [1, 13]

            company_data.append({
                'id': company.id,
                'name': company.name,
                'address': company.address,
                'contact_number': company.contact_number,
                'email': company.email or '',
                'description': company.description or '',
                'website': company.website or '',
                'status': company.status,
                'status_display': company.get_status_display(),  # Add this line
                'is_archived': company.is_archived,
                'created_at': company.created_at.isoformat() if company.created_at else None,
                'updated_at': company.updated_at.isoformat() if company.updated_at else None,
                'can_edit': can_edit,
                'can_archive': can_archive,
                'can_change_status': can_change_status,
                'can_view': True,
                'current_user_type': user_type,
                'is_superuser': is_superuser
            })

        return JsonResponse({
            'companies': company_data,
            'pagination': {
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'current_page': page_obj.number,
                'num_pages': paginator.num_pages,
                'total_count': paginator.count,
                'start_index': page_obj.start_index(),
                'end_index': page_obj.end_index(),
            },
            'current_user_type': request.user.user_type,
            'is_superuser': request.user.is_superuser
        })

    def add_ojt_statistics(self, context):
        # Total companies (including archived)
        context['total_companies'] = OJTCompany.objects.count()

        # Archived companies count
        context['archived_companies'] = OJTCompany.objects.filter(is_archived=True).count()

        # Active companies (non-archived and status = active)
        context['active_companies'] = OJTCompany.objects.filter(
            is_archived=False,
            status='active'
        ).count()

        # Inactive companies (non-archived and status = inactive)
        context['inactive_companies'] = OJTCompany.objects.filter(
            is_archived=False,
            status='inactive'
        ).count()

    def add_organization_statistics(self, context):
        from django.db.models import Sum

        organizations = Organization.objects.filter(
            is_archived=False  # Exclude only archived
        )

        # Total count (excluding only archived)
        context['total_organizations'] = organizations.count()

        # Status counts - include ALL statuses including cancelled
        context['active_organizations_count'] = organizations.filter(_organization_status='active').count()
        context['pending_organizations_count'] = organizations.filter(_organization_status='pending').count()
        context['inactive_organizations_count'] = organizations.filter(_organization_status='inactive').count()
        context['expired_organizations_count'] = organizations.filter(_organization_status='expired').count()
        context['rejected_organizations_count'] = organizations.filter(_organization_status='rejected').count()
        context['cancelled_organizations_count'] = organizations.filter(_organization_status='cancelled').count()

        # Count of archived organizations (for reference)
        context['archived_organizations_count'] = Organization.objects.filter(is_archived=True).count()

        # Type counts
        context['student_organizations_count'] = organizations.filter(organization_type='student').count()
        context['sociocultural_organizations_count'] = organizations.filter(organization_type='sociocultural').count()

        # Total renewals count (sum of all renew_count fields)
        context['total_renewals_count'] = organizations.aggregate(
            total_renewals=Sum('renew_count')
        )['total_renewals'] or 0

        # Organizations with renewals count
        context['organizations_with_renewals'] = organizations.filter(
            renew_count__gt=0
        ).count()

        # Average renewals per organization
        total_orgs = context['total_organizations']
        if total_orgs > 0:
            context['average_renewals'] = round(context['total_renewals_count'] / total_orgs, 1)
        else:
            context['average_renewals'] = 0

    def get_filtered_organizations(self):
        queryset = Organization.objects.filter(is_archived=False)

        # Apply user-based filtering
        if self.request.user.is_superuser or self.request.user.user_type in [1, 10]:
            queryset = queryset
        elif self.request.user.user_type == 15:
            if hasattr(self.request.user, 'organization') and self.request.user.organization:
                queryset = queryset.filter(id=self.request.user.organization.id)
            else:
                queryset = Organization.objects.none()
        else:
            queryset = Organization.objects.none()

        queryset = queryset.order_by('-created_at')

        page = self.request.GET.get('organization_page', 1)
        paginator = Paginator(queryset, 10)

        try:
            return paginator.page(page)
        except PageNotAnInteger:
            return paginator.page(1)
        except EmptyPage:
            return paginator.page(paginator.num_pages)

    def get_filtered_organizations_ajax(self, request):
        search_term = request.GET.get('search', '').strip().lower()
        type_filter = request.GET.get('type', 'all')
        status_filter = request.GET.get('status', 'all')
        page_number = request.GET.get('organization_page', 1)
        per_page = 10

        organizations = Organization.objects.filter(is_archived=False)

        # Apply user-based filtering
        if request.user.is_superuser or request.user.user_type in [1, 10]:
            organizations = organizations
        elif request.user.user_type == 15:
            if hasattr(request.user, 'organization') and request.user.organization:
                organizations = organizations.filter(id=request.user.organization.id)
            else:
                organizations = Organization.objects.none()
        else:
            organizations = Organization.objects.none()

        # Apply search filter
        if search_term:
            organizations = organizations.filter(
                Q(organization_name__icontains=search_term) |
                Q(organization_acronym__icontains=search_term) |
                Q(organization_email__icontains=search_term) |
                Q(organization_adviser_name__icontains=search_term) |
                Q(username__icontains=search_term)
            )

        # Apply type filter
        if type_filter != 'all':
            organizations = organizations.filter(organization_type=type_filter)

        # Apply status filter
        if status_filter != 'all':
            organizations = organizations.filter(_organization_status=status_filter)

        organizations = organizations.order_by('-created_at')

        # Create paginator
        paginator = Paginator(organizations, per_page)

        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        organization_data = []
        for org in page_obj.object_list:
            # Get status display using the correct method name
            status_display = org.get__organization_status_display()

            # Determine if user can perform actions
            can_edit = request.user.is_superuser or request.user.user_type in [1, 10]
            can_archive = can_edit
            can_approve = can_edit and org.organization_status == 'pending' and org.all_requirements_submitted
            can_renew = can_edit and (org.organization_status == 'expired' or org.organization_needs_renewal)
            can_reactivate = can_edit and org.organization_status == 'cancelled'

            organization_data.append({
                'id': org.id,
                'organization_name': org.organization_name,
                'organization_acronym': org.organization_acronym,
                'organization_email': org.organization_email,
                'organization_type': org.organization_type,
                'organization_type_display': org.get_organization_type_display(),
                'organization_status': org.organization_status,  # Use the property
                'organization_status_display': status_display,
                'organization_valid_until': org.organization_valid_until.isoformat() if org.organization_valid_until else None,
                'renew_count': org.renew_count,
                'organization_member_count': org.organization_member_count,
                'organization_logo_url': org.organization_logo.url if org.organization_logo else None,
                'created_at': org.created_at.isoformat(),
                'all_requirements_submitted': org.all_requirements_submitted,
                'organization_needs_renewal': org.organization_needs_renewal,
                'can_edit': can_edit,
                'can_archive': can_archive,
                'can_approve': can_approve,
                'can_renew': can_renew,
                'can_reactivate': can_reactivate,
                'can_view': True
            })

        return JsonResponse({
            'success': True,
            'organizations': organization_data,
            'pagination': {
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'current_page': page_obj.number,
                'num_pages': paginator.num_pages,
                'total_count': paginator.count,
                'start_index': page_obj.start_index(),
                'end_index': page_obj.end_index(),
            }
        })

    def add_certificates_data(self, context):
        if self.request.user.user_type in [1, 10] or self.request.user.is_superuser:
            certificates = Certificate.objects.filter(is_active=True)
            # Get all organizations for the filter dropdown
            context['all_organizations'] = Organization.objects.filter(
                is_archived=False
            ).order_by('organization_name')
        else:
            certificates = Certificate.objects.filter(
                organization__username=self.request.user.username,
                is_active=True
            )
            # For regular users, only show their organization
            context['all_organizations'] = Organization.objects.filter(
                username=self.request.user.username,
                is_archived=False
            ).order_by('organization_name')

        # Apply pagination for certificates
        context['certificates'] = self.get_paginated_certificates()
        context['total_certificates'] = certificates.count()

        # Add certificate statistics
        self.add_certificate_statistics(context)

    def add_certificate_statistics(self, context):
        if self.request.user.user_type in [1, 10] or self.request.user.is_superuser:
            certificates = Certificate.objects.filter(is_active=True)
        else:
            certificates = Certificate.objects.filter(
                organization__username=self.request.user.username,
                is_active=True
            )

        context['total_certificates_count'] = certificates.count()

        # Count by organization type
        context['student_org_certificates'] = certificates.filter(
            organization__organization_type='student'
        ).count()
        context['sociocultural_org_certificates'] = certificates.filter(
            organization__organization_type='sociocultural'
        ).count()

        # Recent certificates (last 30 days)
        thirty_days_ago = timezone.now() - timedelta(days=30)
        context['recent_certificates'] = certificates.filter(
            created_at__gte=thirty_days_ago
        ).count()

    def get_paginated_certificates(self):
        if self.request.user.user_type in [1, 10] or self.request.user.is_superuser:
            certificates_queryset = Certificate.objects.filter(is_active=True)
        else:
            certificates_queryset = Certificate.objects.filter(
                organization__username=self.request.user.username,
                is_active=True
            )

        return self.get_paginated_data(
            certificates_queryset.select_related('organization', 'generated_by').order_by('-issue_date'),
            'certificate_page',
            per_page=12
        )

    def get_filtered_certificates_ajax(self, request):
        search_term = request.GET.get('search', '').strip().lower()
        organization_filter = request.GET.get('organization', 'all')
        date_filter = request.GET.get('date', 'newest')
        type_filter = request.GET.get('type', 'all')
        page_number = request.GET.get('certificate_page', 1)
        per_page = 12

        # Base queryset based on user permissions
        if request.user.user_type in [1, 10] or request.user.is_superuser:
            certificates = Certificate.objects.filter(is_active=True)
        else:
            certificates = Certificate.objects.filter(
                organization__username=request.user.username,
                is_active=True
            )

        # Apply search filter
        if search_term:
            certificates = certificates.filter(
                Q(organization__organization_name__icontains=search_term) |
                Q(organization__organization_acronym__icontains=search_term) |
                Q(venue__icontains=search_term) |
                Q(organization__organization_adviser_name__icontains=search_term)
            )

        # Apply organization filter
        if organization_filter != 'all':
            certificates = certificates.filter(organization_id=organization_filter)

        # Apply organization type filter
        if type_filter != 'all':
            certificates = certificates.filter(organization__organization_type=type_filter)

        # Apply date sorting
        if date_filter == 'newest':
            certificates = certificates.order_by('-issue_date')
        elif date_filter == 'oldest':
            certificates = certificates.order_by('issue_date')
        elif date_filter == 'recently_updated':
            certificates = certificates.order_by('-updated_at')
        elif date_filter == 'recently_created':
            certificates = certificates.order_by('-created_at')

        # Create paginator
        paginator = Paginator(certificates.select_related(
            'organization', 'generated_by'
        ), per_page)

        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        # Prepare data for JSON response
        certificate_data = []
        for certificate in page_obj.object_list:
            # Determine file type for display
            file_type = 'pdf' if certificate.certificate_file and certificate.certificate_file.name.lower().endswith(
                '.pdf') else 'image'

            certificate_data.append({
                'id': certificate.id,
                'organization_name': certificate.organization.organization_name,
                'organization_acronym': certificate.organization.organization_acronym,
                'organization_type': certificate.organization.organization_type,
                'organization_type_display': certificate.organization.get_organization_type_display(),
                'issue_date': certificate.issue_date.isoformat(),
                'venue': certificate.venue,
                'certificate_url': certificate.certificate_file.url if certificate.certificate_file else None,
                'file_type': file_type,
                'created_at': certificate.created_at.isoformat(),
                'generated_by': certificate.generated_by.get_full_name() if certificate.generated_by else 'System',
                'can_download': True,
                'can_view': True
            })

        return JsonResponse({
            'success': True,
            'certificates': certificate_data,
            'pagination': {
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'current_page': page_obj.number,
                'num_pages': paginator.num_pages,
                'total_count': paginator.count,
                'start_index': page_obj.start_index(),
                'end_index': page_obj.end_index(),
            }
        })

    def add_accomplishment_data(self, context):
        if self.request.user.is_superuser or self.request.user.user_type in [1, 10]:
            accomplishment_queryset = AccomplishmentRecord.objects.filter(is_archived=False)
        elif self.request.user.user_type == 15:
            # Organizations can only see their own reports
            if hasattr(self.request.user, 'organization') and self.request.user.organization:
                accomplishment_queryset = AccomplishmentRecord.objects.filter(
                    organization=self.request.user.organization,
                    is_archived=False
                )
            else:
                accomplishment_queryset = AccomplishmentRecord.objects.none()
        else:
            accomplishment_queryset = AccomplishmentRecord.objects.none()

        # Apply pagination
        context['accomplishment_reports'] = self.get_paginated_accomplishment_reports()

        # Add statistics
        self.add_accomplishment_statistics(context)

    def get_paginated_accomplishment_reports(self):
        if self.request.user.is_superuser or self.request.user.user_type in [1, 10]:
            accomplishment_queryset = AccomplishmentRecord.objects.filter(is_archived=False)
        elif self.request.user.user_type == 15:
            if hasattr(self.request.user, 'organization') and self.request.user.organization:
                accomplishment_queryset = AccomplishmentRecord.objects.filter(
                    organization=self.request.user.organization,
                    is_archived=False
                )
            else:
                accomplishment_queryset = AccomplishmentRecord.objects.none()
        else:
            accomplishment_queryset = AccomplishmentRecord.objects.none()

        return self.get_paginated_data(
            accomplishment_queryset.select_related('organization', 'submitted_by').order_by('-date_conducted',
                                                                                            '-created_at'),
            'accomplishment_page'
        )

    def add_accomplishment_statistics(self, context):
        if self.request.user.is_superuser or self.request.user.user_type in [1, 10]:
            reports = AccomplishmentRecord.objects.filter(is_archived=False)
        elif self.request.user.user_type == 15:
            if hasattr(self.request.user, 'organization') and self.request.user.organization:
                reports = AccomplishmentRecord.objects.filter(
                    organization=self.request.user.organization,
                    is_archived=False
                )
            else:
                reports = AccomplishmentRecord.objects.none()
        else:
            reports = AccomplishmentRecord.objects.none()

        context['total_accomplishment_reports'] = reports.count()

        # Count by record type
        context['event_reports_count'] = reports.filter(record_type='event').count()
        context['meeting_reports_count'] = reports.filter(record_type='meeting').count()
        context['training_reports_count'] = reports.filter(record_type='training').count()
        context['community_reports_count'] = reports.filter(record_type='community').count()
        context['achievement_reports_count'] = reports.filter(record_type='achievement').count()
        context['other_reports_count'] = reports.filter(record_type='other').count()

        # Count by semester
        context['first_sem_reports'] = reports.filter(semester='1st').count()
        context['second_sem_reports'] = reports.filter(semester='2nd').count()
        context['summer_reports'] = reports.filter(semester='summer').count()

    def get_filtered_accomplishment_reports_ajax(self, request):
        search_term = request.GET.get('search', '').strip().lower()
        type_filter = request.GET.get('report_type', 'all')
        semester_filter = request.GET.get('semester', 'all')
        school_year_filter = request.GET.get('school_year', 'all')
        page_number = request.GET.get('accomplishment_page', 1)
        per_page = 10

        # Base queryset based on user permissions
        if request.user.is_superuser or request.user.user_type in [1, 10]:
            reports = AccomplishmentRecord.objects.filter(is_archived=False)
        elif request.user.user_type == 15:
            if hasattr(request.user, 'organization') and request.user.organization:
                reports = AccomplishmentRecord.objects.filter(
                    organization=request.user.organization,
                    is_archived=False
                )
            else:
                reports = AccomplishmentRecord.objects.none()
        else:
            reports = AccomplishmentRecord.objects.none()

        # Apply search filter
        if search_term:
            reports = reports.filter(
                Q(title__icontains=search_term) |
                Q(organization__organization_name__icontains=search_term) |
                Q(organization__organization_acronym__icontains=search_term) |
                Q(venue__icontains=search_term) |
                Q(description__icontains=search_term)
            )

        # Apply type filter
        if type_filter != 'all':
            reports = reports.filter(record_type=type_filter)

        # Apply semester filter
        if semester_filter != 'all':
            reports = reports.filter(semester=semester_filter)

        # Apply school year filter
        if school_year_filter != 'all':
            reports = reports.filter(school_year=school_year_filter)

        # Apply default sorting
        reports = reports.order_by('-date_conducted', '-created_at')

        # Create paginator
        paginator = Paginator(reports.select_related(
            'organization', 'submitted_by'
        ), per_page)

        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        # Prepare data for JSON response
        report_data = []
        for report in page_obj.object_list:
            # Determine permission flags
            is_owner = report.organization.user_account == request.user if report.organization and report.organization.user_account else False
            is_admin_user = request.user.is_superuser or request.user.user_type in [1, 10]
            is_organization_user = request.user.user_type == 15

            # Edit permissions: admin OR organization owner
            can_edit = is_admin_user or (is_organization_user and is_owner)

            report_data.append({
                'id': report.id,
                'title': report.title,
                'organization_name': report.organization.organization_name if report.organization else 'N/A',
                'organization_acronym': report.organization.organization_acronym if report.organization else 'N/A',
                'record_type': report.record_type,
                'record_type_display': report.get_record_type_display(),
                'date_conducted': report.date_conducted.isoformat(),
                'venue': report.venue or 'Not specified',
                'semester': report.semester,
                'semester_display': report.get_semester_display(),
                'school_year': report.school_year,
                'number_of_participants': report.number_of_participants,
                'duration_hours': float(report.duration_hours),
                'budget_utilized': float(report.budget_utilized) if report.budget_utilized else None,
                'submitted_by': report.submitted_by.get_full_name() if report.submitted_by else 'Unknown',
                'submitted_at': report.created_at.isoformat(),
                'has_main_report': bool(report.main_report),
                'can_edit': can_edit,
                'can_archive': can_edit,
                'can_view': True,
            })

        return JsonResponse({
            'success': True,
            'accomplishment_reports': report_data,
            'pagination': {
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'current_page': page_obj.number,
                'num_pages': paginator.num_pages,
                'total_count': paginator.count,
                'start_index': page_obj.start_index(),
                'end_index': page_obj.end_index(),
            }
        })
    

# -------------------------------------------------- Users Profile -----------------------------------------------------
class ProfileView(LoginRequiredMixin, View):
    template_name = 'osas/dashboard.html'

    def get(self, request, *args, **kwargs):
        active_tab = request.GET.get('tab', 'basic')
        context = {
            'active_tab': active_tab,
            'profile_form': UserProfileForm(instance=request.user),
            'account_form': AccountInfoForm(instance=request.user),
            'password_form': CustomPasswordChangeForm(request.user),
        }
        return render(request, self.template_name, context)


class UpdateProfileView(LoginRequiredMixin, View):
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        form = UserProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            user = form.save()

            # Handle organization updates if user is an organization
            if user.is_organization and hasattr(user, 'organization_account'):
                organization = user.organization_account
                organization_data = {
                    'organization_name': request.POST.get('organization_name'),
                    'organization_acronym': request.POST.get('organization_acronym'),
                    'organization_type': request.POST.get('organization_type'),
                    'organization_email': request.POST.get('organization_email'),
                    'organization_adviser_name': request.POST.get('organization_adviser_name'),
                    'organization_adviser_department': request.POST.get('organization_adviser_department'),
                    'organization_description': request.POST.get('organization_description'),
                    'organization_mission': request.POST.get('organization_mission'),
                    'organization_vision': request.POST.get('organization_vision'),
                }

                # Update organization fields
                for field, value in organization_data.items():
                    if value is not None:
                        setattr(organization, field, value)
                organization.save()

            # Prepare response data
            response_data = {
                'success': True,
                'user': {
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'gender_display': user.get_gender_display(),
                    'birth_date': user.birth_date.strftime('%B %d, %Y') if user.birth_date else 'Not specified',
                    'address': user.address or 'Not specified',
                    'phone_number': user.phone_number or 'Not specified',
                    'student_number': user.student_number or 'Not specified',
                    'section': user.section or 'Not specified',
                    'department': user.department or 'Not specified',
                },
                'message': 'Profile updated successfully!'
            }

            # Add OSAS-specific fields
            if user.is_osas_unit:
                response_data['user']['position'] = user.get_position_display() or 'Not specified'

            # Add student-specific fields (course is display-only)
            if user.is_student:
                year_level_display = 'Not specified'
                if user.year_level:
                    if user.year_level == "1":
                        year_level_display = "1st Year"
                    elif user.year_level == "2":
                        year_level_display = "2nd Year"
                    elif user.year_level == "3":
                        year_level_display = "3rd Year"
                    elif user.year_level == "4":
                        year_level_display = "4th Year"
                    elif user.year_level == "5":
                        year_level_display = "5th Year"
                    else:
                        year_level_display = user.year_level

                response_data['user']['year_level'] = year_level_display
                response_data['user']['course'] = user.course.name if user.course else 'Not specified'

            # Add profile picture URL if available
            if user.profile_picture:
                response_data['profile_picture_url'] = user.profile_picture.url

            return JsonResponse(response_data)

        return JsonResponse({
            'success': False,
            'errors': form.errors.get_json_data(),
            'message': 'Please correct the errors below.'
        }, status=400)


class UpdateAccountInfoView(LoginRequiredMixin, View):
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        form = AccountInfoForm(request.POST, instance=request.user)
        if form.is_valid():
            user = form.save()
            return JsonResponse({
                'success': True,
                'user': {
                    'username': user.username,
                    'email': user.email,
                    'user_type_display': user.get_user_type_display(),
                    'is_active': user.is_active
                },
                'message': 'Account information updated successfully!'
            })
        return JsonResponse({
            'success': False,
            'errors': form.errors.get_json_data(),
            'message': 'Please correct the errors below.'
        }, status=400)


class CustomPasswordChangeView(PasswordChangeView):
    def form_valid(self, form):
        form.save()
        # Check for AJAX request
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Password updated successfully.'
            })
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        # Handle AJAX invalid form
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = {f: e.get_json_data() for f, e in form.errors.items()}
            return JsonResponse({
                'success': False,
                'message': 'Please correct the errors below.',
                'errors': errors
            }, status=400)
        else:
            return super().form_invalid(form)


# ------------------------------------------------------ API Course ----------------------------------------------------
def get_courses(request):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        courses = Course.objects.all().values('id', 'name')
        return JsonResponse({
            'success': True,
            'courses': list(courses)
        })
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)


# ------------------------------------------------- User Management Section --------------------------------------------
class ApproveUserView(View):
    @method_decorator(login_required)
    @method_decorator(permission_required('auth.change_user', raise_exception=True))
    @method_decorator(csrf_protect)
    @method_decorator(require_POST)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, user_id):
        try:
            user = CustomUser.objects.get(id=user_id)

            if not user.is_verified:
                user.is_verified = True
                user.save()

                self.assign_role_permissions(user)

                # Log the activity
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"{request.user.first_name} approved user: {user.username}"
                )

                try:
                    self.send_approval_email(user)
                    email_status = 'Notification email sent.'
                except Exception as e:
                    logger.error(f"Failed to send approval email: {str(e)}")
                    email_status = 'User approved but email notification failed.'

                return JsonResponse({
                    'success': True,
                    'message': f'User approved successfully. {email_status}'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'User is already verified'
                }, status=400)

        except CustomUser.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'User not found'
            }, status=404)
        except Exception as e:
            logger.error(f"Error in ApproveUserView: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': 'An unexpected error occurred'
            }, status=500)

    def assign_role_permissions(self, user):
        # Clear existing permissions first
        user.user_permissions.clear()

        # Super User (OSAS Staff) gets all permissions
        if user.user_type == 1:
            user.is_superuser = True
            user.is_staff = True
            user.save()
            return

        # Get all permissions
        all_permissions = Permission.objects.all()

        # Function to determine if a permission should be assigned
        def should_assign_permission(perm, role):
            perm_name = perm.name.lower()
            codename = perm.codename.lower()
            content_type = perm.content_type.model.lower()

            # Organization - Give full access to user types 1 and 10
            if 'organization' in content_type or 'organization' in perm_name:
                return role in [1, 10]

            # Students (#14) get NO announcement permissions at all
            if role == 14 and ('announcement' in perm_name or 'announcement' in content_type):
                return False

            # OJT Adviser (16) gets ALL announcement permissions (view, add, change, delete)
            if role == 16 and ('announcement' in perm_name or 'announcement' in content_type):
                return True  # Give all announcement permissions

            # Common permissions for all non-superadmin roles (except students and OJT Adviser)
            if role != 14 and ('announcement' in perm_name or 'announcement' in content_type):
                # Other OSAS roles get announcement permissions too
                return True

            # OJT Permissions
            if ('ojt' in perm_name or 'ojt' in codename or
                    'on-the-job' in perm_name or 'on_the_job' in codename or
                    'ojtcompany' in content_type or 'ojtapplication' in content_type or
                    'ojtreport' in content_type or 'ojtrequirement' in content_type):

                if role in [1, 13, 16]:  # Super Admin, Job Placement, OJT Adviser
                    return True

                # Student (type 14) permissions
                elif role == 14:
                    # View only for OJT Company
                    if 'ojtcompany' in codename:
                        return 'view' in codename

                    # All access for OJT Application
                    elif 'ojtapplication' in codename:
                        return True

                    # All access for OJT Reports
                    elif 'ojtreport' in codename:
                        return True

                    # All access for OJT Requirements
                    elif 'ojtrequirement' in codename:
                        return True

                    else:
                        return False

                else:
                    return False

            # Student Admission permissions for specific roles
            if ('student admission' in perm_name or 'studentadmission' in codename):
                return role in [1, 12, 14]  # Super Admin, Admission, Student

            # Scholarship Role (5) gets all scholarship permissions
            if role == 5 and ('scholarship' in perm_name or 'scholarship' in codename):
                return True

            # All roles get complaint permissions
            if ('complaint' in perm_name or 'complaint' in codename):
                return True

            # Only OSAS Staff, NSTP, and Student can access
            if 'nstpstudentinfo' in codename:
                return role in [1, 2, 14]

            if 'nstpfile' in codename:
                return role in [1, 2]

            # OJT Adviser permissions - FULL access
            if role == 16:
                # OJT Adviser gets ALL permissions for downloadables (view, add, change, delete)
                if 'downloadable' in perm_name or 'downloadable' in content_type:
                    return True

                # Also give access to other common features
                # View content type permission (if exists)
                if 'view content type' in perm_name:
                    return True

                # Return True for other relevant permissions
                # You can add more permissions here as needed

            # Role-specific permissions for other roles
            if role == 2:  # NSTP
                return ('downloadable' in perm_name and ('view' in perm_name or 'add' in perm_name))
            elif role == 3:  # Clinic
                return ('downloadable' in perm_name and ('view' in perm_name or 'add' in perm_name))
            elif role == 4:  # Alumni
                return ('downloadable' in perm_name and ('view' in perm_name or 'add' in perm_name))
            elif role == 5:  # Scholarship
                # Scholarship gets ALL downloadable permissions
                if 'downloadable' in perm_name or 'downloadable' in content_type:
                    return True
                return ('downloadable' in perm_name and ('view' in perm_name or 'add' in perm_name))
            elif role == 6:  # Culture and Arts
                return ('downloadable' in perm_name and ('view' in perm_name or 'add' in perm_name))
            elif role == 7:  # Sports Development
                return ('downloadable' in perm_name and ('view' in perm_name or 'add' in perm_name))
            elif role == 8:  # Guidance Counseling
                return ('downloadable' in perm_name and ('view' in perm_name or 'add' in perm_name))
            elif role == 9:  # Student Welfare Services
                return ('downloadable' in perm_name and ('view' in perm_name or 'add' in perm_name))
            elif role == 10:  # Student Development Services
                # SDS gets all downloadable permissions plus organization (handled above)
                if 'downloadable' in perm_name or 'downloadable' in content_type:
                    return True
                return ('downloadable' in perm_name and ('view' in perm_name or 'add' in perm_name))
            elif role == 11:  # Misdeamenor
                return ('downloadable' in perm_name and ('view' in perm_name or 'add' in perm_name))
            elif role == 12:  # Admission
                return ('downloadable' in perm_name and ('view' in perm_name or 'add' in perm_name))
            elif role == 13:  # Job Placement
                # Job Placement gets all downloadable permissions
                if 'downloadable' in perm_name or 'downloadable' in content_type:
                    return True
                return ('downloadable' in perm_name and ('view' in perm_name or 'add' in perm_name))
            elif role == 14:  # Student
                return (('downloadable' in perm_name and 'view' in perm_name) or
                        ('scholarship application' in perm_name and
                         ('view' in perm_name or 'change' in perm_name or 'delete' in perm_name)))
            else:
                return False

        # Assign permissions based on role
        for perm in all_permissions:
            if should_assign_permission(perm, user.user_type):
                user.user_permissions.add(perm)

        # Staff status for OSAS units and OJT Adviser (non-student, non-superadmin)
        if user.user_type in range(2, 14) or user.user_type == 16:
            user.is_staff = True
            user.save()

    def send_approval_email(self, user):
        subject = 'Your Account Has Been Approved'

        context = {
            'user': user,
            'site_name': 'CvSU Office of the Student Affairs and Services Bacoor City Campus',
        }

        try:
            html_message = render_to_string('emails/account_approved.html', context)
            plain_message = strip_tags(html_message)

            send_mail(
                subject=subject,
                message=plain_message,
                html_message=html_message,
                from_email='CvSU OSAS Bacoor City Campus <noreply@domain-name.com>',
                recipient_list=[user.email],
                fail_silently=False,
            )
        except BadHeaderError:
            raise ValueError("Invalid email header found.")
        except SMTPException as e:
            raise Exception(f"SMTP error occurred: {str(e)}")
        except Exception as e:
            raise Exception(f"Failed to send email: {str(e)}")


class UserListView(LoginRequiredMixin, ListView):
    model = CustomUser
    template_name = 'osas/dashboard.html'
    context_object_name = 'users'

    def get_queryset(self):
        if self.request.user.is_superuser:
            return CustomUser.objects.all().order_by('id')
        return CustomUser.objects.none()

    def get(self, request, *args, **kwargs):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            users = self.get_queryset()
            data = {
                'users': [{
                    'id': user.id,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'username': user.username,
                    'is_active': user.is_active,
                    'profile_picture': user.profile_picture.url if user.profile_picture else None,
                } for user in users]
            }
            return JsonResponse(data)
        return super().get(request, *args, **kwargs)

    def render_to_response(self, context, **response_kwargs):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            html = render_to_string('osas/includes/user_table.html', context)
            return JsonResponse({'html': html})
        return super().render_to_response(context, **response_kwargs)


class UserCreateView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')

        # Initialize the form with the courses queryset
        form = CustomUserCreationForm(request=request)
        form.fields['course'].queryset = Course.objects.all().order_by('name')

        # Get all permissions grouped by model
        permissions = Permission.objects.all()
        grouped_permissions = {}

        for perm in permissions:
            app_label, model = perm.content_type.app_label, perm.content_type.model
            key = f"{app_label}.{model}"
            if key not in grouped_permissions:
                grouped_permissions[key] = []
            grouped_permissions[key].append(perm)

        context = {
            'form': form,
            'grouped_permissions': grouped_permissions,
        }

        return render(request, 'osas/modals/create-user.html', context)

    def post(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({
                'success': False,
                'errors': {'__all__': ['Authentication required']}
            }, status=403)

        form = CustomUserCreationForm(request.POST, request.FILES, request=request)
        if form.is_valid():
            try:
                with transaction.atomic():
                    user = form.save(commit=False)

                    # Set additional fields based on user type
                    user_type = form.cleaned_data.get('user_type')

                    if str(user_type) == '14':  # Student
                        user.student_number = form.cleaned_data.get('student_number')
                        user.course = form.cleaned_data.get('course')
                        user.year_level = form.cleaned_data.get('year_level')
                        user.section = form.cleaned_data.get('section')
                        user.cor_photo = form.cleaned_data.get('cor_photo')

                    elif str(user_type) in [str(i) for i in range(1, 14)]:  # OSAS Staff
                        user.department = form.cleaned_data.get('department')
                        user.osas_position = form.cleaned_data.get('osas_position')
                        user.id_photo = form.cleaned_data.get('id_photo')

                    user.save()

                    # Handle permissions
                    permission_ids = request.POST.getlist('permissions', [])
                    if permission_ids:
                        permissions = Permission.objects.filter(id__in=permission_ids)
                        user.user_permissions.set(permissions)

                    # Log the activity
                    UserActivityLog.objects.create(
                        user=request.user,
                        activity=f"{request.user.first_name} created new user account for {user.username}",
                    )

                    return JsonResponse({
                        'success': True,
                        'user_id': user.id,
                        'username': user.username,
                        'email': user.email,
                        'user_type': user.get_user_type_display()
                    })

            except Exception as e:
                import traceback
                print("Error creating user:", str(e))
                print("Traceback:", traceback.format_exc())
                return JsonResponse({
                    'success': False,
                    'errors': {'__all__': [f'Error creating user: {str(e)}']}
                }, status=400)

        # Debug: Print form errors to console
        print("Form errors:", form.errors)
        print("Form non-field errors:", form.non_field_errors())

        errors = {}
        for field, error_list in form.errors.items():
            errors[field] = [str(error) for error in error_list]

        return JsonResponse({
            'success': False,
            'errors': errors
        }, status=400)


class UserDetailView(LoginRequiredMixin, DetailView):
    model = CustomUser
    context_object_name = 'user'

    def get(self, request, *args, **kwargs):
        user = self.get_object()
        self.object = self.get_object()

        # Get all permissions (both direct and through groups)
        permissions = user.get_all_permissions()
        permission_list = sorted(list(permissions))  # Convert to sorted list
        direct_permission_ids = list(user.user_permissions.values_list('id', flat=True))

        user_data = {
            'first_name': self.object.first_name,
            'last_name': self.object.last_name,
            'username': self.object.username,
            'email': self.object.email,
            'user_type_display': self.object.get_user_type_display(),
            'is_active': self.object.is_active,
            'date_joined': self.object.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
            'last_login': self.object.last_login.strftime('%Y-%m-%d %H:%M:%S') if self.object.last_login else None,
            'profile_picture': self.object.profile_picture.url if self.object.profile_picture else None,
            'permissions': permission_list,
            'permission_ids': direct_permission_ids,
            'gender': self.object.gender,
            'gender_display': self.object.get_gender_display(),
            'birth_date': self.object.birth_date.strftime('%Y-%m-%d') if self.object.birth_date else None,
            'phone_number': self.object.phone_number,
            'address': self.object.address,
            'student_number': self.object.student_number,
            'course': self.object.course.name if self.object.course else None,
            'course_id': self.object.course.id if self.object.course else None,
            'year_level': self.object.year_level,
            'section': self.object.section,
            'department': self.object.department,
            'position': self.object.position,
            'position_display': self.object.get_position_display() if self.object.position else None,
            'is_verified': self.object.is_verified,
            'is_student': self.object.is_student,
            'is_osas_unit': self.object.is_osas_unit,
            'id_photo': self.object.id_photo.url if self.object.id_photo else None,
            'cor_photo': self.object.cor_photo.url if self.object.cor_photo else None,
        }

        return JsonResponse({
            'success': True,
            'user': user_data
        })


class UserUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = CustomUser
    form_class = CustomUserUpdateForm
    template_name = 'osas/dashboard.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def test_func(self):
        return self.request.user.is_superuser

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        permission_ids = list(self.object.user_permissions.values_list('id', flat=True))

        # Get all courses
        courses = Course.objects.all().values('id', 'name')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            user_data = {
                'success': True,
                'user': {
                    'id': self.object.id,
                    'first_name': self.object.first_name,
                    'last_name': self.object.last_name,
                    'username': self.object.username,
                    'email': self.object.email,
                    'user_type': str(self.object.user_type),
                    'is_active': self.object.is_active,
                    'is_superuser': self.object.is_superuser,
                    'profile_picture': self.object.profile_picture.url if self.object.profile_picture else None,
                    'permissions': permission_ids,
                    'gender': self.object.gender,
                    'birth_date': self.object.birth_date.strftime('%Y-%m-%d') if self.object.birth_date else None,
                    'phone_number': self.object.phone_number,
                    'address': self.object.address,
                    'student_number': self.object.student_number,
                    'course': self.object.course.id if self.object.course else None,  # Send course ID
                    'year_level': self.object.year_level,
                    'section': self.object.section,
                    'department': self.object.department,
                    'osas_position': self.object.osas_position,
                    'id_photo': self.object.id_photo.url if self.object.id_photo else None,
                    'id_photo_name': os.path.basename(self.object.id_photo.name) if self.object.id_photo else None,
                    'cor_photo': self.object.cor_photo.url if self.object.cor_photo else None,
                    'cor_photo_name': os.path.basename(self.object.cor_photo.name) if self.object.cor_photo else None,
                },
                'courses': list(courses)  # Make sure this is included
            }
            return JsonResponse(user_data)
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        user = form.save(commit=False)
        user_type = form.cleaned_data.get('user_type')

        # Handle file clear requests
        if self.request.POST.get('profile_picture-clear') == 'true':
            if user.profile_picture:
                user.profile_picture.delete(save=False)
                user.profile_picture = None

        if self.request.POST.get('id_photo-clear') == 'true':
            if user.id_photo:
                user.id_photo.delete(save=False)
                user.id_photo = None

        if self.request.POST.get('cor_photo-clear') == 'true':
            if user.cor_photo:
                user.cor_photo.delete(save=False)
                user.cor_photo = None

        original_user_type = self.object.user_type
        new_user_type = form.cleaned_data.get('user_type')

        # Handle field cleanup when changing user types
        if original_user_type != new_user_type:
            if original_user_type == 14 and new_user_type != 14:  # Student to non-student
                if not form.cleaned_data.get('student_number'):
                    user.student_number = None
                if not form.cleaned_data.get('course'):
                    user.course = None
                if not form.cleaned_data.get('year_level'):
                    user.year_level = None
                if not form.cleaned_data.get('section'):
                    user.section = None

            elif original_user_type in range(1, 14) and new_user_type not in range(1, 14):  # OSAS unit to non-OSAS
                if not form.cleaned_data.get('department'):
                    user.department = None
                if not form.cleaned_data.get('osas_position'):
                    user.osas_position = None

            elif original_user_type == 16 and new_user_type != 16:  # OJT Adviser to non-OJT Adviser
                # OJT Adviser might have department, keep it if changing to OSAS role
                if new_user_type not in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13']:
                    if not form.cleaned_data.get('department'):
                        user.department = None
                    if not form.cleaned_data.get('osas_position'):
                        user.osas_position = None

        # Handle superuser status
        if new_user_type == '1':  # OSAS Staff
            user.is_superuser = True
        else:
            user.is_superuser = False
            if original_user_type == 1 and new_user_type != 1:
                user.user_permissions.clear()

        # Handle file uploads
        for field_name in ['id_photo', 'cor_photo']:
            if field_name in self.request.FILES:
                setattr(user, field_name, self.request.FILES[field_name])

        # Save the user first
        user.save()

        if new_user_type == '16' or 'permissions' in self.request.POST:
            selected_permissions = self.request.POST.getlist('permissions', [])

            if original_user_type == 1 and new_user_type != 1:
                user.user_permissions.clear()

            if new_user_type == '10':
                # Get all organization permissions
                org_permissions = Permission.objects.filter(
                    codename__contains='organization'
                )

                for perm in org_permissions:
                    if str(perm.id) not in selected_permissions:
                        selected_permissions.append(str(perm.id))

            if new_user_type == '16':
                # Get all OJT Company permissions
                ojt_company_permissions = Permission.objects.filter(
                    codename__contains='ojtcompany'
                )
                # Add OJT Company permissions to selected permissions
                for perm in ojt_company_permissions:
                    if str(perm.id) not in selected_permissions:
                        selected_permissions.append(str(perm.id))

                # Get all Announcement permissions
                announcement_permissions = Permission.objects.filter(
                    codename__contains='announcement'
                )
                # Add Announcement permissions to selected permissions
                for perm in announcement_permissions:
                    if str(perm.id) not in selected_permissions:
                        selected_permissions.append(str(perm.id))

                # Get all Downloadable permissions
                downloadable_permissions = Permission.objects.filter(
                    codename__contains='downloadable'
                )
                # Add Downloadable permissions to selected permissions
                for perm in downloadable_permissions:
                    if str(perm.id) not in selected_permissions:
                        selected_permissions.append(str(perm.id))

            # Also set staff status for OJT Adviser
            if new_user_type == '16':
                user.is_staff = True
                user.save()

            if selected_permissions:
                user.user_permissions.set(selected_permissions)

        # Log the activity
        activity_msg = f"{self.request.user.first_name} updated user {user.username}"
        if original_user_type != new_user_type:
            if new_user_type == '1':
                activity_msg += " (granted superuser as OSAS Staff)"
            elif original_user_type == 1 and new_user_type != '1':
                activity_msg += " (removed superuser status)"
            elif new_user_type == '16':
                activity_msg += " (assigned as OJT Adviser with full OJT, Announcement, and Downloadable access)"

        UserActivityLog.objects.create(
            user=self.request.user,
            activity=activity_msg
        )

        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'user_id': user.id,
                'username': user.username,
                'email': user.email,
                'user_type': user.get_user_type_display(),
                'is_superuser': user.is_superuser,
                'is_staff': user.is_staff
            })
        return super().form_valid(form)

    def form_invalid(self, form):
        print("Form errors:", form.errors)
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = {}
            for field, error_list in form.errors.items():
                errors[field] = [str(error) for error in error_list]
            return JsonResponse({
                'success': False,
                'errors': errors
            }, status=400)
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('dashboard')


class UserArchiveView(View):
    def post(self, request):
        if not request.user.is_superuser:
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to archive users'
            }, status=403)

        user_id = request.POST.get('user_id')
        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'User ID is required'
            }, status=400)

        try:
            user_to_archive = CustomUser.objects.get(id=user_id)

            if user_to_archive.id == request.user.id:
                return JsonResponse({
                    'success': False,
                    'error': 'You cannot archive your own account'
                }, status=400)

            if user_to_archive.is_superuser:
                return JsonResponse({
                    'success': False,
                    'error': 'Cannot archive a super-admin account.'
                }, status=400)

            # Archive the user
            user_to_archive.is_archived = True
            user_to_archive.is_active = False
            user_to_archive.archived_at = timezone.now()
            user_to_archive.archived_by = request.user
            user_to_archive.save()

            # If user is an organization (user_type = 15), also archive the linked organization
            if user_to_archive.user_type == 15 and hasattr(user_to_archive, 'organization_account'):
                organization = user_to_archive.organization_account
                organization.is_archived = True
                organization.is_active = False
                organization.archived_at = timezone.now()
                organization.archived_by = request.user

                # Set organization status to inactive when archived
                organization._organization_status = 'inactive'
                organization.save()

                # Log the organization archival activity
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"{request.user.first_name} archived organization: {organization.organization_name} along with its user account"
                )

                return JsonResponse({
                    'success': True,
                    'message': f'User {user_to_archive.username} and linked organization {organization.organization_name} have been archived successfully'
                })

            # Log regular user archival
            UserActivityLog.objects.create(
                user=request.user,
                activity=f"{request.user.first_name} archived the user account for: {user_to_archive.username}"
            )

            return JsonResponse({
                'success': True,
                'message': f'User {user_to_archive.username} has been archived successfully'
            })

        except CustomUser.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'User not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e) if str(e) else 'An error occurred while archiving the user'
            }, status=500)


@require_POST
def export_users(request):
    if request.method == 'POST':
        try:
            export_option = request.POST.get('export_option', 'all')
            file_format = request.POST.get('format', 'excel')
            search_term = request.POST.get('search', '')
            unit_filter = request.POST.get('unit_filter', '')
            verified_filter = request.POST.get('verified_filter', '')
            status_filter = request.POST.get('status_filter', '')

            # Get custom filters
            verification_status = request.POST.get('verification_status', '')
            account_status = request.POST.get('account_status', '')
            unit_custom = request.POST.get('unit', '')

            # Get date range
            start_date = request.POST.get('start_date', '')
            end_date = request.POST.get('end_date', '')

            users = CustomUser.objects.all()

            # Apply filters based on export options
            if export_option == 'filtered':
                # Apply search if provided
                if search_term:
                    users = users.filter(
                        Q(first_name__icontains=search_term) |
                        Q(last_name__icontains=search_term) |
                        Q(username__icontains=search_term) |
                        Q(email__icontains=search_term)
                    )

                # Apply unit filter if provided
                if unit_filter:
                    users = users.filter(user_type=unit_filter)

                # Apply verification filter if provided
                if verified_filter:
                    if verified_filter == 'Verified':
                        users = users.filter(is_verified=True)
                    elif verified_filter == 'Unverified':
                        users = users.filter(is_verified=False)

                # Apply status filter if provided
                if status_filter:
                    if status_filter == 'Active':
                        users = users.filter(is_active=True)
                    elif status_filter == 'Inactive':
                        users = users.filter(is_active=False)

            elif export_option == 'custom':
                # Apply custom filters from the form
                if verification_status:
                    if verification_status == 'verified':
                        users = users.filter(is_verified=True)
                    elif verification_status == 'unverified':
                        users = users.filter(is_verified=False)

                if account_status:
                    if account_status == 'active':
                        users = users.filter(is_active=True)
                    elif account_status == 'inactive':
                        users = users.filter(is_active=False)

                if unit_custom:
                    users = users.filter(user_type=unit_custom)

            # Apply date range if provided
            if start_date:
                users = users.filter(date_joined__gte=start_date)
            if end_date:
                # Add one day to include the entire end date
                try:
                    end_date_obj = timezone.datetime.strptime(end_date, '%Y-%m-%d')
                    end_date_obj = end_date_obj + timezone.timedelta(days=1)
                    users = users.filter(date_joined__lte=end_date_obj)
                except ValueError:
                    # Handle invalid date format
                    pass

            # Check if Excel format is selected
            if file_format == 'excel':
                # Load the template
                template_path = os.path.join(settings.STATICFILES_DIRS[0], 'templates', 'UserTemplate.xlsx')

                # If staticfiles_dirs doesn't work, try alternative path
                if not os.path.exists(template_path):
                    template_path = os.path.join(settings.BASE_DIR, 'static', 'templates', 'UserTemplate.xlsx')

                if not os.path.exists(template_path):
                    return HttpResponse("Template file not found", status=500)

                # Load workbook
                wb = load_workbook(template_path)
                ws = wb.active  # Get the active sheet (Sheet1)

                print(f"Row 7 headers: {ws['A7'].value}, {ws['B7'].value}, {ws['C7'].value}")
                print(f"Row 8 content before: {ws['A8'].value}, {ws['B8'].value}")

                max_row = ws.max_row

                # Clear rows 8 and below (where data should go)
                for row in range(8, max_row + 1):
                    for col in range(1, 14):  # Columns A to M (1-13)
                        ws.cell(row=row, column=col, value=None)

                # Now write data starting from row 8
                start_row = 8

                # Write user data to the template
                for i, user in enumerate(users, start=start_row):
                    # Write the user data
                    ws.cell(row=i, column=1, value=user.id)
                    ws.cell(row=i, column=2, value=user.first_name or '')
                    ws.cell(row=i, column=3, value=user.last_name or '')
                    ws.cell(row=i, column=4, value=user.get_gender_display() if user.gender else '')
                    ws.cell(row=i, column=5, value=user.email or '')
                    ws.cell(row=i, column=6, value=user.phone_number or '')
                    ws.cell(row=i, column=7, value=user.username or '')
                    ws.cell(row=i, column=8, value=user.get_user_type_display() if user.user_type else '')

                    # Get position display value
                    if user.is_osas_unit:
                        position_display = user.get_position_display()
                    else:
                        position_display = user.get_position_display() if hasattr(user, 'get_position_display') else ''
                    ws.cell(row=i, column=9, value=position_display)

                    ws.cell(row=i, column=10, value='Yes' if user.is_verified else 'No')
                    ws.cell(row=i, column=11, value='Active' if user.is_active else 'Inactive')

                    # Format dates properly
                    if user.date_joined:
                        ws.cell(row=i, column=12, value=user.date_joined.strftime('%Y-%m-%d %H:%M'))
                    else:
                        ws.cell(row=i, column=12, value='')

                    if user.last_login:
                        ws.cell(row=i, column=13, value=user.last_login.strftime('%Y-%m-%d %H:%M'))
                    else:
                        ws.cell(row=i, column=13, value='')

                # Debug: Check what we wrote
                print(f"First data row (row 8): {ws['A8'].value}, {ws['B8'].value}, {ws['C8'].value}")

                # Save workbook to a BytesIO object
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                # Create response with template
                response = HttpResponse(
                    output.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="users_export.xlsx"'

                return response

            elif file_format == 'csv':
                data = []
                for user in users:
                    data.append({
                        'ID': user.id,
                        'First Name': user.first_name,
                        'Last Name': user.last_name,
                        'Username': user.username,
                        'Email': user.email,
                        'Unit': user.get_user_type_display(),
                        'Position': user.get_position_display() if user.position else '',
                        'Verified': 'Yes' if user.is_verified else 'No',
                        'Status': 'Active' if user.is_active else 'Inactive',
                        'Date Joined': user.date_joined.strftime('%Y-%m-%d %H:%M'),
                        'Last Login': user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else '',
                        'Gender': user.get_gender_display() if user.gender else '',
                        'Phone Number': user.phone_number or '',
                    })

                # Create DataFrame
                df = pd.DataFrame(data)

                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename="users_export.csv"'
                df.to_csv(response, index=False, encoding='utf-8')

                return response

        except Exception as e:
            # Log the error for debugging
            import traceback
            error_trace = traceback.format_exc()
            print(f"Export error: {str(e)}")
            print(f"Traceback: {error_trace}")
            return HttpResponse(f"Error during export: {str(e)}", status=500)

    return HttpResponse('Invalid request', status=400)


# ----------------------------------------------- Downloadables Section ------------------------------------------------
class DownloadableListView(LoginRequiredMixin, ListView):
    model = Downloadable
    template_name = 'osas/dashboard.html'
    context_object_name = 'downloadables'

    def get_queryset(self):
        return Downloadable.objects.all().order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['downloadable_form'] = DownloadableForm()
        context['selected_category'] = self.request.GET.get('category', '')
        return context

    def get(self, request, *args, **kwargs):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            downloadables = self.get_queryset()
            data = [{
                'id': d.id,
                'title': d.title,
                'description': d.description,
                'file_name': d.get_file_name(),
                'file_url': d.file.url,
                'file_size': d.get_file_size(),
                'category': d.get_category_display(),
                'created_at': d.created_at.strftime('%Y-%m-%d %H:%M'),
                'is_active': d.is_active,
                'created_by': d.created_by.get_full_name(),
            } for d in downloadables]
            return JsonResponse({'downloadables': data})
        return super().get(request, *args, **kwargs)


class DownloadableDetailView(LoginRequiredMixin, DetailView):
    model = Downloadable

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'downloadable': {
                    'id': self.object.id,
                    'title': self.object.title,
                    'description': self.object.description,
                    'file_name': self.object.get_file_name(),
                    'file_url': self.object.file.url,
                    'file_size': self.object.get_file_size(),
                    'category': self.object.get_category_display(),
                    'created_at': self.object.created_at.strftime('%Y-%m-%d %H:%M'),
                    'is_active': self.object.is_active,
                    'created_by': self.object.created_by.get_full_name(),
                }
            })
        return super().get(request, *args, **kwargs)


class DownloadableCreateView(LoginRequiredMixin, CreateView):
    model = Downloadable
    form_class = DownloadableForm
    success_url = reverse_lazy('downloadables')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            downloadable = form.save()

            # Log the activity
            UserActivityLog.objects.create(
                user=self.request.user,
                activity=f"{self.request.user.first_name} created new template titled: {downloadable.title}"
            )

            return JsonResponse({
                'success': True,
                'downloadable': {
                    'id': downloadable.id,
                    'title': downloadable.title,
                    'description': downloadable.description,
                    'file_name': downloadable.get_file_name(),
                    'file_url': downloadable.file.url,
                    'file_size': downloadable.get_file_size(),
                    'category': downloadable.get_category_display(),
                    'created_at': downloadable.created_at.strftime('%Y-%m-%d %H:%M'),
                    'is_active': downloadable.is_active,
                    'created_by': downloadable.created_by.get_full_name(),
                }
            })

        return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = {field: [str(error) for error in errors] for field, errors in form.errors.items()}
            return JsonResponse({'success': False, 'errors': errors}, status=400)
        return super().form_invalid(form)


class DownloadableUpdateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        downloadable = get_object_or_404(Downloadable, pk=kwargs['pk'])
        return JsonResponse({
            'success': True,
            'downloadable': {
                'id': downloadable.id,
                'title': downloadable.title,
                'description': downloadable.description,
                'file_name': downloadable.get_file_name(),
                'file_url': downloadable.file.url,
                'file_size': downloadable.get_file_size(),
                'category': downloadable.category,
                'created_at': downloadable.created_at.strftime('%Y-%m-%d %H:%M'),
                'is_active': downloadable.is_active,
                'created_by': downloadable.created_by.get_full_name(),
            }
        })

    def post(self, request, *args, **kwargs):
        downloadable = get_object_or_404(Downloadable, pk=kwargs['pk'])
        form = DownloadableForm(request.POST, request.FILES, instance=downloadable)

        if form.is_valid():
            downloadable = form.save()

            # Log the activity
            UserActivityLog.objects.create(
                user=self.request.user,
                activity=f"{self.request.user.first_name} updated the template titled: {downloadable.title}"
            )

            return JsonResponse({
                'success': True,
                'downloadable': {
                    'id': downloadable.id,
                    'title': downloadable.title,
                    'description': downloadable.description,
                    'file_name': downloadable.get_file_name(),
                    'file_url': downloadable.file.url,
                    'file_size': downloadable.get_file_size(),
                    'category': downloadable.get_category_display(),
                    'created_at': downloadable.created_at.strftime('%Y-%m-%d %H:%M'),
                    'is_active': downloadable.is_active,
                    'created_by': downloadable.created_by.get_full_name(),
                }
            })

        errors = {field: [str(error) for error in errors] for field, errors in form.errors.items()}
        return JsonResponse({'success': False, 'errors': errors}, status=400)


class DownloadableArchiveView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        downloadable = self.get_object()
        return self.request.user.is_superuser or downloadable.created_by == self.request.user

    def get_object(self):
        return Downloadable.objects.get(pk=self.kwargs['pk'])

    def post(self, request, *args, **kwargs):
        downloadable = self.get_object()
        downloadable.is_archived = True
        downloadable.archived_at = timezone.now()
        downloadable.archived_by = request.user
        downloadable.save()

        # Log the activity
        UserActivityLog.objects.create(
            user=self.request.user,
            activity=f"{self.request.user.first_name} archived the template titled: {downloadable.title}"
        )

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        return redirect('downloadables')


def download_downloadable(request, pk):
    downloadable = get_object_or_404(Downloadable, pk=pk)
    file_path = downloadable.file.path

    if os.path.exists(file_path):
        response = FileResponse(open(file_path, 'rb'))
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        return response
    raise Http404("File not found")


# ---------------------------------------------- Announcement Section --------------------------------------------------
class AnnouncementListView(LoginRequiredMixin, ListView):
    model = Announcement
    template_name = 'osas/dashboard.html'
    context_object_name = 'announcements'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = AnnouncementForm()
        context['COURSE_CHOICES'] = Announcement.COURSE_CHOICES
        return context

    def get(self, request, *args, **kwargs):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            announcements = self.get_queryset()
            data = {
                'announcements': [{
                    'id': a.id,
                    'title': a.title,
                    'category_display': a.get_category_display(),
                    'author_name': a.author.get_full_name(),
                    'created_at': a.created_at.strftime('%Y-%m-%d %H:%M'),
                    'is_published': a.is_published,
                    'permissions': {
                        'view': request.user.has_perm('announcements.view_announcement', a),
                        'change': request.user.has_perm('announcements.change_announcement', a),
                        'delete': request.user.has_perm('announcements.delete_announcement', a),
                    }
                } for a in announcements]
            }
            return JsonResponse(data)
        return super().get(request, *args, **kwargs)


class AnnouncementCreateView(LoginRequiredMixin, CreateView):
    model = Announcement
    form_class = AnnouncementForm
    template_name = 'osas/dashboard.html'
    success_url = reverse_lazy('announcements')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['scholarships'] = Scholarship.objects.filter(is_active=True)
        return context

    def form_valid(self, form):
        form.instance.author = self.request.user
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                announcement = form.save()

                # Log the activity
                UserActivityLog.objects.create(
                    user=self.request.user,
                    activity=f"{self.request.user.first_name} created new announcement titled: {announcement.title}"
                )

                # Handle images
                images = self.request.FILES.getlist('images')
                for image in images:
                    AnnouncementImage.objects.create(announcement=announcement, image=image)

                return JsonResponse({
                    'success': True,
                    'announcement': {
                        'id': announcement.id,
                        'title': announcement.title,
                        'category': announcement.get_category_display(),
                        'author': announcement.author.get_full_name(),
                        'created_at': announcement.created_at.strftime('%Y-%m-%d %H:%M'),
                        'is_published': announcement.is_published,
                    }
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': str(e)
                }, status=400)
        return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Convert form errors to a format the frontend can handle
            errors = {}
            for field, error_list in form.errors.items():
                # Remove the "id_" prefix if present for consistency
                field_name = field.replace('id_', '')
                errors[field_name] = error_list

            return JsonResponse({
                'success': False,
                'errors': errors,
                'message': 'Please correct the errors below.'
            }, status=400)

        courses = self.request.POST.getlist('courses')
        form.instance.courses = courses
        return super().form_invalid(form)


class AnnouncementDetailView(LoginRequiredMixin, DetailView):
    model = Announcement

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            announcement = self.object
            images = announcement.images.all()

            return JsonResponse({
                'success': True,
                'announcement': {
                    'id': announcement.id,
                    'title': announcement.title,
                    'content': announcement.content,
                    'category': announcement.get_category_display(),
                    'category_value': announcement.category,
                    'author_name': announcement.author.get_full_name(),
                    'created_at': announcement.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'is_published': announcement.is_published,
                    'link': announcement.link or '',

                    # Category-specific fields
                    'courses': list(announcement.courses) if announcement.courses else [],
                    'enrollment_start': announcement.enrollment_start.strftime(
                        '%Y-%m-%d %H:%M:%S') if announcement.enrollment_start else None,
                    'enrollment_end': announcement.enrollment_end.strftime(
                        '%Y-%m-%d %H:%M:%S') if announcement.enrollment_end else None,
                    'event_date': announcement.event_date.strftime(
                        '%Y-%m-%d %H:%M:%S') if announcement.event_date else None,
                    'location': announcement.location or '',
                    'suspension_date': announcement.suspension_date.strftime(
                        '%Y-%m-%d') if announcement.suspension_date else None,
                    'until_suspension_date': announcement.until_suspension_date.strftime(
                        '%Y-%m-%d') if announcement.until_suspension_date else None,
                    'contact_info': announcement.contact_info or '',

                    # Scholarship fields
                    'scholarship': {
                        'id': announcement.scholarship.id if announcement.scholarship else None,
                        'name': announcement.scholarship.name if announcement.scholarship else None
                    },
                    'application_start': announcement.application_start.strftime(
                        '%Y-%m-%d %H:%M:%S') if announcement.application_start else None,
                    'application_end': announcement.application_end.strftime(
                        '%Y-%m-%d %H:%M:%S') if announcement.application_end else None,
                    'requirements': announcement.requirements or '',
                    'benefits': announcement.benefits or '',

                    # Images
                    'images': [{
                        'url': image.image.url,
                        'caption': image.caption or ''
                    } for image in images]
                }
            })
        return super().get(request, *args, **kwargs)


class AnnouncementUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Announcement
    form_class = AnnouncementForm
    template_name = 'osas/dashboard.html'

    def test_func(self):
        obj = self.get_object()
        return self.request.user == obj.author or self.request.user.is_superuser

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.GET.get('get_scholarship'):
            # Handle scholarship details request
            scholarship_id = request.GET.get('scholarship_id')
            try:
                scholarship = Scholarship.objects.get(id=scholarship_id)
                return JsonResponse({
                    'success': True,
                    'requirements': scholarship.requirements,
                    'benefits': scholarship.benefits
                })
            except Scholarship.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Scholarship not found'}, status=404)

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            announcement = self.object
            images = announcement.images.all()

            return JsonResponse({
                'success': True,
                'announcement': {
                    'id': announcement.id,
                    'title': announcement.title,
                    'content': announcement.content,
                    'category': announcement.category,
                    'is_published': announcement.is_published,
                    'link': announcement.link or '',

                    # Category-specific fields
                    'courses': list(announcement.courses) if announcement.courses else [],
                    'enrollment_start': announcement.enrollment_start.strftime(
                        '%Y-%m-%d %H:%M:%S') if announcement.enrollment_start else None,
                    'enrollment_end': announcement.enrollment_end.strftime(
                        '%Y-%m-%d %H:%M:%S') if announcement.enrollment_end else None,
                    'event_date': announcement.event_date.strftime(
                        '%Y-%m-%d %H:%M:%S') if announcement.event_date else None,
                    'location': announcement.location or '',
                    'suspension_date': announcement.suspension_date.strftime(
                        '%Y-%m-%d') if announcement.suspension_date else None,
                    'until_suspension_date': announcement.until_suspension_date.strftime(
                        '%Y-%m-%d') if announcement.until_suspension_date else None,
                    'contact_info': announcement.contact_info or '',

                    # Scholarship fields
                    'scholarship': announcement.scholarship.id if announcement.scholarship else None,
                    'application_start': announcement.application_start.strftime(
                        '%Y-%m-%d %H:%M:%S') if announcement.application_start else None,
                    'application_end': announcement.application_end.strftime(
                        '%Y-%m-%d %H:%M:%S') if announcement.application_end else None,
                    'requirements': announcement.requirements or '',
                    'benefits': announcement.benefits or '',

                    # Images
                    'images': [{
                        'id': image.id,
                        'url': image.image.url,
                        'filename': image.image.name.split('/')[-1]
                    } for image in images]
                }
            })
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        announcement = form.save(commit=False)

        remove_image_ids = self.request.POST.getlist('remove_images')
        if remove_image_ids:
            AnnouncementImage.objects.filter(
                id__in=remove_image_ids,
                announcement=announcement
            ).delete()

        images = self.request.FILES.getlist('images')
        for image in images:
            AnnouncementImage.objects.create(announcement=announcement, image=image)

        announcement.save()

        # Log the activity
        UserActivityLog.objects.create(
            user=self.request.user,
            activity=f"{self.request.user.first_name} updated the announcement titled: {announcement.title}"
        )

        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'announcement': {
                    'id': announcement.id,
                    'title': announcement.title,
                    'category': announcement.get_category_display(),
                    'author': announcement.author.get_full_name(),
                    'created_at': announcement.created_at.strftime('%Y-%m-%d %H:%M'),
                    'is_published': announcement.is_published,
                }
            })
        return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = {field: [str(error) for error in error_list] for field, error_list in form.errors.items()}
            return JsonResponse({
                'success': False,
                'errors': errors,
                'message': 'Please correct the errors below.'
            }, status=400)
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('announcements')


class AnnouncementArchiveView(LoginRequiredMixin, UserPassesTestMixin, View):
    def get_object(self):
        return get_object_or_404(Announcement, pk=self.kwargs.get('pk'))

    def test_func(self):
        obj = self.get_object()
        return self.request.user == obj.author or self.request.user.is_superuser

    def post(self, request, *args, **kwargs):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                announcement = self.get_object()
                announcement.is_archived = True
                announcement.archived_at = timezone.now()
                announcement.archived_by = request.user
                announcement.save()

                # Log the activity
                UserActivityLog.objects.create(
                    user=self.request.user,
                    activity=f"{self.request.user.first_name} archived the announcement titled: {announcement.title}"
                )

                return JsonResponse({
                    'success': True,
                    'message': 'Announcement archived successfully'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': str(e)
                }, status=400)
        return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)

    def handle_no_permission(self):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'You do not have permission to archive announcements'
            }, status=403)
        return super().handle_no_permission()


# ------------------------------------------------ Complain Section ----------------------------------------------------
class ComplaintCreateView(LoginRequiredMixin, CreateView):
    model = Complaint
    form_class = ComplaintForm
    template_name = 'core/complaint.html'
    success_url = reverse_lazy('home')

    def get_initial(self):
        initial = super().get_initial()
        if self.request.user.is_authenticated:
            initial.update({
                'complainant_first_name': self.request.user.first_name or '',
                'complainant_last_name': self.request.user.last_name or '',
                'complainant_email': self.request.user.email or '',
                'complainant_phone': self.request.user.phone_number or '',
                'complainant_address': self.request.user.address or '',
            })
        return initial

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Make the fields readonly
        for field in ['complainant_first_name', 'complainant_last_name',
                      'complainant_email', 'complainant_phone', 'complainant_address']:
            form.fields[field].widget.attrs['readonly'] = True
            form.fields[field].widget.attrs['class'] = form.fields[field].widget.attrs.get('class',
                                                                                           '') + ' readonly-field'

        # Ensure the address field is properly rendered as a textarea
        form.fields['complainant_address'].widget = forms.Textarea(attrs={
            'readonly': True,
            'class': 'form-control readonly-field',
            'rows': 3
        })
        return form

    def form_valid(self, form):
        # Clean the readonly fields before saving
        form.instance.complainant_first_name = form.cleaned_data['complainant_first_name'].strip()
        form.instance.complainant_last_name = form.cleaned_data['complainant_last_name'].strip()

        form.instance.status = 'under_review'
        form.instance.created_by = self.request.user
        response = super().form_valid(form)

        # Handle file uploads
        self.handle_file_uploads()

        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'complaint_id': self.object.reference_number,
                'message': 'Complaint submitted successfully!',
                'redirect_url': self.get_success_url()
            })
        else:
            messages.success(self.request, 'Complaint submitted successfully!')
            return response

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'errors': form.errors.get_json_data(),
                'message': 'Please correct the errors below.'
            }, status=400)
        return super().form_invalid(form)

    def handle_file_uploads(self):
        # Handle document uploads
        documents = self.request.FILES.getlist('documents')
        for doc in documents:
            ComplaintDocument.objects.create(
                complaint=self.object,
                file=doc,
                description=self.request.POST.get('document_description', '')
            )

        # Handle image uploads
        images = self.request.FILES.getlist('images')
        for img in images:
            ComplaintImage.objects.create(
                complaint=self.object,
                image=img,
                caption=self.request.POST.get('image_caption', '')
            )

    # Retrieving Footer Content
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        footer_content = FooterContent.objects.first()
        if not footer_content:
            footer_content = FooterContent.objects.create()
        context['footer_content'] = footer_content
        return context


def download_complaint_pdf(request, complaint_id):
    try:
        complaint = Complaint.objects.get(reference_number=complaint_id)
    except Complaint.DoesNotExist:
        return HttpResponse("Complaint not found", status=404)

    # Create a file-like buffer to receive PDF data
    buffer = BytesIO()

    # Create the PDF object with improved settings
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Designs
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Title'],
        fontSize=14,
        leading=18,
        spaceAfter=12,
        textColor=colors.HexColor('#000000'),
        fontName='Helvetica-Bold',
        alignment=1
    )
    heading_style = ParagraphStyle(
        'Heading2',
        parent=styles['Heading2'],
        fontSize=12,
        leading=16,
        spaceAfter=8,
        textColor=colors.HexColor('#000000'),
        fontName='Helvetica-Bold'
    )
    normal_style = ParagraphStyle(
        'BodyText',
        parent=styles['BodyText'],
        fontSize=10,
        leading=12,
        textColor=colors.HexColor('#000000'),
        fontName='Helvetica'
    )
    label_style = ParagraphStyle(
        'Label',
        parent=styles['BodyText'],
        fontSize=10,
        leading=12,
        textColor=colors.HexColor('#000000'),
        fontName='Helvetica-Bold'
    )

    # HEADER SECTION
    p.setFillColor(colors.white)
    p.rect(0, height - 100, width, 100, fill=True, stroke=False)
    logo_x, logo_y = 40, height - 70
    p.setFillColor(colors.white)
    p.rect(logo_x, logo_y, 50, 50, fill=True, stroke=False)

    try:
        p.drawImage('static/images/cvsu-logo.png', logo_x, logo_y,
                    width=50, height=50, preserveAspectRatio=True, mask='auto')
    except:
        p.setFillColor(colors.HexColor('#006847'))
        p.rect(logo_x, logo_y, 50, 50, fill=True, stroke=False)
        p.setFillColor(colors.white)
        p.setFont("Helvetica-Bold", 8)
        p.drawString(logo_x + 5, logo_y + 20, "CvSU")

    p.setFillColor(colors.black)
    p.setFont("Helvetica-Bold", 12)
    p.drawCentredString(width / 2, height - 40, "Office of the Student Affairs and Services")
    p.setFont("Helvetica", 10)
    p.drawCentredString(width / 2, height - 55, "Bacoor City Campus")

    # Document title
    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width / 2, height - 80, "COMPLAINT REPORT")

    # Reference and date
    p.setFont("Helvetica", 9)
    p.drawString(50, height - 95, f"Reference No.: {complaint.reference_number}")
    p.drawRightString(width - 50, height - 95, f"Date: {datetime.now().strftime('%Y-%m-%d')}")

    # Horizontal line separator
    p.line(50, height - 100, width - 50, height - 100)

    # Main Content
    y_position = height - 120

    # Section 1: Basic Information
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y_position, "1. BASIC INFORMATION")
    y_position -= 20

    # Basic information table
    basic_data = [
        [" "],
        ["Status:", complaint.get_status_display()],
        ["Title:", complaint.title],
        ["Date Submitted:", complaint.created_at.strftime('%Y-%m-%d %H:%M')],
        ["Incident Date:", complaint.incident_date.strftime('%Y-%m-%d')],
        ["Incident Time:", complaint.incident_time.strftime('%H:%M') if complaint.incident_time else "N/A"],
        ["Incident Location:", complaint.incident_location]
    ]

    t = Table(basic_data, colWidths=[100, width - 100])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
    ]))

    t.wrapOn(p, width, height)
    t.drawOn(p, 50, y_position - (len(basic_data) * 15))
    y_position -= (len(basic_data) * 15 + 30)

    # Section 2: Parties Involved
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y_position, "2. PARTIES INVOLVED")
    y_position -= 20

    # Two column layout for parties
    col_width = (width - 120) / 2

    # Complainant Information
    p.setFont("Helvetica-Bold", 11)
    p.drawString(50, y_position, "COMPLAINANT")
    y_position -= 15

    complainant_data = [
        ["Name:", f"{complaint.complainant_first_name} {complaint.complainant_last_name}"],
        ["Email:", complaint.complainant_email],
        ["Phone:", complaint.complainant_phone],
        ["Address:", complaint.complainant_address],
        ["Instructor:", complaint.complainant_instructor_name or "N/A"]
    ]

    t = Table(complainant_data, colWidths=[60, col_width - 70])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))

    t.wrapOn(p, width, height)
    t.drawOn(p, 50, y_position - (len(complainant_data) * 15))

    # Respondent Information
    p.setFont("Helvetica-Bold", 11)
    p.drawString(50 + col_width + 20, y_position + 15, "RESPONDENT")

    respondent_data = [
        ["Name:", f"{complaint.respondent_first_name} {complaint.respondent_last_name}"],
        ["Type:", complaint.get_respondent_type_display()],
    ]

    if complaint.respondent_type == 'student':
        respondent_data.extend([
            ["Course:", complaint.respondent_course.name if complaint.respondent_course else "N/A"],
            ["Year:", complaint.respondent_year],
            ["Section:", complaint.respondent_section]
        ])
    else:
        respondent_data.append(["Department:", complaint.respondent_department])

    t = Table(respondent_data, colWidths=[60, col_width - 70])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))

    t.wrapOn(p, width, height)
    t.drawOn(p, 50 + col_width + 20, y_position - (len(respondent_data) * 15))

    y_position -= max((len(complainant_data) * 15), (len(respondent_data) * 15)) + 40

    # Section 3: Complaint Details
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y_position, "3. COMPLAINT DETAILS")
    y_position -= 20

    # Statement and Witnesses
    col_width = (width - 120) / 2

    # Left column: Statement
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y_position, "STATEMENT:")
    statement_y = y_position - 15

    clean_statement = complaint.statement.replace("\n", "<br/>")
    statement_text = f'<para leading=12><font name="Helvetica" size=10>{clean_statement}</font></para>'
    statement = Paragraph(statement_text, style=normal_style)
    statement.wrap(col_width, height)
    statement.drawOn(p, 50, statement_y - statement.height)

    # Right column: Witnesses (if any)
    if complaint.witnesses:
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50 + col_width + 20, y_position, "WITNESSES:")
        witnesses_y = y_position - 15

        witnesses_clean = complaint.witnesses.replace("\n", "<br/>")
        witnesses_text = f'<para leading=12><font name="Helvetica" size=10>{witnesses_clean}</font></para>'
        witnesses = Paragraph(witnesses_text, style=normal_style)
        witnesses.wrap(col_width, height)
        witnesses.drawOn(p, 50 + col_width + 20, witnesses_y - witnesses.height)

        y_position -= max(statement.height + 15, witnesses.height + 15) + 30
    else:
        # No witnesses
        y_position -= statement.height + 30

    # Section 4: Supporting Evidence
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y_position, "4. SUPPORTING EVIDENCE")
    y_position -= 20

    # Get counts for each evidence type
    doc_count = complaint.documents.count()
    img_count = complaint.images.count()

    evidence_data = [
        ["Documents:", f"{'Yes' if doc_count > 0 else 'No'} ({doc_count})"],
        ["Images:", f"{'Yes' if img_count > 0 else 'No'} ({img_count})"]
    ]

    t = Table(evidence_data, colWidths=[100, width - 150])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
    ]))

    t.wrapOn(p, width, height)
    t.drawOn(p, 50, y_position - (len(evidence_data) * 15))
    y_position -= (len(evidence_data) * 15 + 30)

    # Display actual images if they exist
    if img_count > 0:
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, y_position, "ATTACHED IMAGES:")
        y_position -= 20

        # Check if we need a new page
        if y_position < 200:
            p.showPage()
            y_position = height - 50
            p.setFont("Helvetica-Bold", 10)
            p.drawString(50, y_position, "ATTACHED IMAGES (continued):")
            y_position -= 20

        for image in complaint.images.all():
            try:
                # Check if we need a new page before adding image
                if y_position < 300:
                    p.showPage()
                    y_position = height - 50
                    p.setFont("Helvetica-Bold", 10)
                    p.drawString(50, y_position, "ATTACHED IMAGES (continued):")
                    y_position -= 20

                # Draw image caption if exists
                if image.caption:
                    p.setFont("Helvetica", 9)
                    p.drawString(50, y_position, f"Caption: {image.caption}")
                    y_position -= 15

                # Draw the actual image (scaled to fit)
                img_path = image.image.path
                img = ImageReader(img_path)

                # Get image dimensions and scale to fit page width
                img_width, img_height = img.getSize()
                aspect_ratio = img_width / img_height
                max_width = width - 100
                display_width = min(img_width, max_width)
                display_height = display_width / aspect_ratio

                # Check if image will fit on current page
                if y_position - display_height < 50:  # Too close to bottom
                    p.showPage()
                    y_position = height - 50

                p.drawImage(img_path, 50, y_position - display_height,
                            width=display_width, height=display_height,
                            preserveAspectRatio=True, mask='auto')

                y_position -= (display_height + 30)

            except Exception as e:
                print(f"Error processing image: {e}")
                continue

    # Admin Notes
    if complaint.notes and request.user.is_superuser:
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, y_position, "ADMIN NOTES:")
        y_position -= 15

        notes_clean = complaint.notes.replace("\n", "<br/>")
        notes_text = f'<para leading=12><font name="Helvetica" size=10>{notes_clean}</font></para>'
        notes = Paragraph(notes_text, style=normal_style)
        notes.wrap(width - 100, height)
        notes.drawOn(p, 50, y_position - notes.height)
        y_position -= notes.height + 30

    # FOOTER SECTION
    footer_y = 40

    p.setFillColor(colors.white)
    p.rect(0, 0, width, footer_y + 20, fill=True, stroke=False)

    # Footer line
    p.setStrokeColor(colors.black)
    p.line(50, footer_y + 10, width - 50, footer_y + 10)

    # Footer text
    p.setFillColor(colors.black)
    p.setFont("Helvetica", 8)
    p.drawCentredString(width / 2, footer_y,
                        "This is an official document of Cavite State University - Bacoor City Campus")
    p.drawCentredString(width / 2, footer_y - 10, "Office of the Student Affairs and Services")

    # Close the PDF object cleanly
    p.showPage()
    p.save()

    # File response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Complaint_Report_{complaint.reference_number}.pdf"'
    return response


class UpdateComplaintStatusView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'osas.change_complaint'

    @method_decorator(csrf_exempt)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, pk):
        try:
            # Parse JSON data
            data = json.loads(request.body)
            status = data.get('status')
            notes = data.get('notes', '')
            hearing_data = data.get('hearing_data')
            clear_existing_hearings = data.get('clear_existing_hearings', False)

            print(f"Received data: {data}")  # Debug log

            if not status:
                return JsonResponse({
                    'success': False,
                    'error': 'Status is required'
                }, status=400)

            complaint = Complaint.objects.get(pk=pk)

            # Store old status for comparison
            old_status = complaint.status
            old_status_display = complaint.get_status_display()

            # Update complaint status
            complaint.status = status

            # Handle "under_review" status - clear next hearing date
            if status == 'under_review':
                # Delete any upcoming hearings for this complaint
                upcoming_hearings = complaint.hearings.filter(
                    schedule_date__gte=timezone.now().date()
                )
                upcoming_hearings.delete()
                print(f"Cleared upcoming hearings for complaint {complaint.reference_number}")

            # Update notes if provided
            if notes:
                if complaint.notes:
                    complaint.notes += f"\n\n[{datetime.now().strftime('%Y-%m-%d %H:%M')}] {notes}"
                else:
                    complaint.notes = f"[{datetime.now().strftime('%Y-%m-%d %H:%M')}] {notes}"

            complaint.updated_at = datetime.now()

            # Handle hearing creation if applicable (for hearing statuses)
            hearing = None
            if hearing_data and status in ['1st_hearing', '2nd_hearing', 'other_hearing', 'ongoing_hearing']:
                # Convert string to date/time objects
                schedule_date_str = hearing_data.get('schedule_date')
                schedule_time_str = hearing_data.get('schedule_time')

                if schedule_date_str and schedule_time_str:
                    # IMPORTANT: Clear existing hearings before creating new one
                    # This ensures only one hearing exists at a time
                    complaint.hearings.all().delete()
                    print(f"Cleared all existing hearings for complaint {complaint.reference_number}")

                    # Convert to proper Python date/time objects
                    try:
                        schedule_date = datetime.strptime(schedule_date_str, '%Y-%m-%d').date()
                        schedule_time = datetime.strptime(schedule_time_str, '%H:%M').time()

                        # Create new hearing
                        hearing = Hearing.objects.create(
                            complaint=complaint,
                            schedule_date=schedule_date,
                            schedule_time=schedule_time,
                            location=hearing_data.get('location'),
                            description=hearing_data.get('description', ''),
                            created_by=request.user
                        )
                        print(f"New hearing created: {hearing}")
                    except ValueError as e:
                        print(f"Error parsing date/time: {e}")
                        return JsonResponse({
                            'success': False,
                            'error': f'Invalid date/time format: {str(e)}'
                        }, status=400)
                else:
                    return JsonResponse({
                        'success': False,
                        'error': 'Hearing date and time are required for hearing statuses'
                    }, status=400)

            # Save complaint after all updates
            complaint.save()

            # Send email notification
            try:
                self.send_email_notification(complaint, old_status, old_status_display, hearing, notes)
            except Exception as email_error:
                print(f"Email error (non-critical): {email_error}")
                traceback.print_exc()

            return JsonResponse({
                'success': True,
                'message': f'Complaint status updated to {complaint.get_status_display()} successfully'
            })

        except Complaint.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Complaint not found'
            }, status=404)
        except Exception as e:
            print(f"Error in UpdateComplaintStatusView: {str(e)}")
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

    def send_email_notification(self, complaint, old_status, old_status_display, hearing=None, admin_notes=None):
        try:
            # Prepare context for email template
            context = {
                'complainant_name': f"{complaint.complainant_first_name} {complaint.complainant_last_name}",
                'reference_number': complaint.reference_number,
                'complaint_title': complaint.title,
                'new_status': complaint.status,
                'status_display': complaint.get_status_display(),
                'old_status': old_status,
                'old_status_display': old_status_display,
                'admin_notes': admin_notes,
                'hearing_details': None,
            }

            # Add hearing details if applicable
            if hearing:
                # Format date and time
                context['hearing_details'] = {
                    'hearing_type': complaint.get_hearing_type_display(),
                    'date': hearing.schedule_date.strftime('%B %d, %Y'),
                    'time': hearing.schedule_time.strftime('%I:%M %p'),
                    'location': hearing.location,
                    'description': hearing.description,
                }

            print(f"Email context prepared for {complaint.complainant_email}")

            # Render email template
            html_message = render_to_string('emails/complaint_status_update.html', context)
            plain_message = strip_tags(html_message)

            # Email subject
            subject = f"Status Update: Complaint #{complaint.reference_number} - {complaint.get_status_display()}"

            # Send email
            send_mail(
                subject=subject,
                message=plain_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[complaint.complainant_email],
                html_message=html_message,
                fail_silently=False,
            )

            print(f"Email sent successfully to {complaint.complainant_email}")

        except Exception as e:
            print(f"Failed to send email notification: {e}")
            traceback.print_exc()
            

class ComplaintDetailView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'osas.view_complaint'

    def get(self, request, *args, **kwargs):
        try:
            complaint = get_object_or_404(Complaint, pk=kwargs['pk'])
            documents = complaint.documents.all()
            images = complaint.images.all()

            data = {
                'success': True,
                'complaint': {
                    'id': complaint.id,
                    'reference_number': complaint.reference_number,
                    'title': complaint.title,
                    'status': complaint.status,
                    'status_display': complaint.get_status_display(),
                    'created_at': complaint.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': complaint.updated_at.strftime('%Y-%m-%d %H:%M:%S') if complaint.updated_at else None,

                    # Complainant information
                    'complainant_first_name': complaint.complainant_first_name,
                    'complainant_last_name': complaint.complainant_last_name,
                    'complainant_email': complaint.complainant_email,
                    'complainant_phone': complaint.complainant_phone,
                    'complainant_address': complaint.complainant_address,
                    'complainant_instructor_name': complaint.complainant_instructor_name or '',

                    # Respondent information
                    'respondent_first_name': complaint.respondent_first_name,
                    'respondent_last_name': complaint.respondent_last_name,
                    'respondent_type': complaint.respondent_type,
                    'respondent_type_display': complaint.get_respondent_type_display(),
                    'respondent_course': complaint.respondent_course.name if complaint.respondent_course else '',
                    'respondent_year': complaint.respondent_year or '',
                    'respondent_section': complaint.respondent_section or '',
                    'respondent_department': complaint.respondent_department or '',

                    # Complaint details
                    'incident_date': complaint.incident_date.strftime('%Y-%m-%d') if complaint.incident_date else None,
                    'incident_time': complaint.incident_time.strftime('%H:%M:%S') if complaint.incident_time else None,
                    'incident_location': complaint.incident_location,
                    'statement': complaint.statement,
                    'witnesses': complaint.witnesses or '',
                    'notes': complaint.notes or '',

                    # Supporting files
                    'documents': [{
                        'id': doc.id,
                        'file_name': doc.file.name.split('/')[-1],
                        'file_url': doc.file.url,
                        'description': doc.description or ''
                    } for doc in documents],

                    'images': [{
                        'id': img.id,
                        'file_name': img.image.name.split('/')[-1],
                        'file_url': img.image.url,
                        'caption': img.caption or ''
                    } for img in images],
                }
            }
            return JsonResponse(data)

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


class ComplaintEditView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'osas.change_complaint'

    def get(self, request, pk):
        try:
            complaint = get_object_or_404(Complaint, pk=pk)
            documents = complaint.documents.all()
            images = complaint.images.all()

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # AJAX request - return JSON
                data = {
                    'success': True,
                    'complaint': {
                        'id': complaint.id,
                        'reference_number': complaint.reference_number,
                        'title': complaint.title,
                        'status': complaint.status,
                        'complainant_first_name': complaint.complainant_first_name,
                        'complainant_last_name': complaint.complainant_last_name,
                        'complainant_email': complaint.complainant_email,
                        'complainant_phone': complaint.complainant_phone,
                        'complainant_address': complaint.complainant_address,
                        'respondent_type': complaint.respondent_type,
                        'respondent_first_name': complaint.respondent_first_name,
                        'respondent_last_name': complaint.respondent_last_name,
                        'respondent_course': complaint.respondent_course.id if complaint.respondent_course else None,
                        'respondent_year': complaint.respondent_year or '',
                        'respondent_section': complaint.respondent_section or '',
                        'respondent_department': complaint.respondent_department or '',
                        'incident_date': complaint.incident_date.strftime(
                            '%Y-%m-%d') if complaint.incident_date else '',
                        'incident_time': complaint.incident_time.strftime('%H:%M') if complaint.incident_time else '',
                        'incident_location': complaint.incident_location,
                        'statement': complaint.statement,
                        'witnesses': complaint.witnesses or '',
                        'notes': complaint.notes or '',
                        'documents': [
                            {'id': doc.id, 'name': doc.file.name.split('/')[-1], 'url': doc.file.url}
                            for doc in documents
                        ],
                        'images': [
                            {'id': img.id, 'name': img.image.name.split('/')[-1], 'url': img.image.url}
                            for img in images
                        ],
                    },
                    'courses': list(Course.objects.values('id', 'name'))  # Add courses to response
                }
                return JsonResponse(data)
            else:
                # Regular request - render template
                return render(request, 'complaint_edit.html', {
                    'complaint': complaint,
                    'courses': Course.objects.all(),
                    'documents': documents,
                    'images': images
                })

        except Exception as e:
            import traceback
            traceback.print_exc()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)}, status=500)
            raise e

    def post(self, request, pk):
        try:
            complaint = get_object_or_404(Complaint, pk=pk)

            old_status = complaint.status
            new_status = request.POST.get('status', complaint.status)

            # Handle form data
            complaint.status = new_status
            complaint.title = request.POST.get('title', complaint.title)

            # Complainant info
            complaint.complainant_first_name = request.POST.get('complainant_first_name', '')
            complaint.complainant_last_name = request.POST.get('complainant_last_name', '')
            complaint.complainant_email = request.POST.get('complainant_email', '')
            complaint.complainant_phone = request.POST.get('complainant_phone', '')
            complaint.complainant_address = request.POST.get('complainant_address', '')

            # Respondent info
            respondent_type = request.POST.get('respondent_type', 'student')
            complaint.respondent_type = respondent_type
            complaint.respondent_first_name = request.POST.get('respondent_first_name', '')
            complaint.respondent_last_name = request.POST.get('respondent_last_name', '')

            if respondent_type == 'student':
                course_id = request.POST.get('respondent_course')
                complaint.respondent_course = Course.objects.get(id=course_id) if course_id else None
                complaint.respondent_year = request.POST.get('respondent_year', '')
                complaint.respondent_section = request.POST.get('respondent_section', '')
                complaint.respondent_department = ''
            else:
                complaint.respondent_department = request.POST.get('respondent_department', '')
                complaint.respondent_course = None
                complaint.respondent_year = ''
                complaint.respondent_section = ''

            # Complaint details
            incident_date = request.POST.get('incident_date')
            complaint.incident_date = datetime.strptime(incident_date, '%Y-%m-%d').date() if incident_date else None

            incident_time = request.POST.get('incident_time')
            if incident_time and incident_time.strip():
                complaint.incident_time = datetime.strptime(incident_time, '%H:%M').time()
            else:
                complaint.incident_time = None

            complaint.incident_location = request.POST.get('incident_location', '')
            complaint.statement = request.POST.get('statement', '')
            complaint.witnesses = request.POST.get('witnesses', '')
            complaint.notes = request.POST.get('notes', '')

            if old_status != 'under_review' and new_status == 'under_review':
                upcoming_hearings = complaint.hearings.filter(
                    schedule_date__gte=timezone.now().date()
                )
                hearing_count = upcoming_hearings.count()
                upcoming_hearings.delete()
                print(
                    f"Cleared {hearing_count} upcoming hearings for complaint {complaint.reference_number} when changing to under_review")

            if old_status in ['1st_hearing', '2nd_hearing', 'other_hearing', 'ongoing_hearing'] and \
                    new_status in ['resolved', 'canceled']:
                # Delete any upcoming hearings for this complaint
                upcoming_hearings = complaint.hearings.filter(
                    schedule_date__gte=timezone.now().date()
                )
                hearing_count = upcoming_hearings.count()
                upcoming_hearings.delete()
                print(
                    f"Cleared {hearing_count} upcoming hearings for complaint {complaint.reference_number} when changing to {new_status}")

            complaint.save()

            # Handle file uploads
            for file in request.FILES.getlist('documents'):
                ComplaintDocument.objects.create(complaint=complaint, file=file)

            for file in request.FILES.getlist('images'):
                ComplaintImage.objects.create(complaint=complaint, image=file)

            return JsonResponse({'success': True})

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


class RemoveEvidenceView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'osas.change_complaint'

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            evidence_id = data.get('evidence_id')
            evidence_type = data.get('evidence_type')

            if evidence_type == 'document':
                evidence = get_object_or_404(ComplaintDocument, id=evidence_id)
            elif evidence_type == 'image':
                evidence = get_object_or_404(ComplaintImage, id=evidence_id)
            else:
                return JsonResponse({'success': False, 'error': 'Invalid evidence type'}, status=400)

            evidence.delete()
            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


class ArchiveComplaintView(LoginRequiredMixin, View):
    def post(self, request, pk):
        try:
            complaint = Complaint.objects.get(pk=pk)

            if complaint.is_archived:
                return JsonResponse({
                    'success': False,
                    'error': 'Complaint is already archived'
                }, status=400)

            complaint.is_archived = True
            complaint.status = 'canceled'
            complaint.archived_at = timezone.now()
            complaint.archived_by = request.user

            # Add optional notes
            notes = request.POST.get('notes', '')
            if notes:
                if complaint.notes:
                    complaint.notes += f"\n\nArchival Notes ({timezone.now().strftime('%Y-%m-%d %H:%M')}):\n{notes}"
                else:
                    complaint.notes = f"Archival Notes ({timezone.now().strftime('%Y-%m-%d %H:%M')}):\n{notes}"

            complaint.save()

            return JsonResponse({
                'success': True,
                'message': 'Complaint archived and status set to Canceled'
            })

        except Complaint.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Complaint not found'
            }, status=404)

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


def export_complaints(request):
    if request.method == 'POST':
        try:
            export_option = request.POST.get('export_option', 'all_statuses')
            file_format = request.POST.get('format', 'excel')
            source_section = request.POST.get('source_section', 'under_review')

            # Start with all complaints
            complaints = Complaint.objects.all().select_related(
                'respondent_course',
                'created_by',
                'archived_by'
            )

            # Apply filters based on export options
            if export_option == 'all_statuses':
                # No additional filter needed for status
                pass
            elif export_option == 'by_status':
                status = request.POST.get('status', 'all')
                if status != 'all':
                    complaints = complaints.filter(status=status)
            elif export_option == 'by_respondent':
                respondent_type = request.POST.get('respondent_type', 'all')
                if respondent_type != 'all':
                    complaints = complaints.filter(respondent_type=respondent_type)
            elif export_option == 'custom':
                # Apply multiple custom filters
                custom_status = request.POST.get('custom_status', '')
                custom_respondent_type = request.POST.get('custom_respondent_type', '')
                include_archived = request.POST.get('include_archived', 'no')
                has_hearing = request.POST.get('has_hearing', '')

                # Apply status filter
                if custom_status:
                    complaints = complaints.filter(status=custom_status)

                # Apply respondent type filter
                if custom_respondent_type:
                    complaints = complaints.filter(respondent_type=custom_respondent_type)

                # Apply archived filter
                if include_archived == 'no':
                    complaints = complaints.filter(is_archived=False)
                elif include_archived == 'archived_only':
                    complaints = complaints.filter(is_archived=True)
                # 'yes' includes both archived and non-archived

                # Apply hearing filter
                if has_hearing == 'yes':
                    # Get complaints that have at least one hearing
                    complaints = complaints.filter(hearings__isnull=False).distinct()
                elif has_hearing == 'no':
                    # Get complaints that have no hearings
                    complaints = complaints.filter(hearings__isnull=True)

            # Apply date range if provided
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            date_type = request.POST.get('date_type', 'incident')

            if start_date:
                if date_type == 'incident':
                    complaints = complaints.filter(incident_date__gte=start_date)
                elif date_type == 'submission':
                    complaints = complaints.filter(created_at__gte=start_date)
                elif date_type == 'updated':
                    complaints = complaints.filter(updated_at__gte=start_date)

            if end_date:
                try:
                    # Add one day to include the entire end date
                    end_date_obj = timezone.datetime.strptime(end_date, '%Y-%m-%d')
                    end_date_obj = end_date_obj + timezone.timedelta(days=1)

                    if date_type == 'incident':
                        complaints = complaints.filter(incident_date__lte=end_date_obj)
                    elif date_type == 'submission':
                        complaints = complaints.filter(created_at__lte=end_date_obj)
                    elif date_type == 'updated':
                        complaints = complaints.filter(updated_at__lte=end_date_obj)
                except ValueError:
                    # Handle invalid date format
                    pass

            # Check if Excel format is selected
            if file_format == 'excel':
                # Load the template
                template_path = os.path.join(settings.STATICFILES_DIRS[0], 'templates', 'ComplaintTemplate.xlsx')

                # If staticfiles_dirs doesn't work, try alternative path
                if not os.path.exists(template_path):
                    template_path = os.path.join(settings.BASE_DIR, 'static', 'templates', 'ComplaintTemplate.xlsx')

                if not os.path.exists(template_path):
                    return HttpResponse("Template file not found", status=500)

                # Load workbook
                wb = load_workbook(template_path)
                ws = wb.active  # Get the active sheet (Sheet1)

                # Clear rows 8 and below for data
                max_row = ws.max_row
                for row in range(8, max_row + 1):
                    for col in range(1, 17):  # Columns A to P (1-16)
                        ws.cell(row=row, column=col, value=None)

                # Start writing data at row 8
                start_row = 8

                # Write complaint data to the template
                for i, complaint in enumerate(complaints, start=start_row):
                    # Get respondent course
                    respondent_course = str(complaint.respondent_course) if complaint.respondent_course else ''

                    # Get respondent department
                    respondent_department = complaint.respondent_department or ''

                    # Format incident time
                    incident_time = ''
                    if complaint.incident_time:
                        incident_time = complaint.incident_time.strftime('%H:%M')

                    # Column A: Reference No
                    ws.cell(row=i, column=1, value=complaint.reference_number or '')

                    # Column B: Complainant Name
                    complainant_name = f"{complaint.complainant_last_name}, {complaint.complainant_first_name}"
                    ws.cell(row=i, column=2, value=complainant_name)

                    # Column C: Complainant Email
                    ws.cell(row=i, column=3, value=complaint.complainant_email or '')

                    # Column D: Complainant Phone
                    ws.cell(row=i, column=4, value=complaint.complainant_phone or '')

                    # Column E: Respondent Name
                    respondent_name = f"{complaint.respondent_last_name}, {complaint.respondent_first_name}"
                    ws.cell(row=i, column=5, value=respondent_name)

                    # Column F: Respondent Type
                    ws.cell(row=i, column=6, value=complaint.get_respondent_type_display())

                    # Column G: Respondent Year
                    ws.cell(row=i, column=7, value=complaint.respondent_year or '')

                    # Column H: Respondent Section
                    ws.cell(row=i, column=8, value=complaint.respondent_section or '')

                    # Column I: Respondent Course
                    ws.cell(row=i, column=9, value=respondent_course)

                    # Column J: Respondent Department
                    ws.cell(row=i, column=10, value=respondent_department)

                    # Column K: Complaint Title
                    ws.cell(row=i, column=11, value=complaint.title)

                    # Column L: Incident Date
                    if complaint.incident_date:
                        ws.cell(row=i, column=12, value=complaint.incident_date.strftime('%Y-%m-%d'))
                    else:
                        ws.cell(row=i, column=12, value='')

                    # Column M: Incident Time
                    ws.cell(row=i, column=13, value=incident_time)

                    # Column N: Incident Location
                    ws.cell(row=i, column=14, value=complaint.incident_location or '')

                    # Column O: Date Submitted
                    if complaint.created_at:
                        ws.cell(row=i, column=15, value=complaint.created_at.strftime('%Y-%m-%d %H:%M'))
                    else:
                        ws.cell(row=i, column=15, value='')

                    # Column P: Status
                    ws.cell(row=i, column=16, value=complaint.get_status_display())

                # Save workbook to a BytesIO object
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                # Create response with template
                response = HttpResponse(
                    output.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="complaints_export.xlsx"'

                return response

            elif file_format == 'csv':
                # Prepare comprehensive data for CSV export
                data = []
                for complaint in complaints:
                    # Get hearing information
                    next_hearing = complaint.next_hearing
                    next_hearing_display = ''
                    if next_hearing:
                        next_hearing_display = f"{next_hearing.schedule_date.strftime('%Y-%m-%d')} {next_hearing.schedule_time.strftime('%H:%M')}"

                    # Get hearing count
                    hearing_count = complaint.hearings.count()

                    # Get document and image counts
                    document_count = complaint.documents.count()
                    image_count = complaint.images.count()

                    # Get complainant instructor name
                    complainant_instructor = complaint.complainant_instructor_name or ''

                    # Get statement (first 100 chars)
                    statement_preview = complaint.statement[:100] + '...' if len(
                        complaint.statement) > 100 else complaint.statement

                    # Get witnesses
                    witnesses = complaint.witnesses or ''

                    # Get notes (first 100 chars)
                    notes_preview = complaint.notes[:100] + '...' if complaint.notes and len(
                        complaint.notes) > 100 else (complaint.notes or '')

                    # Get archived information
                    archived_by = ''
                    archived_at = ''
                    if complaint.is_archived and complaint.archived_by:
                        archived_by = f"{complaint.archived_by.last_name}, {complaint.archived_by.first_name}"
                        if complaint.archived_at:
                            archived_at = complaint.archived_at.strftime('%Y-%m-%d %H:%M')

                    data.append({
                        'Reference Number': complaint.reference_number,
                        'Title': complaint.title,

                        # Complainant Information
                        'Complainant': f"{complaint.complainant_last_name}, {complaint.complainant_first_name}",
                        'Complainant Email': complaint.complainant_email,
                        'Complainant Phone': complaint.complainant_phone,
                        'Complainant Address': complaint.complainant_address,
                        'Complainant Instructor': complainant_instructor,

                        # Respondent Information
                        'Respondent Type': complaint.get_respondent_type_display(),
                        'Respondent': f"{complaint.respondent_last_name}, {complaint.respondent_first_name}",
                        'Respondent Course': str(complaint.respondent_course) if complaint.respondent_course else '',
                        'Respondent Year': complaint.respondent_year or '',
                        'Respondent Section': complaint.respondent_section or '',
                        'Respondent Department': complaint.respondent_department or '',

                        # Complaint Details
                        'Incident Date': complaint.incident_date.strftime('%Y-%m-%d'),
                        'Incident Time': complaint.incident_time.strftime('%H:%M') if complaint.incident_time else '',
                        'Incident Location': complaint.incident_location,
                        'Statement Preview': statement_preview,
                        'Witnesses': witnesses,
                        'Notes Preview': notes_preview,

                        # Status and Dates
                        'Status': complaint.get_status_display(),
                        'Date Submitted': complaint.created_at.strftime('%Y-%m-%d %H:%M'),
                        'Last Updated': complaint.updated_at.strftime('%Y-%m-%d %H:%M'),
                        'Date Resolved': complaint.updated_at.strftime(
                            '%Y-%m-%d %H:%M') if complaint.status == 'resolved' else '',

                        # Hearing Information
                        'Next Hearing': next_hearing_display,
                        'Hearing Count': hearing_count,

                        # Evidence Information
                        'Document Count': document_count,
                        'Image Count': image_count,

                        # Created By
                        'Created By': f"{complaint.created_by.last_name}, {complaint.created_by.first_name}" if complaint.created_by else '',

                        # Archived Information
                        'Archived Status': 'Yes' if complaint.is_archived else 'No',
                        'Archived By': archived_by,
                        'Archived At': archived_at,
                    })

                # Create DataFrame
                df = pd.DataFrame(data)

                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename="complaints_export.csv"'
                df.to_csv(response, index=False, encoding='utf-8')

                return response

        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"Complaint export error: {str(e)}")
            print(f"Traceback: {error_trace}")
            return HttpResponse(f"Error during export: {str(e)}", status=500)

    return HttpResponse('Invalid request', status=400)


# ------------------------------------------------ Activity Log Section ------------------------------------------------
@require_GET
def recent_activities(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    # Get filter parameters
    activity_type = request.GET.get('type', 'all')
    user_filter = request.GET.get('user', 'all')
    search_query = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    page_number = request.GET.get('page', 1)

    # Base queryset
    activities = UserActivityLog.objects.select_related('user').order_by('-timestamp')

    # Apply filters
    if activity_type != 'all':
        type_filters = {
            'create': Q(activity__icontains='create') | Q(activity__icontains='created'),
            'update': Q(activity__icontains='update') | Q(activity__icontains='updated'),
            'archive': Q(activity__icontains='archive') | Q(activity__icontains='archived'),
        }
        if activity_type in type_filters:
            activities = activities.filter(type_filters[activity_type])

    if user_filter != 'all' and user_filter != 'all':
        activities = activities.filter(user_id=user_filter)

    if search_query:
        activities = activities.filter(
            Q(activity__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query)
        )

    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            activities = activities.filter(timestamp__date__gte=date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            activities = activities.filter(timestamp__date__lte=date_to_obj)
        except ValueError:
            pass

    # Paginate results for normal display
    paginator = Paginator(activities, 50)
    page_obj = paginator.get_page(page_number)

    activity_list = []
    for activity in page_obj:
        activity_list.append({
            'id': activity.id,
            'user': activity.user.get_full_name() or activity.user.username,
            'activity': activity.activity,
            'timestamp': activity.timestamp.isoformat(),
            'user_type': activity.user.get_user_type_display(),
            'ip_address': getattr(activity, 'ip_address', None),
            'details': getattr(activity, 'details', None),
        })

    return JsonResponse({
        'activities': activity_list,
        'has_next': page_obj.has_next(),
        'has_previous': page_obj.has_previous(),
        'current_page': page_obj.number,
        'total_pages': paginator.num_pages,
        'total_activities': paginator.count
    })


@require_GET
def export_activities(request):
    if not request.user.is_authenticated:
        return HttpResponse('Unauthorized', status=401)

    # Check if user has permission to export
    if not (request.user.is_superuser or request.user.user_type == 1):
        return HttpResponse('Forbidden: Insufficient permissions', status=403)

    # Get filter parameters
    activity_type = request.GET.get('type', 'all')
    user_filter = request.GET.get('user', 'all')
    search_query = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Base queryset
    activities = UserActivityLog.objects.select_related('user').order_by('-timestamp')

    # Apply filters
    if activity_type != 'all':
        type_filters = {
            'create': Q(activity__icontains='create') | Q(activity__icontains='created'),
            'update': Q(activity__icontains='update') | Q(activity__icontains='updated'),
            'archive': Q(activity__icontains='archive') | Q(activity__icontains='archived'),
        }
        if activity_type in type_filters:
            activities = activities.filter(type_filters[activity_type])

    if user_filter != 'all' and user_filter != 'all':
        activities = activities.filter(user_id=user_filter)

    if search_query:
        activities = activities.filter(
            Q(activity__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query)
        )

    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            activities = activities.filter(timestamp__date__gte=date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            activities = activities.filter(timestamp__date__lte=date_to_obj)
        except ValueError:
            pass

    # Create CSV response with proper Excel formatting
    response = HttpResponse(
        content_type='text/csv; charset=utf-8-sig',
        headers={'Content-Disposition': 'attachment; filename="activity_log_export.csv"'},
    )

    writer = csv.writer(response)

    # Write headers
    writer.writerow(['ID', 'User', 'Activity', 'User Type', 'Timestamp', 'Date', 'Time'])

    # Write data rows with actual data from activities
    for activity in activities:
        timestamp = activity.timestamp

        # Format the timestamp properly
        full_timestamp = timestamp.strftime('%Y-%m-%d %H:%M:%S')
        date_str = timestamp.strftime('%Y-%m-%d')
        time_str = timestamp.strftime('%H:%M:%S')

        writer.writerow([
            activity.id,
            activity.user.get_full_name() or activity.user.username,
            activity.activity,
            activity.user.get_user_type_display(),
            full_timestamp,
            date_str,
            time_str
        ])

    return response


@require_GET
def activity_users(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    # Get unique users who have activities
    user_ids = UserActivityLog.objects.values_list('user_id', flat=True).distinct()
    users = CustomUser.objects.filter(id__in=user_ids)

    user_list = [{'id': 'all', 'name': 'All Users'}]
    for user in users:
        user_list.append({
            'id': user.id,
            'name': user.get_full_name() or user.username
        })

    return JsonResponse(user_list, safe=False)


@require_GET
def activity_log_overview(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    # Calculate today's activities
    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_activities = UserActivityLog.objects.filter(
        timestamp__gte=today_start
    ).count()

    # Total activities
    total_activities = UserActivityLog.objects.count()

    # Activities by type (last 7 days)
    week_ago = timezone.now() - timedelta(days=7)
    recent_activities = UserActivityLog.objects.filter(timestamp__gte=week_ago)

    activity_types = {
        'creations': recent_activities.filter(
            Q(activity__icontains='create') | Q(activity__icontains='created')).count(),
        'updates': recent_activities.filter(Q(activity__icontains='update') | Q(activity__icontains='updated')).count(),
        'archives': recent_activities.filter(
            Q(activity__icontains='archive') | Q(activity__icontains='archived')).count(),
    }

    return JsonResponse({
        'today_activities': today_activities,
        'total_activities': total_activities,
        'recent_activity_types': activity_types,
    })


# --------------------------------------------------- Archived Section -------------------------------------------------
class ArchivedItemDetailView(LoginRequiredMixin, View):
    def get_file_type_icon(self, filename):
        ext = filename.split('.')[-1].lower() if '.' in filename else ''
        if ext in ['pdf']:
            return 'fas fa-file-pdf'
        elif ext in ['doc', 'docx']:
            return 'fas fa-file-word'
        elif ext in ['xls', 'xlsx']:
            return 'fas fa-file-excel'
        elif ext in ['ppt', 'pptx']:
            return 'fas fa-file-powerpoint'
        elif ext in ['jpg', 'jpeg', 'png', 'gif', 'bmp']:
            return 'fas fa-file-image'
        elif ext in ['zip', 'rar', '7z']:
            return 'fas fa-file-archive'
        else:
            return 'fas fa-file'

    def get(self, request, item_type, pk):
        if item_type == 'user':
            item = get_object_or_404(CustomUser, pk=pk, is_archived=True)

            # Base user data
            data = {
                'id': item.id,
                'username': item.username,
                'first_name': item.first_name,
                'last_name': item.last_name,
                'email': item.email,
                'user_type_display': item.get_user_type_display(),
                'gender_display': item.get_gender_display(),
                'birth_date': item.birth_date.strftime('%B %d, %Y') if item.birth_date else None,
                'phone_number': item.phone_number,
                'address': item.address,
                'is_active': item.is_active,
                'is_verified': item.is_verified,
                'date_joined': item.date_joined.strftime('%B %d, %Y %H:%M'),
                'last_login': item.last_login.strftime('%B %d, %Y %H:%M') if item.last_login else None,
                'archived_at': item.archived_at.strftime('%B %d, %Y %H:%M') if item.archived_at else None,
                'archived_by': item.archived_by.get_full_name() if item.archived_by else None,
                'profile_picture': item.profile_picture.url if item.profile_picture else None,

                # Student-specific fields
                'student_number': item.student_number,
                'course': str(item.course) if item.course else None,
                'year_level': item.year_level,
                'section': item.section,

                # OSAS staff-specific fields
                'department': item.department,
                'position': item.position,

                # Verification documents
                'id_photo_url': item.id_photo.url if item.id_photo else None,
                'cor_photo_url': item.cor_photo.url if item.cor_photo else None,

                # Type indicators
                'is_student': item.is_student,
                'is_osas_unit': item.is_osas_unit,
                'is_organization': item.is_organization,
            }

            # Add display name based on user type
            if item.is_organization and hasattr(item, 'organization_account'):
                org = item.organization_account
                data['display_name'] = org.organization_name
            else:
                data['display_name'] = f"{item.first_name} {item.last_name}".strip() or item.username

            # Add organization-specific data if this is an organization user
            if item.is_organization and hasattr(item, 'organization_account'):
                org = item.organization_account
                data.update({
                    'organization_name': org.organization_name,
                    'organization_acronym': org.organization_acronym,
                    'organization_type': org.get_organization_type_display(),
                    'organization_email': org.organization_email,
                    'organization_status': org.organization_status,
                    'organization_status_display': dict(org.ORGANIZATION_STATUS_CHOICES).get(org.organization_status,
                                                                                             'Unknown'),
                    'organization_valid_from': org.organization_valid_from.strftime(
                        '%B %d, %Y') if org.organization_valid_from else None,
                    'organization_valid_until': org.organization_valid_until.strftime(
                        '%B %d, %Y') if org.organization_valid_until else None,
                    'organization_adviser_name': org.organization_adviser_name,
                    'organization_adviser_department': org.organization_adviser_department,
                    'organization_adviser_email': org.organization_adviser_email,
                    'organization_adviser_phone': org.organization_adviser_phone,
                    'organization_member_count': org.organization_member_count,
                    'organization_has_minimum_members': org.organization_has_minimum_members,
                    'organization_members': org.organization_members or [],

                    # Organization document URLs
                    'organization_logo_url': org.organization_logo.url if org.organization_logo else None,
                    'organization_calendar_activities_url': org.organization_calendar_activities.url if org.organization_calendar_activities else None,
                    'organization_adviser_cv_url': org.organization_adviser_cv.url if org.organization_adviser_cv else None,
                    'organization_cog_url': org.organization_cog.url if org.organization_cog else None,
                    'organization_group_picture_url': org.organization_group_picture.url if org.organization_group_picture else None,
                    'organization_cbl_url': org.organization_cbl.url if org.organization_cbl else None,
                    'organization_list_members_url': org.organization_list_members.url if org.organization_list_members else None,
                    'organization_acceptance_letter_url': org.organization_acceptance_letter.url if org.organization_acceptance_letter else None,
                    'organization_ar_url': org.organization_ar.url if org.organization_ar else None,
                    'organization_previous_calendar_url': org.organization_previous_calendar.url if org.organization_previous_calendar else None,
                    'organization_financial_report_url': org.organization_financial_report.url if org.organization_financial_report else None,
                    'organization_coa_url': org.organization_coa.url if org.organization_coa else None,
                    'organization_member_biodata_url': org.organization_member_biodata.url if org.organization_member_biodata else None,
                    'organization_good_moral_url': org.organization_good_moral.url if org.organization_good_moral else None,
                })

        elif item_type == 'announcement':
            item = get_object_or_404(Announcement, pk=pk, is_archived=True)
            data = {
                'id': item.id,
                'title': item.title,
                'content': item.content,
                'category_display': item.get_category_display(),
                'author_name': item.author.get_full_name(),
                'created_at': item.created_at.strftime('%B %d, %Y %H:%M'),
                'publish_date': item.publish_date.strftime('%B %d, %Y %H:%M') if item.publish_date else None,
                'link': item.link,
                'archived_at': item.archived_at.strftime('%B %d, %Y %H:%M') if item.archived_at else None,
                'archived_by': item.archived_by.get_full_name() if item.archived_by else None,
                'images': [{
                    'url': img.image.url,
                    'caption': img.caption
                } for img in item.images.all()]
            }

        elif item_type == 'downloadable':
            item = get_object_or_404(Downloadable, pk=pk, is_archived=True)
            data = {
                'id': item.id,
                'title': item.title,
                'description': item.description,
                'category_display': item.get_category_display(),
                'file_name': item.get_file_name(),
                'file_size': item.get_file_size(),
                'file_url': item.file.url,
                'created_at': item.created_at.strftime('%B %d, %Y %H:%M'),
                'created_by': item.created_by.get_full_name(),
                'archived_at': item.archived_at.strftime('%B %d, %Y %H:%M') if item.archived_at else None,
                'archived_by': item.archived_by.get_full_name() if item.archived_by else None,
            }

        elif item_type == 'complaint':
            item = get_object_or_404(Complaint, pk=pk, is_archived=True)
            # Permission check
            if not (request.user.is_superuser or
                    request.user.user_type in [1, 11] or
                    item.created_by == request.user or
                    item.archived_by == request.user):
                return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

            # Prepare course data
            respondent_course_data = None

            if item.respondent_course:
                respondent_course_data = {
                    'id': item.respondent_course.id,
                    'name': item.respondent_course.name,
                    'subtext': item.respondent_course.subtext,
                    'logo_url': item.respondent_course.logo.url if item.respondent_course.logo else None
                }

            data = {
                'id': item.id,
                'reference_number': item.reference_number,
                'title': item.title,
                'status': item.status,
                'status_display': item.get_status_display(),
                'statement': item.statement,
                'notes': item.notes,
                'complainant_first_name': item.complainant_first_name,
                'complainant_last_name': item.complainant_last_name,
                'complainant_email': item.complainant_email,
                'complainant_phone': item.complainant_phone,
                'complainant_address': item.complainant_address,
                'complainant_instructor_name': item.complainant_instructor_name,
                'respondent_first_name': item.respondent_first_name,
                'respondent_last_name': item.respondent_last_name,
                'respondent_type': item.respondent_type,
                'respondent_type_display': item.get_respondent_type_display(),
                'respondent_course': respondent_course_data,
                'respondent_year': item.respondent_year,
                'respondent_section': item.respondent_section,
                'respondent_department': item.respondent_department,
                'incident_date': item.incident_date.strftime('%B %d, %Y') if item.incident_date else None,
                'incident_time': item.incident_time.strftime('%H:%M') if item.incident_time else None,
                'incident_location': item.incident_location,
                'witnesses': item.witnesses,
                'created_at': item.created_at.strftime('%B %d, %Y %H:%M'),
                'updated_at': item.updated_at.strftime('%B %d, %Y %H:%M') if item.updated_at else None,
                'archived_at': item.archived_at.strftime('%B %d, %Y %H:%M') if item.archived_at else None,
                'archived_by': item.archived_by.get_full_name() if item.archived_by else None,
                'created_by': item.created_by.get_full_name() if item.created_by else None,

                'documents': [{
                    'id': doc.id,
                    'file_url': doc.file.url,
                    'file_name': doc.file.name.split('/')[-1],
                    'description': doc.description,
                    'uploaded_at': doc.uploaded_at.strftime('%B %d, %Y %H:%M')
                } for doc in item.documents.all()],

                'images': [{
                    'id': img.id,
                    'image_url': img.image.url,
                    'caption': img.caption,
                    'uploaded_at': img.uploaded_at.strftime('%B %d, %Y %H:%M')
                } for img in item.images.all()],
            }
        elif item_type == 'scholarship':
            item = get_object_or_404(Scholarship, pk=pk, is_archived=True)
            data = {
                'id': item.id,
                'name': item.name,
                'description': item.description,
                'scholarship_type': item.scholarship_type,
                'scholarship_type_display': item.get_scholarship_type_display(),
                'benefits': item.benefits,
                'requirements': item.requirements,
                'is_active': item.is_active,
                'created_at': item.created_at.strftime('%B %d, %Y %H:%M'),
                'updated_at': item.updated_at.strftime('%B %d, %Y %H:%M') if item.updated_at else None,
                'created_by': item.created_by.get_full_name() if item.created_by else None,
                'archived_at': item.archived_at.strftime('%B %d, %Y %H:%M') if item.archived_at else None,
                'archived_by': item.archived_by.get_full_name() if item.archived_by else None,
                'application_form': {
                    'title': item.application_form.title if item.application_form else None,
                    'file_url': item.application_form.file.url if item.application_form else None,
                    'file_name': item.application_form.get_file_name() if item.application_form else None,
                } if item.application_form else None
            }
        elif item_type == 'scholarship-application':
            item = get_object_or_404(ScholarshipApplication, pk=pk, is_archived=True)
            data = {
                'id': item.id,
                'status': item.status,
                'status_display': item.get_status_display(),
                'application_date': item.application_date.strftime('%B %d, %Y %H:%M'),
                'status_update_date': item.status_update_date.strftime(
                    '%B %d, %Y %H:%M') if item.status_update_date else None,
                'notes': item.notes,
                'archived_at': item.archived_at.strftime('%B %d, %Y %H:%M') if item.archived_at else None,
                'archived_by': item.archived_by.get_full_name() if item.archived_by else None,
                'student': {
                    'full_name': item.student.get_full_name(),
                    'student_number': item.student.student_number,
                    'course': str(item.student.course) if item.student.course else None,
                    'email': item.student.email,
                    'phone_number': item.student.phone_number,
                    'profile_picture': item.student.profile_picture.url if item.student.profile_picture else None,
                },
                'scholarship': {
                    'name': item.scholarship.name,
                    'description': item.scholarship.description,
                    'type_display': item.scholarship.get_scholarship_type_display(),
                },
                'application_form_url': item.application_form.url if item.application_form else None,
                'cog_url': item.cog.url if item.cog else None,
                'cor_url': item.cor.url if item.cor else None,
                'id_photo_url': item.id_photo.url if item.id_photo else None,
                'other_documents_url': item.other_documents.url if item.other_documents else None,
            }
        elif item_type == 'admission':
            item = get_object_or_404(StudentAdmission, pk=pk, is_archived=True)
            data = {
                'id': item.id,
                'control_no': item.control_no,
                'student_type_display': item.get_student_type_display(),
                'course': {
                    'id': item.course.id if item.course else None,
                    'name': item.course.name if item.course else None,
                },
                'status_display': item.get_status_display(),
                'date': item.date.strftime('%B %d, %Y'),
                'created_at': item.created_at.strftime('%B %d, %Y %H:%M'),
                'updated_at': item.updated_at.strftime('%B %d, %Y %H:%M') if item.updated_at else None,
                'archived_at': item.archived_at.strftime('%B %d, %Y %H:%M') if item.archived_at else None,
                'archived_by': item.archived_by.get_full_name() if item.archived_by else None,
                'remarks': item.remarks,
                'admission_portal_registration': item.admission_portal_registration,

                # Student type specific fields
                'strand': item.strand,
                'grade11_report_card': item.grade11_report_card.url if item.grade11_report_card else None,
                'certificate_of_enrollment': item.certificate_of_enrollment.url if item.certificate_of_enrollment else None,
                'grade12_report_card': item.grade12_report_card.url if item.grade12_report_card else None,
                'form137': item.form137.url if item.form137 else None,

                # Transferee specific fields
                'curriculum_type_display': item.get_curriculum_type_display() if item.curriculum_type else None,
                'first_year_first_semester_display': item.get_first_year_first_semester_display() if item.first_year_first_semester else None,
                'first_year_second_semester_display': item.get_first_year_second_semester_display() if item.first_year_second_semester else None,
                'second_year_first_semester_display': item.get_second_year_first_semester_display() if item.second_year_first_semester else None,
                'other_semester_info': item.other_semester_info,
                'transcript_of_grades': item.transcript_of_grades.url if item.transcript_of_grades else None,
                'good_moral_certificate': item.good_moral_certificate.url if item.good_moral_certificate else None,
                'honorable_dismissal': item.honorable_dismissal.url if item.honorable_dismissal else None,
                'nbi_police_clearance': item.nbi_police_clearance.url if item.nbi_police_clearance else None,
            }
        elif item_type == 'nstp-student':
            item = get_object_or_404(NSTPStudentInfo, pk=pk, is_archived=True)
            data = {
                'id': item.id,
                'first_name': item.first_name,
                'last_name': item.last_name,
                'middle_name': item.middle_name,
                'student_number': item.student_number,
                'program': item.program,
                'gender': item.gender,
                'gender_display': item.get_gender_display(),
                'birth_date': item.birth_date.strftime('%B %d, %Y') if item.birth_date else None,
                'contact_number': item.contact_number,
                'email_address': item.email_address,
                'street_or_barangay': item.street_or_barangay,
                'municipality_or_city': item.municipality_or_city,
                'province': item.province,
                'approval_status': item.approval_status,
                'approval_status_display': item.get_approval_status_display(),
                'semester': item.semester,
                'semester_display': item.get_semester_display(),
                'academic_year': item.academic_year,
                'remarks': item.remarks,
                'archived_at': item.archived_at.strftime('%B %d, %Y %H:%M') if item.archived_at else None,
                'archived_by': item.archived_by.get_full_name() if item.archived_by else None,
                'user': {
                    'full_name': item.user.get_full_name(),
                    'profile_picture': item.user.profile_picture.url if item.user.profile_picture else None,
                    'email': item.user.email,
                }
            }
        elif item_type == 'nstp-file':
            item = get_object_or_404(NSTPFile, pk=pk, is_archived=True)
            data = {
                'id': item.id,
                'title': item.title,
                'description': item.description,
                'category_display': item.get_category_display(),
                'semester_display': item.get_semester_display(),
                'school_year': item.school_year,
                'file_name': item.get_file_name(),
                'file_size': item.get_file_size(),
                'file_url': item.file.url,
                'created_at': item.created_at.strftime('%B %d, %Y %H:%M'),
                'created_by': item.created_by.get_full_name(),
                'updated_at': item.updated_at.strftime('%B %d, %Y %H:%M') if item.updated_at else None,
                'archived_at': item.archived_at.strftime('%B %d, %Y %H:%M') if item.archived_at else None,
                'archived_by': item.archived_by.get_full_name() if item.archived_by else None,
            }
        elif item_type == 'ojt-company':
            item = get_object_or_404(OJTCompany, pk=pk, is_archived=True)
            data = {
                'id': item.id,
                'name': item.name,
                'address': item.address,
                'contact_number': item.contact_number,
                'description': item.description or "",
                'website': item.website or "",
                'email': item.email or "",
                'status': item.display_status,  # Use display_status property
                'created_at': item.created_at.strftime('%B %d, %Y %H:%M'),
                'updated_at': item.updated_at.strftime('%B %d, %Y %H:%M') if item.updated_at else "",
                'status_updated_at': item.status_updated_at.strftime(
                    '%B %d, %Y %H:%M') if item.status_updated_at else "",
                'status_updated_by': item.status_updated_by.get_full_name() if item.status_updated_by else "",
                'archived_at': item.archived_at.strftime('%B %d, %Y %H:%M') if item.archived_at else "",
                'archived_by': item.archived_by.get_full_name() if item.archived_by else "System",
            }
        elif item_type == 'organization':
            item = get_object_or_404(Organization, pk=pk, is_archived=True)
            data = {
                'id': item.id,
                'organization_name': item.organization_name,
                'organization_acronym': item.organization_acronym,
                'organization_description': item.organization_description,
                'organization_mission': item.organization_mission,
                'organization_vision': item.organization_vision,
                'organization_type_display': item.get_organization_type_display(),
                'organization_email': item.organization_email,
                'organization_status': item.organization_status,
                'organization_status_display': dict(item.ORGANIZATION_STATUS_CHOICES).get(item.organization_status,
                                                                                          'Unknown'),
                'organization_valid_from': item.organization_valid_from.strftime(
                    '%B %d, %Y') if item.organization_valid_from else None,
                'organization_valid_until': item.organization_valid_until.strftime(
                    '%B %d, %Y') if item.organization_valid_until else None,
                'organization_adviser_name': item.organization_adviser_name,
                'organization_adviser_department': item.organization_adviser_department,
                'organization_adviser_email': item.organization_adviser_email,
                'organization_adviser_phone': item.organization_adviser_phone,
                'organization_member_count': item.organization_member_count,
                'organization_has_minimum_members': item.organization_has_minimum_members,
                'all_requirements_submitted': item.all_requirements_submitted,
                'created_at': item.created_at.strftime('%B %d, %Y %H:%M'),
                'updated_at': item.updated_at.strftime('%B %d, %Y %H:%M') if item.updated_at else None,
                'archived_at': item.archived_at.strftime('%B %d, %Y %H:%M') if item.archived_at else None,
                'archived_by': item.archived_by.get_full_name() if item.archived_by else None,
                'organization_members': item.organization_members or [],

                # Document URLs
                'organization_logo_url': item.organization_logo.url if item.organization_logo else None,
                'organization_calendar_activities_url': item.organization_calendar_activities.url if item.organization_calendar_activities else None,
                'organization_adviser_cv_url': item.organization_adviser_cv.url if item.organization_adviser_cv else None,
                'organization_cog_url': item.organization_cog.url if item.organization_cog else None,
                'organization_group_picture_url': item.organization_group_picture.url if item.organization_group_picture else None,
                'organization_cbl_url': item.organization_cbl.url if item.organization_cbl else None,
                'organization_list_members_url': item.organization_list_members.url if item.organization_list_members else None,
                'organization_acceptance_letter_url': item.organization_acceptance_letter.url if item.organization_acceptance_letter else None,
                'organization_ar_url': item.organization_ar.url if item.organization_ar else None,
                'organization_previous_calendar_url': item.organization_previous_calendar.url if item.organization_previous_calendar else None,
                'organization_financial_report_url': item.organization_financial_report.url if item.organization_financial_report else None,
                'organization_coa_url': item.organization_coa.url if item.organization_coa else None,
                'organization_member_biodata_url': item.organization_member_biodata.url if item.organization_member_biodata else None,
                'organization_good_moral_url': item.organization_good_moral.url if item.organization_good_moral else None,
            }
        elif item_type == 'accomplishment-report':
            item = get_object_or_404(AccomplishmentRecord, pk=pk, is_archived=True)

            # Check permissions based on your requirements
            if not (request.user.is_superuser or
                    request.user.user_type in [1, 10] or
                    (request.user.user_type == 15 and
                     item.organization and
                     item.organization.user_account == request.user) or
                    item.archived_by == request.user):
                return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

            data = {
                'id': item.id,
                'title': item.title,
                'record_type': item.record_type,
                'record_type_display': item.get_record_type_display(),
                'date_conducted': item.date_conducted.strftime('%B %d, %Y'),
                'venue': item.venue,
                'semester': item.semester,
                'semester_display': item.get_semester_display(),
                'school_year': item.school_year,
                'objectives': item.objectives,
                'outcomes': item.outcomes,
                'number_of_participants': item.number_of_participants,
                'duration_hours': float(item.duration_hours),
                'budget_utilized': float(item.budget_utilized) if item.budget_utilized else None,
                'archive_reason': item.archive_reason,
                'created_at': item.created_at.strftime('%B %d, %Y %H:%M'),
                'updated_at': item.updated_at.strftime('%B %d, %Y %H:%M') if item.updated_at else None,
                'archived_at': item.archived_at.strftime('%B %d, %Y %H:%M') if item.archived_at else None,
                'archived_by': item.archived_by.get_full_name() if item.archived_by else None,
                'submitted_by': item.submitted_by.get_full_name() if item.submitted_by else None,
                'organization': {
                    'name': item.organization.organization_name if item.organization else None,
                    'acronym': item.organization.organization_acronym if item.organization else None,
                } if item.organization else None,
                'main_report': {
                    'url': item.main_report.url if item.main_report else None,
                    'name': item.main_report.name.split('/')[-1] if item.main_report else None,
                },
                'supporting_files': [{
                    'id': file.id,
                    'url': file.file.url,
                    'name': file.filename,
                    'description': file.description,
                    'uploaded_at': file.uploaded_at.strftime('%B %d, %Y %H:%M')
                } for file in item.supporting_files.all()]
            }
        else:
            return JsonResponse({'success': False, 'error': 'Invalid item type'}, status=400)

        return JsonResponse({'success': True, item_type: data})


class RetrieveArchivedItemView(LoginRequiredMixin, View):
    def post(self, request, item_type, pk):
        if not request.user.is_authenticated:
            return JsonResponse({'success': False, 'error': 'Authentication required'}, status=401)

        try:
            if item_type == 'user':
                if not request.user.is_superuser:
                    return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

                item = get_object_or_404(CustomUser, pk=pk, is_archived=True)

                from django.db import transaction
                with transaction.atomic():
                    # Retrieve the user
                    item.is_archived = False
                    item.is_active = True
                    item.archived_at = None
                    item.archived_by = None
                    item.save()

                    organization_retrieved = False
                    if item.user_type == 15 and hasattr(item, 'organization_account'):
                        organization = item.organization_account

                        # Retrieve the organization
                        organization.is_archived = False
                        organization.is_active = True  # Make organization active
                        organization.archived_at = None
                        organization.archived_by = None

                        # Set organization status to 'cancelled' as requested
                        organization._organization_status = 'cancelled'

                        organization.save()
                        organization_retrieved = True

                # Log the activity
                if item.user_type == 15 and organization_retrieved:
                    UserActivityLog.objects.create(
                        user=request.user,
                        activity=f"{request.user.first_name} retrieved organization user: {item.get_full_name()} and linked organization: {organization.organization_name} (Status set to: cancelled)"
                    )
                else:
                    UserActivityLog.objects.create(
                        user=request.user,
                        activity=f"{request.user.first_name} retrieved user: {item.get_full_name()}"
                    )

            elif item_type == 'announcement':
                item = get_object_or_404(Announcement, pk=pk, is_archived=True)

                # Check permissions - author or superuser can retrieve
                if not (request.user.is_superuser or item.author == request.user):
                    return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

                item.is_archived = False
                item.archived_at = None
                item.archived_by = None
                item.save()

                # Log the activity
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"{request.user.first_name} retrieved announcement: {item.title}"
                )

            elif item_type == 'downloadable':
                item = get_object_or_404(Downloadable, pk=pk, is_archived=True)

                # Check permissions - creator or superuser can retrieve
                if not (request.user.is_superuser or item.created_by == request.user):
                    return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

                item.is_archived = False
                item.archived_at = None
                item.archived_by = None
                item.save()

                # Log the activity
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"{request.user.first_name} retrieved downloadable: {item.title}"
                )

            elif item_type == 'complaint':
                item = get_object_or_404(Complaint, pk=pk, is_archived=True)
                if not (request.user.is_superuser or
                        request.user.user_type in [1, 11] or
                        item.created_by == request.user or
                        item.archived_by == request.user):
                    return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

                item.is_archived = False
                item.status = 'under_review'
                item.archived_at = None
                item.archived_by = None

                retrieval_note = f"\n\nRetrieved on {timezone.now().strftime('%Y-%m-%d %H:%M')} by {request.user.get_full_name()}"
                if item.notes:
                    item.notes += retrieval_note
                else:
                    item.notes = retrieval_note.strip()
                item.save()
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"{request.user.first_name} retrieved complaint: {item.reference_number} and set status to Under Review"
                )
            elif item_type == 'scholarship':
                item = get_object_or_404(Scholarship, pk=pk, is_archived=True)

                if not (request.user.is_superuser or
                        request.user.user_type in [1, 5] or
                        item.archived_by == request.user):
                    return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

                item.is_archived = False
                item.archived_at = None
                item.archived_by = None
                item.save()

                # Log the activity
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"{request.user.first_name} retrieved scholarship: {item.name}"
                )

            elif item_type == 'scholarship-application':
                item = get_object_or_404(ScholarshipApplication, pk=pk, is_archived=True)

                if not (request.user.is_superuser or
                        request.user.user_type in [1, 5] or
                        item.archived_by == request.user):
                    return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

                item.is_archived = False
                item.archived_at = None
                item.archived_by = None
                item.save()

                # Log the activity
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"{request.user.first_name} retrieved scholarship application for {item.student.get_full_name()}"
                )
            elif item_type == 'admission':
                item = get_object_or_404(StudentAdmission, pk=pk, is_archived=True)

                # Check permissions
                if not (request.user.is_superuser or
                        request.user.user_type in [1, 12] or
                        item.archived_by == request.user):
                    return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

                item.is_archived = False
                item.archived_at = None
                item.archived_by = None
                item.save()

                # Log the activity
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"{request.user.first_name} retrieved student admission: {item.control_no}"
                )
            elif item_type == 'nstp-student':
                item = get_object_or_404(NSTPStudentInfo, pk=pk, is_archived=True)

                if not (request.user.is_superuser or
                        request.user.user_type in [1, 2] or
                        item.archived_by == request.user):
                    return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

                item.is_archived = False
                item.archived_at = None
                item.archived_by = None
                item.save()

                # Log the activity
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"{request.user.first_name} retrieved NSTP student: {item.last_name}, {item.first_name} ({item.student_number})"
                )
            elif item_type == 'nstp-file':
                item = get_object_or_404(NSTPFile, pk=pk, is_archived=True)

                if not (request.user.is_superuser or
                        item.created_by == request.user or
                        request.user.user_type in [1, 2] or
                        item.archived_by == request.user):
                    return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

                item.is_archived = False
                item.archived_at = None
                item.archived_by = None
                item.save()

                # Log the activity
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"{request.user.first_name} retrieved NSTP file: {item.title}"
                )
            elif item_type == 'ojt-company':
                item = get_object_or_404(OJTCompany, pk=pk, is_archived=True)

                # Check permissions - superuser or the user who archived it can retrieve
                if not (request.user.is_superuser or
                        request.user.user_type in [1, 13] or
                        item.archived_by == request.user):
                    return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

                item.is_archived = False
                item.archived_at = None
                item.archived_by = None
                item.save()

                # Log the activity
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"{request.user.first_name} retrieved OJT company: {item.name}"
                )
            elif item_type == 'organization':
                item = get_object_or_404(Organization, pk=pk, is_archived=True)
                if not (request.user.is_superuser or
                        request.user.user_type == 10 or  # Student Development Services
                        item.archived_by == request.user):
                    return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

                from django.db import transaction
                with transaction.atomic():
                    # Retrieve the organization
                    item.is_archived = False
                    item.archived_at = None
                    item.archived_by = None
                    item.is_active = True
                    item._organization_status = 'cancelled'

                    item.save()

                    # Also retrieve the linked user account if it exists
                    user_retrieved = False
                    if item.user_account:
                        user_account = item.user_account
                        user_account.is_archived = False
                        user_account.is_active = True
                        user_account.archived_at = None
                        user_account.archived_by = None
                        user_account.save()
                        user_retrieved = True

                # Log the activity
                if user_retrieved:
                    UserActivityLog.objects.create(
                        user=request.user,
                        activity=f"{request.user.first_name} retrieved organization: {item.organization_name} and its linked user account - Status set to CANCELLED"
                    )
                else:
                    UserActivityLog.objects.create(
                        user=request.user,
                        activity=f"{request.user.first_name} retrieved organization: {item.organization_name} - Status set to CANCELLED"
                    )
            elif item_type == 'accomplishment-report':
                item = get_object_or_404(AccomplishmentRecord, pk=pk, is_archived=True)

                if not (request.user.is_superuser or
                        request.user.user_type in [1, 10] or
                        (request.user.user_type == 15 and item.archived_by == request.user)):
                    return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

                item.is_archived = False
                item.archived_at = None
                item.archived_by = None
                item.archive_reason = None
                item.save()

                # Log the activity
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"{request.user.first_name} retrieved accomplishment report: {item.title}"
                )
            else:
                return JsonResponse({'success': False, 'error': 'Invalid item type'}, status=400)

            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ----------------------------------------------- Calendar Section -----------------------------------------------------
class AnnouncementCalendarAPI(LoginRequiredMixin, View):
    def get(self, request):
        # Get all published announcements
        announcements = Announcement.objects.filter(
            is_published=True,
            is_archived=False
        ).order_by('-created_at')

        announcements_data = []
        for announcement in announcements:
            created_at = timezone.localtime(announcement.created_at)

            # Base data for all announcements
            base_data = {
                'id': announcement.id,
                'title': announcement.title,
                'content': announcement.content,
                'category': announcement.category,
                'created_at': created_at.isoformat(),
                'author': announcement.author.get_full_name(),
            }

            # For Basic and Emergency show on creation date
            if announcement.category in ['BASIC', 'EMERGENCY']:
                announcements_data.append({
                    **base_data,
                    'date': created_at.strftime('%Y-%m-%d')
                })

            # Show on event date
            elif announcement.category == 'EVENT' and announcement.event_date:
                event_date = timezone.localtime(announcement.event_date)
                announcements_data.append({
                    **base_data,
                    'date': event_date.strftime('%Y-%m-%d'),
                    'event_date': event_date.strftime('%Y-%m-%d'),
                })

            # Show during enrollment period
            elif announcement.category == 'ENROLLMENT':
                if announcement.enrollment_start and announcement.enrollment_end:
                    current_date = announcement.enrollment_start.date()
                    end_date = announcement.enrollment_end.date()
                    while current_date <= end_date:
                        announcements_data.append({
                            **base_data,
                            'date': current_date.strftime('%Y-%m-%d'),
                            'enrollment_start': announcement.enrollment_start.strftime('%Y-%m-%d'),
                            'enrollment_end': announcement.enrollment_end.strftime('%Y-%m-%d'),
                        })
                        current_date += timedelta(days=1)

            # Show during suspension period
            elif announcement.category == 'SUSPENSION':
                if announcement.suspension_date:
                    current_date = announcement.suspension_date
                    end_date = announcement.until_suspension_date if announcement.until_suspension_date else announcement.suspension_date
                    while current_date <= end_date:
                        announcements_data.append({
                            **base_data,
                            'date': current_date.strftime('%Y-%m-%d'),
                            'suspension_date': current_date.strftime('%Y-%m-%d'),
                            'until_suspension_date': end_date.strftime(
                                '%Y-%m-%d') if announcement.until_suspension_date else None,
                        })
                        current_date += timedelta(days=1)

            # Show during scholarship application period
            elif announcement.category == 'SCHOLARSHIP':
                if announcement.application_start and announcement.application_end:
                    current_date = announcement.application_start.date()
                    end_date = announcement.application_end.date()
                    while current_date <= end_date:
                        announcements_data.append({
                            **base_data,
                            'date': current_date.strftime('%Y-%m-%d'),
                            'application_start': announcement.application_start.strftime('%Y-%m-%d'),
                            'application_end': announcement.application_end.strftime('%Y-%m-%d'),
                        })
                        current_date += timedelta(days=1)

        return JsonResponse(announcements_data, safe=False)


# ------------------------------------------------- Scholarship Section ------------------------------------------------
class ScholarshipCreateView(LoginRequiredMixin, CreateView):
    model = Scholarship
    form_class = ScholarshipForm
    template_name = 'osas/modals/create-scholarship.html'

    def form_valid(self, form):
        try:
            scholarship = form.save(commit=False)
            scholarship.created_by = self.request.user

            # Handle file upload by creating a Downloadable instance
            if 'application_form' in self.request.FILES:
                uploaded_file = self.request.FILES['application_form']

                # Create a new Downloadable instance
                downloadable = Downloadable.objects.create(
                    title=f"{form.cleaned_data['name']}",
                    description=f"Application form for {form.cleaned_data['name']} scholarship",
                    file=uploaded_file,
                    category='scholarship_forms',
                    is_active=True,
                    created_by=self.request.user
                )
                scholarship.application_form = downloadable

            scholarship.save()

            return JsonResponse({
                'success': True,
                'message': 'Scholarship created successfully!',
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)

    def form_invalid(self, form):
        errors = {}
        for field, error_list in form.errors.items():
            label = form.fields[field].label or field
            errors[field] = {
                'label': label,
                'messages': [str(error) for error in error_list]
            }

        return JsonResponse({
            'success': False,
            'errors': errors,
            'message': 'Please correct the errors below.'
        }, status=400)


class ScholarshipDetailView(LoginRequiredMixin, DetailView):
    model = Scholarship

    def get(self, request, *args, **kwargs):
        try:
            scholarship = self.get_object()

            # Handle application form data safely
            application_form = None
            if scholarship.application_form:
                application_form = {
                    'id': scholarship.application_form.id,
                    'title': scholarship.application_form.title,
                    'file_url': scholarship.application_form.file.url,
                    'file_name': scholarship.application_form.get_file_name(),
                    'file_size': scholarship.application_form.get_file_size(),
                }

            data = {
                'success': True,
                'scholarship': {
                    'id': scholarship.id,
                    'name': scholarship.name,
                    'description': scholarship.description,
                    'scholarship_type': scholarship.get_scholarship_type_display(),
                    'benefits': scholarship.benefits,
                    'requirements': scholarship.requirements,
                    'is_active': scholarship.is_active,
                    'created_at': scholarship.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': scholarship.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'created_by': scholarship.created_by.get_full_name() if scholarship.created_by else None,
                    'is_archived': scholarship.is_archived,
                    'archived_at': scholarship.archived_at.strftime('%Y-%m-%d %H:%M:%S') if scholarship.archived_at else None,
                    'archived_by': scholarship.archived_by.get_full_name() if scholarship.archived_by else None,
                    'application_form': application_form,
                }
            }
            return JsonResponse(data)

        except Scholarship.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Scholarship not found'}, status=404)
        except Exception as e:
            import traceback
            traceback.print_exc()  # Log the full traceback
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


class ScholarshipUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        try:
            scholarship = get_object_or_404(Scholarship, pk=pk)
            form = ScholarshipForm(request.POST, instance=scholarship)

            if form.is_valid():
                updated_scholarship = form.save()

                # Log the activity
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"{request.user.first_name} updated scholarship: {updated_scholarship.name}"
                )

                return JsonResponse({
                    'success': True,
                    'message': 'Scholarship updated successfully!',
                    'scholarship': {
                        'id': updated_scholarship.id,
                        'name': updated_scholarship.name,
                        'scholarship_type': updated_scholarship.get_scholarship_type_display(),
                        'description': updated_scholarship.description,
                        'benefits': updated_scholarship.benefits,
                        'requirements': updated_scholarship.requirements,
                        'is_active': updated_scholarship.is_active,
                    }
                })
            else:
                errors = {field: [str(error) for error in error_list] for field, error_list in form.errors.items()}
                return JsonResponse({
                    'success': False,
                    'errors': errors,
                    'message': 'Please correct the errors below.'
                }, status=400)

        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)


class ScholarshipArchiveView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not request.user.has_perm('osas.delete_scholarship'):
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to archive scholarships'
            }, status=403)

        try:
            scholarship = Scholarship.objects.get(id=pk)

            # Archive the scholarship
            scholarship.is_archived = True
            scholarship.archived_at = timezone.now()
            scholarship.archived_by = request.user
            scholarship.save()

            # Check if necessary it will Also archive all related applications (Uncomment once the application finish)
            #ScholarshipApplication.objects.filter(scholarship=scholarship).update(
                #is_archived=True,
                #archived_at=timezone.now(),
                #archived_by=request.user
            #)

            # Log the activity
            UserActivityLog.objects.create(
                user=request.user,
                activity=f"{request.user.first_name} archived scholarship: {scholarship.name}"
            )

            return JsonResponse({'success': True})

        except Scholarship.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Scholarship not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e) if str(e) else 'An error occurred while archiving the scholarship'
            }, status=500)


# --------------------------------------- Scholarship Application Section ----------------------------------------------
class ScholarshipApplicationApproveView(View):
    def post(self, request, *args, **kwargs):
        application = get_object_or_404(ScholarshipApplication, pk=kwargs['pk'])

        # Check permissions
        if not request.user.user_type in [1, 5]:
            return JsonResponse({
                'success': False,
                'message': 'You do not have permission to approve applications.'
            }, status=403)

        decision = request.POST.get('decision')
        notes = request.POST.get('notes', '')

        if decision not in ['approved', 'rejected']:
            return JsonResponse({
                'success': False,
                'message': 'Invalid decision.'
            }, status=400)

        # Update the application
        application.status = decision
        application.notes = notes
        application.status_updated_by = request.user
        application.status_update_date = timezone.now()
        application.save()

        # Send email notification
        self.send_decision_email(application, decision, notes)

        # Log the activity
        UserActivityLog.objects.create(
            user=request.user,
            activity=f"{request.user.get_full_name()} {decision} scholarship application {application.id}"
        )

        return JsonResponse({
            'success': True,
            'message': f'Application {decision} successfully.'
        })

    def send_decision_email(self, application, decision, notes):
        student = application.student
        subject = f"Your Scholarship Application for {application.scholarship.name} has been {decision}"

        # Prepare context for email template
        context = {
            'student_name': student.get_full_name(),
            'scholarship_name': application.scholarship.name,
            'decision': decision,
            'notes': notes,
            'status_update_date': application.status_update_date.strftime("%B %d, %Y"),
            'status_updated_by': application.status_updated_by.get_full_name(),
        }

        # Render HTML email template
        html_message = render_to_string('emails/scholarship_decision.html', context)
        plain_message = strip_tags(html_message)

        # Send email
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=None,
            recipient_list=[student.email],
            html_message=html_message,
            fail_silently=False,
        )


class ScholarshipApplicationView(LoginRequiredMixin, CreateView):
    model = ScholarshipApplication
    form_class = ScholarshipApplicationForm
    template_name = 'core/scholarship-application.html'

    def get_success_url(self):
        return reverse_lazy('dashboard') + '?tab=basic#scholarship-applications'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get active scholarships (no slot filtering needed anymore)
        active_scholarships = Scholarship.objects.filter(
            is_active=True,
            is_archived=False
        ).order_by('name')

        # Get existing applications with their scholarship IDs
        existing_applications = ScholarshipApplication.objects.filter(
            student=self.request.user,
            status__in=['pending', 'under_review', 'approved']
        ).select_related('scholarship')

        # Create a set of scholarship IDs the user has already applied for
        existing_scholarship_ids = {app.scholarship.id for app in existing_applications}

        context['active_scholarships'] = active_scholarships
        context['existing_scholarship_ids'] = existing_scholarship_ids
        context['existing_applications'] = existing_applications
        context['user'] = self.request.user

        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # Check for existing pending applications for the same scholarship
        existing_app = ScholarshipApplication.objects.filter(
            student=self.request.user,
            scholarship=form.cleaned_data['scholarship'],
            status__in=['pending', 'under_review', 'approved']
        ).first()

        if existing_app:
            messages.warning(self.request,
                             f"You already have an existing application for this scholarship (Status: {existing_app.get_status_display()}).")
            return redirect('scholarship_application_status', pk=existing_app.pk)

        form.instance.student = self.request.user
        response = super().form_valid(form)
        messages.success(self.request, "Your scholarship application has been submitted successfully!")
        return response

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.warning(request, "You need to login to access the scholarship application.")
            return redirect('home')
        return super().dispatch(request, *args, **kwargs)


class ScholarshipApplicationStatusView(DetailView):
    model = ScholarshipApplication
    template_name = 'core/scholarship-application-status.html'
    context_object_name = 'application'

    def get_object(self, queryset=None):
        # Get the application ID from the URL
        pk = self.kwargs.get('pk')

        application = get_object_or_404(ScholarshipApplication, pk=pk)

        # Check if the current user is allowed to view this application
        user = self.request.user
        if not (user.is_superuser or
                user.user_type in [1, 5, 14] or
                application.student == user):
            raise Http404("You don't have permission to view this application")

        return application

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        application = self.object

        # Add existing applications for the same student to context
        context['existing_applications'] = ScholarshipApplication.objects.filter(
            student=application.student
        ).order_by('-application_date')

        return context


class ScholarshipApplicationUpdateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        application = get_object_or_404(ScholarshipApplication, pk=kwargs['pk'])
        scholarships = Scholarship.objects.filter(is_active=True, is_archived=False)

        # Get student details
        student = application.student

        return JsonResponse({
            'success': True,
            'application': {
                'id': application.id,
                'student_id': student.id,
                'student_name': student.get_full_name(),
                'student_number': student.student_number,
                'course': student.course.name if student.course else None,
                'course_id': student.course.id if student.course else None,
                'year_level': student.year_level,
                'section': student.section,
                'scholarship_id': application.scholarship.id,
                'scholarship_name': application.scholarship.name,
                'application_date': application.application_date.strftime('%Y-%m-%d %H:%M'),
                'status': application.status,
                'notes': application.notes,
                'application_form_url': application.application_form.url,
                'application_form_name': application.application_form.name.split('/')[-1],
                'cog_url': application.cog.url,
                'cog_name': application.cog.name.split('/')[-1],
                'cor_url': application.cor.url,
                'cor_name': application.cor.name.split('/')[-1],
                'id_photo_url': application.id_photo.url,
                'id_photo_name': application.id_photo.name.split('/')[-1],
                'other_documents_url': application.other_documents.url if application.other_documents else None,
                'other_documents_name': application.other_documents.name.split('/')[
                    -1] if application.other_documents else None,
                'status_updated_by': application.status_updated_by.get_full_name() if application.status_updated_by else None,
                'status_update_date': application.status_update_date.strftime(
                    '%Y-%m-%d %H:%M') if application.status_update_date else None,
            },
            'scholarships': [{'id': s.id, 'name': s.name} for s in scholarships]
        })

    def post(self, request, *args, **kwargs):
        application = get_object_or_404(ScholarshipApplication, pk=kwargs['pk'])

        # Get previous values for logging
        previous_status = application.status
        previous_scholarship = application.scholarship

        # Create form instance with POST data and FILES
        form = ScholarshipApplicationEditForm(request.POST, request.FILES, instance=application)

        if form.is_valid():
            application = form.save(commit=False)

            new_scholarship = form.cleaned_data.get('scholarship')
            if new_scholarship and new_scholarship != previous_scholarship:
                if ScholarshipApplication.objects.filter(
                        student=application.student,
                        scholarship=new_scholarship
                ).exclude(pk=application.pk).exists():
                    return JsonResponse({
                        'success': False,
                        'message': 'This student already has an application for the selected scholarship program.'
                    }, status=400)

            if 'status' in form.changed_data:
                new_status = form.cleaned_data['status']

                application.status_updated_by = request.user
                application.status_update_date = timezone.now()

            try:
                application.save()
            except IntegrityError:
                return JsonResponse({
                    'success': False,
                    'message': 'This student already has an application for the selected scholarship program.'
                }, status=400)

            # Log changes
            changes = []
            if previous_status != application.status:
                changes.append(f"status from {previous_status} to {application.status}")
            if previous_scholarship != application.scholarship:
                changes.append(f"scholarship from {previous_scholarship.name} to {application.scholarship.name}")

            if changes:
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"{request.user.first_name} updated application {application.id}: " + ", ".join(changes)
                )

            return JsonResponse({
                'success': True,
                'application': {
                    'id': application.id,
                    'student_name': application.student.get_full_name(),
                    'scholarship_name': application.scholarship.name,
                    'status': application.status,
                    'status_display': application.get_status_display(),
                    'notes': application.notes,
                    'status_updated_by': application.status_updated_by.get_full_name() if application.status_updated_by else None,
                    'status_update_date': application.status_update_date.strftime(
                        '%Y-%m-%d %H:%M') if application.status_update_date else None,
                }
            })

        errors = {field: [str(error) for error in form.errors.get(field, [])] for field in form.errors}
        return JsonResponse({'success': False, 'errors': errors, 'message': 'Please correct the errors below.'},
                            status=400)


class ScholarshipApplicationArchiveView(View):
    def post(self, request, *args, **kwargs):
        try:
            # Get application ID from URL parameter
            application_id = self.kwargs.get('pk')

            # Get the application object
            application = get_object_or_404(ScholarshipApplication, id=application_id)

            # Archive the application with user and timestamp
            application.is_archived = True
            application.archived_at = timezone.now()
            application.archived_by = request.user
            application.save(update_fields=['is_archived', 'archived_at', 'archived_by'])

            # Add success message (if using messages framework)
            messages.success(request, 'Application archived successfully')

            return JsonResponse({
                'success': True,
                'message': 'Application archived successfully'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)


def export_scholarship_applications(request):
    if request.method == 'POST':
        try:
            print("Export function called")

            # Get export parameters
            export_option = request.POST.get('exportOption', 'all')
            include_incomplete = request.POST.get('include_incomplete') == '1'

            # Get current filters
            search_term = request.POST.get('search', '')
            scholarship_filter = request.POST.get('scholarship_filter', 'all')
            status_filter = request.POST.get('status_filter', 'all')
            date_filter = request.POST.get('date_filter', 'all')
            sort_column = request.POST.get('sort', 'application_date')
            sort_direction = request.POST.get('direction', 'desc')

            # Start with all applications (non-archived)
            applications = ScholarshipApplication.objects.filter(
                is_archived=False
            ).select_related(
                'student', 'scholarship', 'student__course'
            )

            # Apply filters based on export options
            if export_option == 'scholarship':
                scholarship_id = request.POST.get('scholarship')
                if scholarship_id and scholarship_id != '':
                    applications = applications.filter(scholarship_id=scholarship_id)
            elif export_option == 'scholarship-status':
                scholarship_id = request.POST.get('scholarship_status')
                status = request.POST.get('status_scholarship')
                if scholarship_id and scholarship_id != '':
                    applications = applications.filter(scholarship_id=scholarship_id)
                if status and status != 'all':
                    applications = applications.filter(status=status)

            # Apply date range if provided
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            if start_date:
                applications = applications.filter(application_date__date__gte=start_date)
            if end_date:
                applications = applications.filter(application_date__date__lte=end_date)

            # Apply search if provided
            if search_term:
                applications = applications.filter(
                    Q(student__first_name__icontains=search_term) |
                    Q(student__last_name__icontains=search_term) |
                    Q(scholarship__name__icontains=search_term)
                )

            # Apply additional filters if provided
            if scholarship_filter != 'all':
                applications = applications.filter(scholarship_id=scholarship_filter)

            if status_filter != 'all':
                applications = applications.filter(status=status_filter)

            # Apply date filter if provided
            if date_filter != 'all':
                now = timezone.now()
                if date_filter == 'today':
                    applications = applications.filter(application_date__date=now.date())
                elif date_filter == 'week':
                    start_of_week = now - timezone.timedelta(days=now.weekday())
                    applications = applications.filter(application_date__gte=start_of_week)
                elif date_filter == 'month':
                    applications = applications.filter(
                        application_date__month=now.month,
                        application_date__year=now.year
                    )

            # Apply sorting
            sort_field = sort_column
            if sort_column == 'student':
                sort_field = 'student__last_name'
            elif sort_column == 'scholarship':
                sort_field = 'scholarship__name'
            elif sort_column == 'date':
                sort_field = 'application_date'

            if sort_direction == 'desc':
                sort_field = f'-{sort_field}'

            applications = applications.order_by(sort_field)

            # Only filter incomplete if checkbox is unchecked
            if not include_incomplete:
                # Check for required fields
                applications = applications.filter(
                    student__last_name__isnull=False,
                    student__first_name__isnull=False,
                    student__student_number__isnull=False,
                    student__course__isnull=False,
                    student__year_level__isnull=False
                ).exclude(
                    Q(student__last_name='') |
                    Q(student__first_name='') |
                    Q(student__student_number='') |
                    Q(student__year_level='')
                )

            print(f"Final applications count: {applications.count()}")

            # Load the template Excel file
            template_path = os.path.join(settings.BASE_DIR, 'static', 'templates', 'ScholarshipTemplate.xlsx')

            if not os.path.exists(template_path):
                raise FileNotFoundError(f"Template file not found at: {template_path}")

            print(f"Loading template from: {template_path}")

            # Load the workbook using openpyxl
            workbook = load_workbook(template_path)

            # Get the 'Annex 1' sheet
            if 'Annex 1' not in workbook.sheetnames:
                raise ValueError("Template must contain 'Annex 1' sheet")

            sheet = workbook['Annex 1']

            # Data starts at row 13
            data_start_row = 13

            # Clear data from row 13 downwards
            max_rows_to_clear = data_start_row + applications.count() + 50

            for row in range(data_start_row, max_rows_to_clear + 1):
                # Clear columns A-M (1-13)
                for col in range(1, 14):
                    try:
                        cell = sheet.cell(row=row, column=col)
                        # Only clear regular cells, not merged cells
                        if cell.__class__.__name__ != 'MergedCell':
                            cell.value = None
                    except:
                        continue

            # Populate data starting from row 13
            successful_exports = 0

            # Set alignment style
            from openpyxl.styles import Alignment
            alignment = Alignment(horizontal='left', vertical='center')

            for index, app in enumerate(applications):
                row_num = data_start_row + index
                student = app.student

                try:
                    # A: SEQ (Sequence Number) - Column 1
                    cell = sheet.cell(row=row_num, column=1, value=index + 1)
                    cell.alignment = alignment

                    # B: STUDENT ID - Column 2
                    cell = sheet.cell(row=row_num, column=2, value=student.student_number or '')
                    cell.alignment = alignment

                    # C: LAST NAME - Column 3
                    cell = sheet.cell(row=row_num, column=3, value=student.last_name or '')
                    cell.alignment = alignment

                    # D: GIVEN NAME - Column 4
                    cell = sheet.cell(row=row_num, column=4, value=student.first_name or '')
                    cell.alignment = alignment

                    # E: SEX - Column 5
                    gender_map = {
                        'M': 'Male',
                        'F': 'Female',
                        'O': 'Other',
                        'N': ''
                    }
                    sex = gender_map.get(student.gender, '')
                    cell = sheet.cell(row=row_num, column=5, value=sex)
                    cell.alignment = alignment

                    # F: BIRTHDATE (dd/mm/yyyy) - Column 6
                    if student.birth_date:
                        try:
                            if isinstance(student.birth_date, str):
                                date_str = student.birth_date.split()[0] if ' ' in student.birth_date else student.birth_date
                                for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']:
                                    try:
                                        birthdate = datetime.strptime(date_str, fmt)
                                        break
                                    except:
                                        continue
                                else:
                                    birthdate = None
                            else:
                                birthdate = student.birth_date

                            if birthdate:
                                cell = sheet.cell(row=row_num, column=6, value=birthdate)
                                cell.number_format = 'DD/MM/YYYY'
                            else:
                                cell = sheet.cell(row=row_num, column=6, value='')
                        except Exception as e:
                            cell = sheet.cell(row=row_num, column=6, value='')
                    else:
                        cell = sheet.cell(row=row_num, column=6, value='')
                    cell.alignment = alignment

                    # G: COMPLETE PROGRAM NAME - Column 7
                    program_name = student.course.name if student.course else ''
                    cell = sheet.cell(row=row_num, column=7, value=program_name)
                    cell.alignment = alignment

                    # H: YEAR LEVEL - Column 8
                    year_level = student.year_level or ''
                    # Clean year level
                    if year_level:
                        import re
                        numbers = re.findall(r'\d+', str(year_level))
                        if numbers:
                            year_level = numbers[0]
                    cell = sheet.cell(row=row_num, column=8, value=year_level)
                    cell.alignment = alignment

                    # I: STREET & BARANGAY - Column 9
                    address = student.address or ''
                    cell = sheet.cell(row=row_num, column=9, value=address)
                    cell.alignment = alignment

                    # J: (Empty/Reserved) - Column 10
                    # K: (Empty/Reserved) - Column 11

                    # L: CONTACT NUMBER - Column 12
                    contact_number = student.phone_number or ''
                    cell = sheet.cell(row=row_num, column=12, value=contact_number)
                    cell.alignment = alignment

                    # M: EMAIL ADDRESS - Column 13
                    email = student.email or ''
                    cell = sheet.cell(row=row_num, column=13, value=email)
                    cell.alignment = alignment

                    successful_exports += 1

                except Exception as e:
                    print(f"Error processing application for {student.get_full_name()}: {e}")
                    continue

            # Create HTTP response with Excel file
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            # Generate filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Scholarship_Applications_{timestamp}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            # Save workbook to response
            workbook.save(response)

            print(f"Export completed with {successful_exports} applications")
            return response

        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"Export error: {str(e)}")
            return HttpResponse(f"Error during export: {str(e)}", status=500)

    return HttpResponse('Invalid request method. Use POST.', status=400)


# -------------------------------------------------- Admission Section -------------------------------------------------
class AdmissionApproveView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        admission = get_object_or_404(StudentAdmission, pk=kwargs['pk'])
        return JsonResponse({
            'success': True,
            'admission': {
                'id': admission.id,
                'control_no': admission.control_no,
                'student_type': admission.student_type,
                'student_type_display': admission.get_student_type_display(),
                'course': admission.course.name if admission.course else 'Not specified',
                'status': admission.status,
                'status_display': admission.get_status_display(),
                'first_name': admission.first_name,
                'last_name': admission.last_name,
                'strand': admission.strand,
            }
        })

    def post(self, request, *args, **kwargs):
        admission = get_object_or_404(StudentAdmission, pk=kwargs['pk'])

        # First update the status to check requirements
        admission.update_status()

        # If requirements aren't complete, return error
        if admission.status != 'complete':
            return JsonResponse({
                'success': False,
                'message': 'Cannot approve admission with incomplete requirements',
                'status': admission.status,
                'remarks': admission.remarks
            }, status=400)

        # Only proceed with approval if requirements are complete
        admission.status = 'done'
        notes = request.POST.get('notes', '')

        # Append admin notes to existing remarks if provided
        if notes:
            if admission.remarks:
                admission.remarks = f"{admission.remarks}\n\nAdmin Notes: {notes}"
            else:
                admission.remarks = f"Admin Notes: {notes}"

        admission.save()

        # Log the activity
        UserActivityLog.objects.create(
            user=request.user,
            activity=f"{request.user.get_full_name()} approved admission {admission.control_no}"
        )

        return JsonResponse({
            'success': True,
            'message': 'Admission approved successfully.',
            'status': 'done'
        })


class StudentAdmissionCreateView(LoginRequiredMixin, CreateView):
    model = StudentAdmission
    form_class = StudentAdmissionForm
    template_name = 'core/create-admission.html'
    success_url = reverse_lazy('admission_summary')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        if not form.cleaned_data.get('first_name'):
            form.instance.first_name = self.request.user.first_name
        if not form.cleaned_data.get('last_name'):
            form.instance.last_name = self.request.user.last_name

        form.instance.user = self.request.user
        form.instance.created_by = self.request.user

        response = super().form_valid(form)

        self.object.update_status()
        self.request.session['recent_admission_id'] = self.object.id

        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['courses'] = Course.objects.all()
        context['student_type_choices'] = StudentAdmission.STUDENT_TYPE_CHOICES
        context['curriculum_choices'] = StudentAdmission.CURRICULUM_CHOICES
        context['semester_status_choices'] = StudentAdmission.SEMESTER_STATUS_CHOICES
        return context

    def get_success_url(self):
        return reverse_lazy('admission_summary', kwargs={'pk': self.object.id})


class StudentAdmissionSummaryView(LoginRequiredMixin, DetailView):
    model = StudentAdmission
    template_name = 'core/view-admission.html'
    context_object_name = 'admission'

    def get_object(self, queryset=None):
        if 'pk' in self.kwargs:
            return super().get_object(queryset)
        return StudentAdmission.objects.get(pk=self.request.session.get('recent_admission_id'))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        admission = context['admission']

        # Create a list of document fields with their display names and status
        document_fields = [
            {'name': 'grade11_report_card', 'display': 'Grade 11 Report Card', 'file': admission.grade11_report_card},
            {'name': 'certificate_of_enrollment', 'display': 'Certificate of Enrollment',
             'file': admission.certificate_of_enrollment},
            {'name': 'grade12_report_card', 'display': 'Grade 12 Report Card', 'file': admission.grade12_report_card},
            {'name': 'form137', 'display': 'Form 137', 'file': admission.form137},
            {'name': 'transcript_of_grades', 'display': 'Transcript of Grades', 'file': admission.transcript_of_grades},
            {'name': 'good_moral_certificate', 'display': 'Good Moral Certificate',
             'file': admission.good_moral_certificate},
            {'name': 'honorable_dismissal', 'display': 'Honorable Dismissal', 'file': admission.honorable_dismissal},
            {'name': 'nbi_police_clearance', 'display': 'NBI/Police Clearance', 'file': admission.nbi_police_clearance},
        ]

        context['document_fields'] = document_fields
        context['full_name'] = f"{admission.first_name} {admission.last_name}"
        context['course_name'] = admission.course.name if admission.course else "No course selected"
        return context


class AdmissionDetailView(LoginRequiredMixin, DetailView):
    model = StudentAdmission

    def get_queryset(self):
        return super().get_queryset().select_related('course', 'user')

    def get(self, request, *args, **kwargs):
        try:
            admission = self.get_object()

            # Prepare file data
            def get_file_data(file_field):
                if file_field:
                    return {
                        'url': file_field.url,
                        'name': file_field.name.split('/')[-1],
                        'size': self.format_file_size(file_field.size)
                    }
                return None

            # Get student name if user is associated
            student_name = None
            if admission.user:
                student_name = {
                    'first_name': admission.user.first_name,
                    'last_name': admission.user.last_name,
                    'full_name': f"{admission.user.last_name}, {admission.user.first_name}"
                }

            # Prepare course data
            course_data = None
            if admission.course:
                course_data = {
                    'id': admission.course.id,
                    'name': admission.course.name,
                    'subtext': admission.course.subtext,
                    'full_display': f"{admission.course.name} ({admission.course.subtext})" if admission.course.subtext else admission.course.name
                }

            data = {
                'success': True,
                'admission': {
                    'id': admission.id,
                    'control_no': admission.control_no,
                    'student_name': student_name,
                    'student_type': admission.get_student_type_display(),
                    'student_type_code': admission.student_type,
                    'course': course_data,
                    'status': admission.get_status_display(),
                    'status_code': admission.status,
                    'remarks': admission.remarks,
                    'created_at': admission.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': admission.updated_at.strftime('%Y-%m-%d %H:%M:%S'),

                    # Common fields
                    'admission_portal_registration': admission.admission_portal_registration,
                    'strand': admission.strand,

                    # Current Grade 12 files
                    'grade11_report_card': get_file_data(admission.grade11_report_card),
                    'certificate_of_enrollment': get_file_data(admission.certificate_of_enrollment),

                    # SHS Graduate files
                    'grade12_report_card': get_file_data(admission.grade12_report_card),
                    'form137': get_file_data(admission.form137),

                    # Transferee fields
                    'curriculum_type': admission.get_curriculum_type_display() if admission.curriculum_type else None,
                    'first_year_first_semester': admission.get_first_year_first_semester_display() if admission.first_year_first_semester else None,
                    'first_year_second_semester': admission.get_first_year_second_semester_display() if admission.first_year_second_semester else None,
                    'second_year_first_semester': admission.get_second_year_first_semester_display() if admission.second_year_first_semester else None,
                    'other_semester_info': admission.other_semester_info,
                    'transcript_of_grades': get_file_data(admission.transcript_of_grades),
                    'good_moral_certificate': get_file_data(admission.good_moral_certificate),
                    'honorable_dismissal': get_file_data(admission.honorable_dismissal),
                    'nbi_police_clearance': get_file_data(admission.nbi_police_clearance),
                }
            }
            return JsonResponse(data)

        except StudentAdmission.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Admission application not found'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    def format_file_size(self, size):
        """Convert file size to human-readable format"""
        if not size:
            return "0 KB"
        for unit in ['bytes', 'KB', 'MB', 'GB']:
            if size < 1024.0:
                return f"{size:.1f} {unit}"
            size /= 1024.0
        return f"{size:.1f} TB"


class AdmissionUpdateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        admission = get_object_or_404(StudentAdmission, pk=kwargs['pk'])
        courses = Course.objects.all()

        # Get required fields based on student type
        required_fields = admission.get_required_fields()
        requirements_status = []

        for field in required_fields:
            if field == 'admission_portal_registration':
                requirements_status.append({
                    'name': 'Admission Portal Registration',
                    'completed': getattr(admission, field),
                    'field': field
                })
            else:
                # For file fields
                file_field = getattr(admission, field)
                field_info = {
                    'name': admission._meta.get_field(field).verbose_name,
                    'completed': bool(file_field),
                    'field': field
                }

                # Only add file info if this is a FileField/ImageField
                field_type = admission._meta.get_field(field).get_internal_type()
                if field_type in ('FileField', 'ImageField'):
                    field_info.update({
                        'file_url': file_field.url if file_field else None,
                        'file_name': os.path.basename(file_field.name) if file_field else None
                    })

                requirements_status.append(field_info)

        # Prepare course data
        course_data = None
        if admission.course:
            course_data = {
                'id': admission.course.id,
                'name': admission.course.name
            }

        return JsonResponse({
            'success': True,
            'admission': {
                'id': admission.id,
                'control_no': admission.control_no,
                'student_type': admission.student_type,
                'student_type_display': admission.get_student_type_display(),
                'course': course_data,
                'status': admission.status,
                'status_display': admission.get_status_display(),
                'admission_portal_registration': admission.admission_portal_registration,
                'remarks': admission.remarks,
                'created_at': admission.created_at.strftime('%Y-%m-%d %H:%M'),
                'date': admission.date.strftime('%Y-%m-%d') if admission.date else '',
                'strand': admission.strand,
                'is_archived': admission.is_archived,
                'curriculum_type': admission.curriculum_type,
                'first_year_first_semester': admission.first_year_first_semester,
                'first_year_second_semester': admission.first_year_second_semester,
                'second_year_first_semester': admission.second_year_first_semester,
                'other_semester_info': admission.other_semester_info,
                'first_name': admission.first_name,
                'last_name': admission.last_name,
                'requirements_status': requirements_status
            },
            'courses': [{'id': course.id, 'name': course.name} for course in courses]
        })

    def post(self, request, *args, **kwargs):
        admission = get_object_or_404(StudentAdmission, pk=kwargs['pk'])

        # Update basic fields
        admission.control_no = request.POST.get('control_no')
        admission.student_type = request.POST.get('student_type')

        # Handle course field - get Course instance
        course_id = request.POST.get('course')
        if course_id:
            try:
                admission.course = Course.objects.get(id=course_id)
            except Course.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid course selected'
                }, status=400)
        else:
            admission.course = None

        admission.status = request.POST.get('status')
        admission.remarks = request.POST.get('remarks')
        admission.strand = request.POST.get('strand')
        admission.admission_portal_registration = request.POST.get('admission_portal_registration') == 'true'
        admission.is_archived = request.POST.get('is_archived') == 'true'

        # Handle date field
        date_str = request.POST.get('date')
        if date_str:
            try:
                admission.date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid date format'
                }, status=400)

        # Handle transferee specific fields
        if admission.student_type == 'transferee':
            admission.curriculum_type = request.POST.get('curriculum_type')
            admission.first_year_first_semester = request.POST.get('first_year_first_semester')
            admission.first_year_second_semester = request.POST.get('first_year_second_semester')
            admission.second_year_first_semester = request.POST.get('second_year_first_semester')
            admission.other_semester_info = request.POST.get('other_semester_info')

        # Handle file uploads/removals
        required_fields = admission.get_required_fields()

        for field in required_fields:
            if field == 'admission_portal_registration':
                continue

            # Check if this is a file field
            field_type = admission._meta.get_field(field).get_internal_type()
            if field_type in ('FileField', 'ImageField'):
                # Check if file should be removed
                if f'remove_{field}' in request.POST and request.POST[f'remove_{field}'] == 'true':
                    current_file = getattr(admission, field)
                    if current_file:
                        current_file.delete(save=False)
                    setattr(admission, field, None)

                # Handle new file upload
                if field in request.FILES:
                    # Delete old file if exists
                    old_file = getattr(admission, field)
                    if old_file:
                        old_file.delete(save=False)
                    setattr(admission, field, request.FILES[field])

        # Save the admission
        try:
            admission.save()
            admission.update_status()
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error saving admission: {str(e)}'
            }, status=400)

        # Log the activity
        UserActivityLog.objects.create(
            user=request.user,
            activity=f"{request.user.get_full_name()} updated admission application {admission.control_no}"
        )

        return JsonResponse({
            'success': True,
            'message': 'Admission updated successfully',
            'admission': {
                'id': admission.id,
                'control_no': admission.control_no,
                'status': admission.status,
                'status_display': admission.get_status_display(),
                'student_type_display': admission.get_student_type_display(),
                'course': {
                    'id': admission.course.id if admission.course else None,
                    'name': admission.course.name if admission.course else None
                },
                'remarks': admission.remarks,
                'updated_at': admission.updated_at.strftime('%Y-%m-%d %H:%M')
            }
        })


class StudentAdmissionArchiveView(View):
    def get(self, request, *args, **kwargs):
        try:
            admission_id = self.kwargs.get('pk')
            admission = get_object_or_404(StudentAdmission, id=admission_id)

            data = {
                'success': True,
                'admission': {
                    'control_no': admission.control_no,
                    'student_name': f"{admission.last_name}, {admission.first_name}",
                    'student_type': admission.get_student_type_display(),
                    'course': str(admission.course),
                    'status': admission.get_status_display(),
                    'created_at': admission.created_at.strftime("%b %d, %Y"),
                    'remarks': admission.remarks or '-'
                }
            }
            return JsonResponse(data)

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)

    def post(self, request, *args, **kwargs):
        """Existing archive functionality"""
        try:
            admission_id = self.kwargs.get('pk')
            admission = get_object_or_404(StudentAdmission, id=admission_id)
            admission.is_archived = True
            admission.archived_at = timezone.now()
            admission.archived_by = request.user
            admission.save(update_fields=['is_archived', 'archived_at', 'archived_by'])

            return JsonResponse({
                'success': True,
                'message': 'Archived successfully',
                'archived_at': admission.archived_at.strftime("%b %d, %Y %H:%M"),
                'archived_by': request.user.get_full_name() or request.user.username
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)


def export_admissions(request):
    if request.method == 'POST':
        try:
            export_option = request.POST.get('export_option', 'all')
            file_format = request.POST.get('format', 'excel')
            search_term = request.POST.get('search', '')
            student_type_filter = request.POST.get('student_type_filter', '')

            applications = StudentAdmission.objects.filter(is_archived=False).select_related('user', 'course')

            # Apply filters based on export options
            if export_option == 'filtered':
                # Apply search if provided
                if search_term:
                    applications = applications.filter(
                        Q(control_no__icontains=search_term) |
                        Q(user__first_name__icontains=search_term) |
                        Q(user__last_name__icontains=search_term)
                    )

                # Apply student type filter if provided
                if student_type_filter:
                    applications = applications.filter(student_type=student_type_filter)

            elif export_option == 'custom':
                # Apply custom filters from the form
                student_type = request.POST.get('student_type', '')
                status = request.POST.get('status', '')

                if student_type:
                    applications = applications.filter(student_type=student_type)
                if status:
                    applications = applications.filter(status=status)

            # Apply date range if provided
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')

            if start_date:
                applications = applications.filter(created_at__gte=start_date)
            if end_date:
                # Add one day to include the entire end date
                try:
                    end_date_obj = timezone.datetime.strptime(end_date, '%Y-%m-%d')
                    end_date_obj = end_date_obj + timezone.timedelta(days=1)
                    applications = applications.filter(created_at__lte=end_date_obj)
                except ValueError:
                    # Handle invalid date format
                    pass

            # Check if Excel format is selected
            if file_format == 'excel':
                # Load the template
                template_path = os.path.join(settings.STATICFILES_DIRS[0], 'templates', 'AdmissionTemplate.xlsx')

                # If staticfiles_dirs doesn't work, try alternative path
                if not os.path.exists(template_path):
                    template_path = os.path.join(settings.BASE_DIR, 'static', 'templates', 'AdmissionTemplate.xlsx')

                if not os.path.exists(template_path):
                    return HttpResponse("Template file not found", status=500)

                # Load workbook
                wb = load_workbook(template_path)
                ws = wb.active

                max_row = ws.max_row
                for row in range(8, max_row + 1):
                    for col in range(1, 9):  # Columns A to H (1-8)
                        ws.cell(row=row, column=col, value=None)

                # Start writing data at row 8
                start_row = 8

                # Write admission data to the template
                for i, app in enumerate(applications, start=start_row):
                    # Get student name
                    student_name = ''
                    if app.user:
                        if app.user.last_name and app.user.first_name:
                            student_name = f"{app.user.last_name}, {app.user.first_name}"
                        elif app.last_name and app.first_name:
                            student_name = f"{app.last_name}, {app.first_name}"

                    # Get email
                    email = app.user.email if app.user and app.user.email else ''

                    # Get course name
                    course_name = str(app.course) if app.course else ''

                    ws.cell(row=i, column=1, value=app.control_no)  # Control No.
                    ws.cell(row=i, column=2, value=student_name)  # Student Name
                    ws.cell(row=i, column=3, value=app.get_student_type_display())  # Student Type
                    ws.cell(row=i, column=4, value=course_name)  # Course
                    ws.cell(row=i, column=5, value=email)  # Email
                    ws.cell(row=i, column=6, value=app.get_status_display())  # Status

                    # Date Submitted
                    if app.created_at:
                        ws.cell(row=i, column=7, value=app.created_at.strftime('%Y-%m-%d %H:%M'))
                    else:
                        ws.cell(row=i, column=7, value='')

                    # Remarks
                    ws.cell(row=i, column=8, value=app.remarks or '')  # Remarks

                # Save workbook to a BytesIO object
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                # Create response with template
                response = HttpResponse(
                    output.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="admissions_export.xlsx"'

                return response

            elif file_format == 'csv':
                # Prepare data for CSV export (original logic)
                data = []
                for app in applications:
                    # Get phone number with fallback
                    phone = ''
                    if app.user and hasattr(app.user, 'phone'):
                        phone = app.user.phone

                    data.append({
                        'Control No': app.control_no,
                        'Student Name': f"{app.user.last_name}, {app.user.first_name}" if app.user else '',
                        'Student Type': app.get_student_type_display(),
                        'Course': str(app.course) if app.course else '',
                        'Status': app.get_status_display(),
                        'Remarks': app.remarks or '',
                        'Date Submitted': app.created_at.strftime('%Y-%m-%d %H:%M'),
                        'Email': app.user.email if app.user and app.user.email else '',
                        'Phone': phone
                    })

                # Create DataFrame
                df = pd.DataFrame(data)

                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename="admissions_export.csv"'
                df.to_csv(response, index=False, encoding='utf-8')

                return response

        except Exception as e:
            # Log the error for debugging
            import traceback
            error_trace = traceback.format_exc()
            print(f"Export error: {str(e)}")
            print(f"Traceback: {error_trace}")
            return HttpResponse(f"Error during export: {str(e)}", status=500)

    return HttpResponse('Invalid request', status=400)


# ------------------------------------------------------- NSTP Section -------------------------------------------------
class NSTPApproveView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        enlistment = get_object_or_404(NSTPStudentInfo, pk=kwargs['pk'])
        return JsonResponse({
            'success': True,
            'enlistment': {
                'id': enlistment.id,
                'student_number': enlistment.student_number,
                'first_name': enlistment.first_name,
                'last_name': enlistment.last_name,
                'program': enlistment.program,
                'semester': enlistment.get_semester_display(),
                'academic_year': enlistment.academic_year,
                'approval_status': enlistment.approval_status,
                'approval_status_display': enlistment.get_approval_status_display(),
                'remarks': enlistment.remarks or '',
            }
        })

    def post(self, request, *args, **kwargs):
        enlistment = get_object_or_404(NSTPStudentInfo, pk=kwargs['pk'])
        action = request.POST.get('action', 'approve')
        notes = request.POST.get('notes', '')

        if action == 'approve':
            enlistment.approval_status = 'approved'
            status_message = 'approved'
        elif action == 'reject':
            enlistment.approval_status = 'rejected'
            status_message = 'rejected'
        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid action specified'
            }, status=400)

        # Save admin notes
        if notes:
            enlistment.remarks = notes

        enlistment.save()

        # Log the activity
        UserActivityLog.objects.create(
            user=request.user,
            activity=f"{request.user.get_full_name()} {status_message} NSTP enlistment for {enlistment.student_number}"
        )

        return JsonResponse({
            'success': True,
            'message': f'NSTP enlistment {status_message} successfully.',
            'approval_status': enlistment.approval_status
        })


class NSTPStudentInfoCreateView(LoginRequiredMixin, CreateView):
    model = NSTPStudentInfo
    form_class = NSTPStudentInfoForm
    template_name = 'core/create-nstp.html'

    def get_success_url(self):
        return reverse('nstp_success', kwargs={'pk': self.object.pk})

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def post(self, request, *args, **kwargs):
        print("POST data:", request.POST)  # Debugging
        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        if not self.request.user.is_authenticated:
            form.add_error(None, "You must be logged in to submit this form.")
            return self.form_invalid(form)

        form.instance.user = self.request.user

        try:
            # Additional validation if needed
            if form.cleaned_data.get('academic_year') and '-' not in form.cleaned_data['academic_year']:
                form.add_error('academic_year', "Academic year must be in format YYYY-YYYY")
                return self.form_invalid(form)

            self.object = form.save(commit=False)
            self.object.user = self.request.user

            # Check for existing registration for this semester/year
            existing = NSTPStudentInfo.objects.filter(
                user=self.request.user,
                semester=self.object.semester,
                academic_year=self.object.academic_year
            ).exists()

            if existing:
                form.add_error(None, "You already have a registration for this semester and academic year.")
                return self.form_invalid(form)

            self.object.save()

            if hasattr(form, 'save_m2m'):  # Only if there are M2M fields
                form.save_m2m()

            messages.success(self.request, 'NSTP information submitted successfully!')
            return HttpResponseRedirect(self.get_success_url())

        except ValidationError as e:
            form.add_error(None, e)
            return self.form_invalid(form)
        except Exception as e:
            import traceback
            traceback.print_exc()
            form.add_error(None, f"An error occurred: {str(e)}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        print("Form errors:", form.errors)  # Debugging
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'NSTP Student Information Form'

        if self.request.user.is_authenticated:
            context['existing_applications'] = NSTPStudentInfo.objects.filter(
                user=self.request.user
            ).values('semester', 'academic_year')

        return context


class NSTPDetailSuccessView(DetailView): #Used as Summary or Success page after registration
    model = NSTPStudentInfo
    template_name = 'core/nstp_detail.html'
    context_object_name = 'enlistment'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'NSTP Registration Details'
        return context

    def get_object(self, queryset=None):
        # Get the specific registration by PK or the most recent one
        if 'pk' in self.kwargs:
            return super().get_object(queryset)
        return NSTPStudentInfo.objects.filter(user=self.request.user).latest('created_at')


class NSTPStudentDetailView(LoginRequiredMixin, DetailView):
    model = NSTPStudentInfo

    def get(self, request, *args, **kwargs):
        try:
            nstp_student = self.get_object()

            # Prepare data structure
            data = {
                'success': True,
                'nstp_student': {
                    'id': nstp_student.id,
                    'student_number': nstp_student.student_number,
                    'full_name': f"{nstp_student.last_name}, {nstp_student.first_name}",
                    'first_name': nstp_student.first_name,
                    'last_name': nstp_student.last_name,
                    'middle_name': nstp_student.middle_name,
                    'program': nstp_student.program,
                    'gender': nstp_student.get_gender_display(),
                    'gender_code': nstp_student.gender,
                    'birth_date': nstp_student.birth_date.strftime('%Y-%m-%d') if nstp_student.birth_date else None,
                    'contact_number': nstp_student.contact_number,
                    'email_address': nstp_student.email_address,

                    # Address information
                    'street_or_barangay': nstp_student.street_or_barangay,
                    'municipality_or_city': nstp_student.municipality_or_city,
                    'province': nstp_student.province,
                    'full_address': self.format_full_address(nstp_student),

                    # NSTP information
                    'semester': nstp_student.get_semester_display(),
                    'semester_code': nstp_student.semester,
                    'academic_year': nstp_student.academic_year,

                    # Approval status and remarks
                    'approval_status': nstp_student.get_approval_status_display(),
                    'approval_status_code': nstp_student.approval_status,
                    'remarks': nstp_student.remarks,
                    'is_archived': nstp_student.is_archived,
                    'archived_at': nstp_student.archived_at.strftime(
                        '%Y-%m-%d %H:%M:%S') if nstp_student.archived_at else None,

                    # Timestamps
                    'created_at': nstp_student.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': nstp_student.updated_at.strftime('%Y-%m-%d %H:%M:%S'),

                    # User reference
                    'user_id': nstp_student.user.id if nstp_student.user else None,
                    'username': nstp_student.user.username if nstp_student.user else None,
                    'archived_by': nstp_student.archived_by.username if nstp_student.archived_by else None,
                }
            }
            return JsonResponse(data)

        except NSTPStudentInfo.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'NSTP student record not found'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    def format_full_address(self, nstp_student):
        parts = []
        if nstp_student.street_or_barangay:
            parts.append(nstp_student.street_or_barangay)
        if nstp_student.municipality_or_city:
            parts.append(nstp_student.municipality_or_city)
        if nstp_student.province:
            parts.append(nstp_student.province)
        return ', '.join(parts) if parts else 'No address provided'


class NSTPStudentUpdateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        nstp_student = get_object_or_404(NSTPStudentInfo, pk=kwargs['pk'])

        response_data = {
            'success': True,
            'nstp_student': {
                'id': nstp_student.id,
                'student_number': nstp_student.student_number,
                'last_name': nstp_student.last_name,
                'first_name': nstp_student.first_name,
                'middle_name': nstp_student.middle_name,
                'program': nstp_student.program,
                'gender': nstp_student.gender,
                'birth_date': nstp_student.birth_date.strftime('%Y-%m-%d') if nstp_student.birth_date else '',
                'contact_number': nstp_student.contact_number,
                'email_address': nstp_student.email_address,
                'street_or_barangay': nstp_student.street_or_barangay,
                'municipality_or_city': nstp_student.municipality_or_city,
                'province': nstp_student.province,
                'semester': nstp_student.semester,
                'academic_year': nstp_student.academic_year,
                'approval_status': nstp_student.approval_status,
                'remarks': nstp_student.remarks,
                'is_archived': nstp_student.is_archived,
                'created_at': nstp_student.created_at.strftime('%Y-%m-%d %H:%M'),
                'updated_at': nstp_student.updated_at.strftime('%Y-%m-%d %H:%M'),
            },
            'hide_approval_section': request.user.user_type == 14  # Student user type
        }

        return JsonResponse(response_data)

    def post(self, request, *args, **kwargs):
        nstp_student = get_object_or_404(NSTPStudentInfo, pk=kwargs['pk'])

        # Update basic fields
        nstp_student.student_number = request.POST.get('student_number')
        nstp_student.last_name = request.POST.get('last_name')
        nstp_student.first_name = request.POST.get('first_name')
        nstp_student.middle_name = request.POST.get('middle_name')
        nstp_student.program = request.POST.get('program')
        nstp_student.gender = request.POST.get('gender')
        nstp_student.contact_number = request.POST.get('contact_number')
        nstp_student.email_address = request.POST.get('email_address')
        nstp_student.street_or_barangay = request.POST.get('street_or_barangay')
        nstp_student.municipality_or_city = request.POST.get('municipality_or_city')
        nstp_student.province = request.POST.get('province')
        nstp_student.semester = request.POST.get('semester')
        nstp_student.academic_year = request.POST.get('academic_year')
        nstp_student.remarks = request.POST.get('remarks', '')
        nstp_student.is_archived = request.POST.get('is_archived') == 'true'

        # Only update approval status if user is not a student (type 14)
        if request.user.user_type != 14:
            nstp_student.approval_status = request.POST.get('approval_status', 'pending')

        # Handle birth date
        birth_date_str = request.POST.get('birth_date')
        if birth_date_str:
            nstp_student.birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()

        nstp_student.save()

        # Log the activity
        UserActivityLog.objects.create(
            user=request.user,
            activity=f"{request.user.first_name} updated NSTP enlistment for {nstp_student.student_number}"
        )

        return JsonResponse({
            'success': True,
            'nstp_student': {
                'id': nstp_student.id,
                'student_number': nstp_student.student_number,
                'full_name': f"{nstp_student.last_name}, {nstp_student.first_name}",
                'program': nstp_student.program,
                'semester': nstp_student.semester,
                'academic_year': nstp_student.academic_year,
                'approval_status': nstp_student.approval_status,
                'updated_at': nstp_student.updated_at.strftime('%Y-%m-%d %H:%M')
            }
        })


class NSTPEnlistmentArchiveView(View):
    def get(self, request, *args, **kwargs):
        try:
            enlistment_id = self.kwargs.get('pk')
            enlistment = get_object_or_404(NSTPStudentInfo, id=enlistment_id)
            program = enlistment.program or '-'
            created_at = enlistment.created_at.strftime("%b %d, %Y") if enlistment.created_at else 'Date not available'

            data = {
                'success': True,
                'enlistment': {
                    'student_number': enlistment.student_number or '-',
                    'program': program,
                    'semester': enlistment.get_semester_display(),
                    'created_at': created_at,
                }
            }
            return JsonResponse(data)

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)

    def post(self, request, *args, **kwargs):
        try:
            enlistment_id = self.kwargs.get('pk')
            if not enlistment_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Missing enlistment ID'
                }, status=400)

            enlistment = get_object_or_404(NSTPStudentInfo, id=enlistment_id)

            if enlistment.is_archived:
                return JsonResponse({
                    'success': False,
                    'error': 'This enlistment is already archived'
                }, status=400)

            enlistment.is_archived = True
            enlistment.archived_at = timezone.now()
            enlistment.archived_by = request.user
            enlistment.save(update_fields=['is_archived', 'archived_at', 'archived_by'])

            return JsonResponse({
                'success': True,
                'message': 'NSTP enlistment archived successfully',
                'archived_at': enlistment.archived_at.strftime("%b %d, %Y %H:%M"),
                'archived_by': request.user.get_full_name() or request.user.username
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)


class NSTPExportTemplateView(View):
    def post(self, request, *args, **kwargs):
        # Get filter parameters from request
        search_term = request.POST.get('search', '')
        semester_filter = request.POST.get('semester', '')
        academic_year = request.POST.get('academic_year', '')
        export_semester = request.POST.get('export_semester', '')
        export_title = request.POST.get('export_title', 'Title')
        include_search = request.POST.get('include_search', 'false') == 'true'

        # Filter NSTP records
        nstp_records = NSTPStudentInfo.objects.filter(is_archived=False)

        # Apply filters if provided
        if academic_year:
            nstp_records = nstp_records.filter(academic_year=academic_year)
        if export_semester:
            nstp_records = nstp_records.filter(semester=export_semester)

        # Apply search filters if include_search is checked
        if include_search:
            if search_term:
                nstp_records = nstp_records.filter(
                    Q(student_number__icontains=search_term) |
                    Q(last_name__icontains=search_term) |
                    Q(first_name__icontains=search_term) |
                    Q(program__icontains=search_term)
                )
            if semester_filter:
                nstp_records = nstp_records.filter(semester=semester_filter)

        # Get template path
        template_path = os.path.join(settings.BASE_DIR, 'static', 'templates', 'NSTPTemplate.xlsx')

        try:
            # Load the template workbook
            wb = load_workbook(template_path)
            ws = wb["NSTP EL"]

            # Set title and academic year/semester
            ws['A5'] = export_title  # Title
            if export_semester and academic_year:
                ws['A6'] = f"{export_semester}, Academic Year: {academic_year}"
            elif export_semester:
                ws['A6'] = f"{export_semester}"
            elif academic_year:
                ws['A6'] = f"Academic Year: {academic_year}"
            else:
                ws['A6'] = "All Records"

            ws['C8'] = "Cavite State University - Bacoor City Campus"  # Name of HEI
            ws['C9'] = "Lily St., Phase 2, Soldiers Hills IV, Molino VI, Bacoor City, Cavite"  # Address

            ws['K8'] = "Region:"
            ws['K8'].font = Font(bold=True)
            ws['L8'] = "IV (CALABARZON)"  # Default region value

            ws['K9'] = "NSTP Component:"
            ws['K9'].font = Font(bold=True)
            ws['L9'] = "CWTS"  # Default NSTP component value

            # Define styles for new rows
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Fill student data starting from row 13
            start_row = 13
            max_template_rows = 45  # Original template rows
            total_records = len(nstp_records)

            if total_records > max_template_rows:
                rows_to_add = total_records - max_template_rows
                ws.insert_rows(start_row + max_template_rows, amount=rows_to_add)

                # Apply formatting to new rows
                for row in range(start_row + max_template_rows, start_row + max_template_rows + rows_to_add):
                    for col in range(1, 14):  # Columns A to M
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border
                        if col in [1, 6, 7]:  # Center align for No., Program, Sex
                            cell.alignment = Alignment(horizontal='center')

            # Fill all student records
            for idx, record in enumerate(nstp_records, start=1):
                row = start_row + idx - 1

                ws[f'A{row}'] = idx  # No.
                ws[f'B{row}'] = record.student_number
                ws[f'C{row}'] = record.last_name
                ws[f'D{row}'] = record.first_name
                ws[f'E{row}'] = record.middle_name or ''
                ws[f'F{row}'] = record.program or ''
                ws[f'G{row}'] = record.get_gender_display() if record.gender else ''
                ws[f'H{row}'] = record.birth_date.strftime('%m/%d/%Y') if record.birth_date else ''
                ws[f'I{row}'] = record.street_or_barangay or ''
                ws[f'J{row}'] = record.municipality_or_city or ''
                ws[f'K{row}'] = record.province or ''
                ws[f'L{row}'] = record.contact_number or ''
                ws[f'M{row}'] = record.email_address or ''

            # Set optimized column widths
            col_widths = {
                'A': 5,   # No.
                'B': 15,  # Student Number
                'C': 20,  # Last Name (Surname)
                'D': 15,  # First Name
                'E': 15,  # Middle Name
                'F': 25,  # Program
                'G': 8,   # Sex
                'H': 12,  # Birthdate
                'I': 25,  # Street/Barangay
                'J': 20,  # Municipality/City
                'K': 15,  # Province
                'L': 15,  # Contact Number
                'M': 25   # Email Address
            }

            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width

            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = "NSTP_Enrollment"
            if export_semester:
                filename += f"_{export_semester.replace(' ', '_')}"
            if academic_year:
                filename += f"_{academic_year.replace('-', '_')}"
            if not export_semester and not academic_year:
                filename += "_All_Records"
            filename += ".xlsx"

            response['Content-Disposition'] = f'attachment; filename={filename}'
            wb.save(response)

            return response

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class NSTPFileCreateView(LoginRequiredMixin, CreateView):
    model = NSTPFile
    form_class = NSTPFileForm
    template_name = 'osas/modals/create-nstp-files.html'

    def form_valid(self, form):
        try:
            nstp_file = form.save(commit=False)
            nstp_file.created_by = self.request.user
            nstp_file.save()

            return JsonResponse({
                'success': True,
                'message': 'NSTP file uploaded successfully!',
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)

    def form_invalid(self, form):
        errors = {}
        for field, error_list in form.errors.items():
            label = form.fields[field].label or field
            errors[field] = {
                'label': label,
                'messages': [str(error) for error in error_list]
            }

        return JsonResponse({
            'success': False,
            'errors': errors,
            'message': 'Please correct the errors below.'
        }, status=400)


class NSTPFileDetailView(LoginRequiredMixin, DetailView):
    model = NSTPFile

    def get(self, request, *args, **kwargs):
        try:
            nstp_file = self.get_object()

            data = {
                'success': True,
                'file': {
                    'id': nstp_file.id,
                    'title': nstp_file.title,
                    'description': nstp_file.description,
                    'category_display': nstp_file.get_category_display(),
                    'category_value': nstp_file.category,
                    'semester_display': nstp_file.get_semester_display(),
                    'semester_value': nstp_file.semester,
                    'school_year': nstp_file.school_year,
                    'file_url': nstp_file.file.url,
                    'file_name': nstp_file.get_file_name(),
                    'file_size': nstp_file.file.size,
                    'file_type': self.get_file_type(nstp_file.file.name),
                    'created_at': nstp_file.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': nstp_file.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'created_by': nstp_file.created_by.get_full_name() if nstp_file.created_by else None,
                }
            }
            return JsonResponse(data)

        except NSTPFile.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'NSTP file not found'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    def get_file_type(self, filename):
        extension = filename.split('.')[-1].lower()
        if extension in ['pdf']:
            return 'pdf'
        elif extension in ['doc', 'docx']:
            return 'word'
        elif extension in ['xls', 'xlsx']:
            return 'excel'
        elif extension in ['ppt', 'pptx']:
            return 'powerpoint'
        elif extension in ['jpg', 'jpeg', 'png']:
            return 'image'
        else:
            return 'file'


class NSTPFileUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        try:
            nstp_file = get_object_or_404(NSTPFile, pk=pk)
            form = NSTPFileForm(request.POST, request.FILES, instance=nstp_file)

            if form.is_valid():
                updated_file = form.save()

                return JsonResponse({
                    'success': True,
                    'message': 'File updated successfully!',
                    'file': {
                        'id': updated_file.id,
                        'title': updated_file.title,
                        'description': updated_file.description,
                        'category': updated_file.get_category_display(),
                        'semester': updated_file.get_semester_display(),
                        'school_year': updated_file.school_year,
                        'file_name': updated_file.get_file_name(),
                    }
                })
            else:
                errors = {field: errors for field, errors in form.errors.items()}
                return JsonResponse({
                    'success': False,
                    'errors': errors,
                    'message': 'Please correct the errors below'
                }, status=400)

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


class NSTPFileArchiveView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not (request.user.is_superuser or
                request.user.user_type in [1, 2] or
                request.user.has_perm('nstp.delete_nstpfile')):
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to archive NSTP files'
            }, status=403)

        try:
            nstp_file = NSTPFile.objects.get(id=pk)

            # Archive the file
            nstp_file.is_archived = True
            nstp_file.archived_at = timezone.now()
            nstp_file.archived_by = request.user
            nstp_file.save()

            # Log the activity
            UserActivityLog.objects.create(
                user=request.user,
                activity=f"{request.user.first_name} archived NSTP file: {nstp_file.title}"
            )

            return JsonResponse({'success': True})

        except NSTPFile.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'NSTP file not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e) if str(e) else 'An error occurred while archiving the NSTP file'
            }, status=500)


# ----------------------------------------------- OJT Company Section --------------------------------------------------
class CompanyCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = OJTCompany
    form_class = OJTCompanyForm
    success_url = reverse_lazy('ojt-company')

    def test_func(self):
        return self.request.user.is_superuser or self.request.user.user_type in [1, 13, 16]

    def handle_no_permission(self):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to create companies.'
            }, status=403)
        return super().handle_no_permission()

    def form_valid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            company = form.save()

            # Log the activity
            UserActivityLog.objects.create(
                user=self.request.user,
                activity=f"{self.request.user.first_name} created new company: {company.name}"
            )

            return JsonResponse({
                'success': True,
                'company': {
                    'id': company.id,
                    'name': company.name,
                    'address': company.address,
                    'contact_number': company.contact_number,
                }
            })

        return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = {field: [str(error) for error in errors] for field, errors in form.errors.items()}
            return JsonResponse({'success': False, 'errors': errors}, status=400)
        return super().form_invalid(form)


class OJTCompanyDetailView(LoginRequiredMixin, DetailView):
    model = OJTCompany

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'company': {
                    'id': self.object.id,
                    'name': self.object.name,
                    'address': self.object.address,
                    'contact_number': self.object.contact_number,
                    'email': self.object.email,
                    'website': self.object.website,
                    'description': self.object.description,
                    'status': self.object.status,
                    'is_archived': self.object.is_archived,
                    'created_at': self.object.created_at.strftime('%Y-%m-%d %H:%M'),
                    'updated_at': self.object.updated_at.strftime('%Y-%m-%d %H:%M'),
                }
            })
        return super().get(request, *args, **kwargs)


class OJTCompanyUpdateView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.is_superuser or self.request.user.user_type in [1, 13, 16]

    def handle_no_permission(self):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to edit companies.'
            }, status=403)
        return super().handle_no_permission()

    def get(self, request, *args, **kwargs):
        company = get_object_or_404(OJTCompany, pk=kwargs['pk'])
        return JsonResponse({
            'success': True,
            'company': {
                'id': company.id,
                'name': company.name,
                'address': company.address,
                'contact_number': company.contact_number,
                'status': company.status,
                'is_archived': company.is_archived,
                'description': company.description or '',
                'email': company.email or '',
                'website': company.website or '',
                'created_at': company.created_at.strftime('%Y-%m-%d %H:%M'),
                'updated_at': company.updated_at.strftime('%Y-%m-%d %H:%M'),
                'status_updated_at': company.status_updated_at.strftime(
                    '%Y-%m-%d %H:%M') if company.status_updated_at else None,
                'status_updated_by': company.status_updated_by.get_full_name() if company.status_updated_by else None,
            }
        })

    def post(self, request, *args, **kwargs):
        company = get_object_or_404(OJTCompany, pk=kwargs['pk'])

        # Create form with request data and instance
        form_data = request.POST.copy()
        form = OJTCompanyForm(form_data, instance=company)

        if form.is_valid():
            # Get original status before saving
            original_status = company.status

            # Save the company with the request context for status tracking
            company = form.save(commit=False)

            # Check if status has changed
            if original_status != form.cleaned_data['status']:
                company.status_updated_at = timezone.now()
                company.status_updated_by = request.user

            company.save()

            # Log the activity including status change if applicable
            log_message = f"{request.user.first_name} updated company: {company.name}"
            if original_status != company.status:
                log_message += f" (Status changed from {original_status} to {company.status})"

            UserActivityLog.objects.create(
                user=request.user,
                activity=log_message
            )

            return JsonResponse({
                'success': True,
                'company': {
                    'id': company.id,
                    'name': company.name,
                    'address': company.address,
                    'contact_number': company.contact_number,
                    'status': company.status,
                    'is_archived': company.is_archived,
                    'description': company.description or '',
                    'email': company.email or '',
                    'website': company.website or '',
                },
                'status_changed': original_status != company.status,
                'message': 'Company updated successfully'
            })

        errors = {field: [str(error) for error in errors] for field, errors in form.errors.items()}
        return JsonResponse({'success': False, 'errors': errors}, status=400)


class OJTCompanyArchiveView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.is_superuser or self.request.user.user_type in [1, 13, 16]

    def handle_no_permission(self):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to archive companies.'
            }, status=403)
        return super().handle_no_permission()

    def get_object(self):
        return get_object_or_404(OJTCompany, pk=self.kwargs['pk'])

    def post(self, request, *args, **kwargs):
        company = self.get_object()

        # Check if company is already archived
        if company.is_archived:
            return JsonResponse({
                'success': False,
                'error': f'Company "{company.name}" is already archived.'
            }, status=400)

        try:
            # Archive the company
            company.is_archived = True
            company.archived_at = timezone.now()
            company.archived_by = request.user
            company.save()

            # Log the activity
            UserActivityLog.objects.create(
                user=request.user,
                activity=f"{request.user.first_name} archived company: {company.name}"
            )

            return JsonResponse({
                'success': True,
                'message': f'Company "{company.name}" has been archived successfully.',
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Failed to archive company: {str(e)}'
            }, status=500)


class OJTCompanyExportView(View):
    def post(self, request, *args, **kwargs):
        try:
            # Get export parameters - try multiple parameter names
            export_option = request.POST.get('export_option', 'all')

            # Debug: Print all POST data
            print("=== OJT Export Debug ===")
            print("All POST data:", dict(request.POST))
            print("Export option from form:", export_option)

            # Check for the new form structure
            if 'export_type' in request.POST:
                export_type = request.POST.get('export_type', 'all')
                print("Export type:", export_type)

                if export_type == 'by_status':
                    if 'status_filter' in request.POST:
                        status_filter = request.POST.get('status_filter', 'active')
                        print("Status filter:", status_filter)
                        export_option = status_filter  # 'active' or 'inactive'
                else:
                    export_option = 'all'

            print("Final export option:", export_option)

            export_title = request.POST.get('export_title', 'OJT Companies Report')

            # Load template file
            template_path = os.path.join(settings.BASE_DIR, 'static', 'templates', 'OJTTemplate.xlsx')

            if not os.path.exists(template_path):
                return JsonResponse(
                    {'error': 'Export template not found. Please ensure OJTTemplate.xlsx exists in static/templates/'},
                    status=500)

            wb = load_workbook(template_path)
            ws = wb['Sheet1']

            # Update report title in cell A2
            ws['A2'] = export_title

            # Get companies based on export option
            # IMPORTANT: We need to handle three different cases:
            # 1. All active/inactive companies (non-archived)
            # 2. Only active companies (non-archived and status=active)
            # 3. Only inactive companies (non-archived and status=inactive)

            print(f"Export option: {export_option}")

            if export_option == 'all':
                # Get ALL non-archived companies (both active and inactive status)
                companies = OJTCompany.objects.filter(is_archived=False)
                print(f"Getting all non-archived companies (both active and inactive)")

            elif export_option == 'active':
                # Get only active AND non-archived companies
                companies = OJTCompany.objects.filter(
                    is_archived=False,
                    status='active'
                )
                print(f"Getting active non-archived companies")

            elif export_option == 'inactive':
                # Get only inactive AND non-archived companies
                companies = OJTCompany.objects.filter(
                    is_archived=False,
                    status='inactive'
                )
                print(f"Getting inactive non-archived companies")
            else:
                # Fallback: get all non-archived
                companies = OJTCompany.objects.filter(is_archived=False)
                print(f"Unknown export option, falling back to all non-archived")

            companies = companies.order_by('name')
            companies_count = companies.count()

            print(f"Found {companies_count} companies for export")

            # Debug: Print sample companies
            for i, company in enumerate(companies[:3]):
                print(f"Company {i + 1}: {company.name}, Status: {company.status}, Archived: {company.is_archived}")

            # Data starts at row 7
            data_start_row = 7

            # Clear only the data area (rows 7 and below for columns A-G)
            max_rows_to_clear = data_start_row + companies_count + 50

            for row in range(data_start_row, max_rows_to_clear + 1):
                for col in range(1, 8):  # Columns A-G
                    try:
                        cell = ws.cell(row=row, column=col)
                        # Only clear if it's not a merged cell
                        if cell.value is not None or cell.has_style:
                            cell.value = None
                    except:
                        continue

            # Populate data
            for idx, company in enumerate(companies):
                row_num = data_start_row + idx

                # Map data to template columns
                ws.cell(row=row_num, column=1, value=company.name or "")
                ws.cell(row=row_num, column=2, value=company.address or "")
                ws.cell(row=row_num, column=3, value=company.contact_number or "")
                ws.cell(row=row_num, column=4, value=company.email or "")
                ws.cell(row=row_num, column=5, value=company.description or "")

                # Status - use actual status field, not display_status
                # The display_status property shows "Archived" if is_archived=True
                # But we're filtering out archived companies, so we can use the status field
                status_display = company.get_status_display() if company.status else ""
                ws.cell(row=row_num, column=6, value=status_display)

                ws.cell(row=row_num, column=7, value=company.website or "")

            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            # Generate filename
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            if export_option == 'all':
                status_text = "All"
            elif export_option == 'active':
                status_text = "Active"
            elif export_option == 'inactive':
                status_text = "Inactive"
            else:
                status_text = "All"

            filename = f"OJT_Companies_{status_text}_{timestamp}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            wb.save(response)
            print(f"Export completed: {filename}, {companies_count} companies")
            return response

        except Exception as e:
            print(f"Error exporting OJT companies: {str(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({'error': f'An unexpected error occurred: {str(e)}'}, status=500)


# ---------------------------------------------- Organization Section --------------------------------------------------
class OrganizationCreateView(LoginRequiredMixin, CreateView):
    model = Organization
    form_class = OrganizationCreateForm
    template_name = 'osas/dashboard.html'
    success_url = reverse_lazy('organization-management')

    def post(self, request, *args, **kwargs):
        print("DEBUG: OrganizationCreateView POST received")
        form = self.get_form()

        if not form.is_valid():
            print(f"DEBUG: Form is invalid. Errors: {form.errors}")

            # Convert form errors to a simpler structure
            errors = {}
            for field, error_list in form.errors.items():
                if field == '__all__':
                    errors[field] = [str(error) for error in error_list]
                else:
                    errors[field] = [str(error) for error in error_list]

            return JsonResponse({
                'success': False,
                'message': 'Please correct the errors below.',
                'errors': errors
            }, status=400)

        return self.form_valid(form)

    def form_valid(self, form):
        try:
            print("DEBUG: Form is valid, saving...")
            organization = form.save(commit=False)

            raw_password = form.cleaned_data.get('password')
            print(f"DEBUG: Raw password from form: {raw_password}")

            if not raw_password:
                return JsonResponse({
                    'success': False,
                    'message': 'Password is required',
                    'errors': {'password': ['Password is required']}
                }, status=400)

            # Create user account
            user = CustomUser.objects.create_user(
                username=organization.username,
                email=organization.email,
                password=raw_password,
                user_type=15,
                first_name=organization.organization_name,
                is_active=True,
                is_verified=True,
                is_staff=True
            )

            print(f"DEBUG: User created with ID: {user.id}")

            # Set organization password and link user account
            organization.set_password(raw_password)
            organization.user_account = user

            # Handle organization members from the hidden JSON field
            members_json = self.request.POST.get('organization_members_json')
            print(f"DEBUG: Received members JSON: {members_json}")

            if members_json:
                try:
                    members_data = json.loads(members_json)
                    print(f"DEBUG: Parsed members data: {members_data}")

                    # Validate we have at least 3 members
                    if len(members_data) < 3:
                        return JsonResponse({
                            'success': False,
                            'message': 'Organization must have at least 3 members.',
                            'errors': {'organization_members_json': ['Organization must have at least 3 members.']}
                        }, status=400)

                    organization.organization_members = members_data
                    print(f"DEBUG: Successfully set organization members")

                except json.JSONDecodeError as e:
                    print(f"DEBUG: JSON decode error: {e}")
                    return JsonResponse({
                        'success': False,
                        'message': 'Invalid members data format.',
                        'errors': {'organization_members_json': ['Invalid members data format.']}
                    }, status=400)

            # Save the organization
            organization.save()

            print(f"DEBUG: Organization saved with ID: {organization.id}")

            # Log the activity
            UserActivityLog.objects.create(
                user=self.request.user,
                activity=f"Created new organization: {organization.organization_name} (ID: {organization.id}) - Type: {organization.organization_type}"
            )

            return JsonResponse({
                'success': True,
                'message': 'Organization created successfully!',
                'organization_id': organization.id
            })

        except Exception as e:
            print(f"DEBUG: Error saving organization: {str(e)}")
            import traceback
            print(traceback.format_exc())
            return JsonResponse({
                'success': False,
                'message': f'Error creating organization: {str(e)}',
                'errors': {'__all__': [str(e)]}
            }, status=400)

@require_POST
@csrf_exempt
def check_username_availability(request):
    try:
        data = json.loads(request.body)
        username = data.get('username', '').strip()

        if not username:
            return JsonResponse({
                'available': False,
                'message': 'Username is required'
            })

        if len(username) < 3:
            return JsonResponse({
                'available': False,
                'message': 'Username must be at least 3 characters long'
            })

        # Check if username contains only allowed characters
        import re
        if not re.match(r'^[a-zA-Z0-9_\.]+$', username):
            return JsonResponse({
                'available': False,
                'message': 'Username can only contain letters, numbers, underscores, and periods'
            })

        # Check both models for username availability
        username_exists_in_organization = Organization.objects.filter(username__iexact=username).exists()
        username_exists_in_customuser = CustomUser.objects.filter(username__iexact=username).exists()

        if username_exists_in_organization or username_exists_in_customuser:
            return JsonResponse({
                'available': False,
                'message': 'This username is already taken'
            })

        return JsonResponse({
            'available': True,
            'message': 'Username is available'
        })

    except Exception as e:
        print(f"Error checking username: {str(e)}")
        return JsonResponse({
            'available': False,
            'message': 'Error checking username availability'
        }, status=500)

@require_POST
@csrf_exempt
def check_email_availability(request):
    try:
        data = json.loads(request.body)
        email = data.get('email', '').strip().lower()

        if not email:
            return JsonResponse({
                'available': False,
                'message': 'Email is required'
            })

        # Basic email format validation
        import re
        email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_regex, email):
            return JsonResponse({
                'available': False,
                'message': 'Please enter a valid email address'
            })

        # Check ONLY CustomUser model for email (personal email)
        if CustomUser.objects.filter(email__iexact=email).exists():
            return JsonResponse({
                'available': False,
                'message': 'This email is already registered'
            })

        return JsonResponse({
            'available': True,
            'message': 'Email is available'
        })

    except Exception as e:
        print(f"Error checking email: {str(e)}")
        return JsonResponse({
            'available': False,
            'message': 'Error checking email availability'
        }, status=500)

@require_POST
@csrf_exempt
def check_organization_email_availability(request):
    try:
        data = json.loads(request.body)
        email = data.get('email', '').strip().lower()

        if not email:
            return JsonResponse({
                'available': False,
                'message': 'Organization email is required'
            })

        # Basic email format validation
        import re
        email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_regex, email):
            return JsonResponse({
                'available': False,
                'message': 'Please enter a valid email address'
            })

        # Check ONLY Organization model for organization email
        if Organization.objects.filter(organization_email__iexact=email).exists():
            return JsonResponse({
                'available': False,
                'message': 'This organization email is already registered'
            })

        return JsonResponse({
            'available': True,
            'message': 'Organization email is available'
        })

    except Exception as e:
        print(f"Error checking organization email: {str(e)}")
        return JsonResponse({
            'available': False,
            'message': 'Error checking organization email availability'
        }, status=500)


class OrganizationDetailView(LoginRequiredMixin, DetailView):
    model = Organization

    def get_queryset(self):
        return Organization.objects.select_related('organization_approved_by')

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Get organization data
            organization_data = {
                'id': self.object.id,
                'organization_name': self.object.organization_name,
                'organization_acronym': self.object.organization_acronym,
                'organization_type': self.object.organization_type,
                'organization_type_display': self.object.get_organization_type_display(),
                'organization_status': self.object.organization_status,  # Use the property
                'organization_status_display': self.get_status_display(self.object.organization_status),
                'organization_email': self.object.organization_email,
                'username': self.object.username,
                'organization_description': self.object.organization_description,
                'organization_mission': self.object.organization_mission,
                'organization_vision': self.object.organization_vision,

                # Adviser information
                'organization_adviser_name': self.object.organization_adviser_name,
                'organization_adviser_department': self.object.organization_adviser_department,
                'organization_adviser_email': self.object.organization_adviser_email,
                'organization_adviser_phone': self.object.organization_adviser_phone,

                # Co-Adviser information (NEW)
                'organization_coadviser_name': self.object.organization_coadviser_name,
                'organization_coadviser_department': self.object.organization_coadviser_department,
                'organization_coadviser_email': self.object.organization_coadviser_email,
                'organization_coadviser_phone': self.object.organization_coadviser_phone,

                # Dates and validity
                'organization_valid_from': self.object.organization_valid_from.strftime(
                    '%Y-%m-%d') if self.object.organization_valid_from else None,
                'organization_valid_until': self.object.organization_valid_until.strftime(
                    '%Y-%m-%d') if self.object.organization_valid_until else None,
                'current_school_year': self.object.current_school_year,
                'organization_member_count': self.object.organization_member_count,

                # Approval information
                'organization_approved_by': self.object.organization_approved_by.get_full_name() if self.object.organization_approved_by else None,
                'organization_approved_at': self.object.organization_approved_at.strftime(
                    '%Y-%m-%d %H:%M') if self.object.organization_approved_at else None,
                'organization_rejection_reason': self.object.organization_rejection_reason,

                # Logo URL
                'organization_logo_url': self.object.organization_logo.url if self.object.organization_logo else None,

                # Members data
                'organization_members': self.object.organization_members or [],

                # Documents data
                'documents': self.get_documents_data(),
                'total_documents': self.get_total_documents_count(),

                # Requirements status
                'all_requirements_submitted': self.object.all_requirements_submitted,
                'organization_has_minimum_members': self.object.organization_has_minimum_members,
                'can_be_approved': self.object.can_be_approved,
            }

            return JsonResponse({
                'success': True,
                'organization': organization_data
            })
        return super().get(request, *args, **kwargs)

    def get_status_display(self, status):
        status_map = {
            'active': 'Active',
            'pending': 'Pending Registration',
            'inactive': 'Inactive',
            'expired': 'Expired',
            'rejected': 'Rejected',
            'cancelled': 'Cancelled',
        }
        return status_map.get(status, status)

    def get_documents_data(self):
        """Get all organization documents with their status"""
        documents = []

        # Define all document fields
        document_fields = [
            ('organization_calendar_activities', 'Calendar of Activities'),
            ('organization_logo', 'Organization Logo'),
            ('organization_adviser_cv', 'Adviser CV'),
            ('organization_cog', 'Certificate of Grades'),
            ('organization_group_picture', 'Group Picture'),
            ('organization_cbl', 'Constitution and By-Laws'),
            ('organization_list_members', 'List of Members'),
            ('organization_acceptance_letter', 'Acceptance Letter'),
            ('organization_ar', 'Accomplishment Report'),
            ('organization_previous_calendar', 'Previous Calendar'),
            ('organization_financial_report', 'Financial Report'),
            ('organization_coa', 'Certificate of Assessment'),
            ('organization_member_biodata', 'Member Biodata'),
            ('organization_good_moral', 'Good Moral Certificate'),
        ]

        for field_name, display_name in document_fields:
            field = getattr(self.object, field_name)
            if field:
                documents.append({
                    'name': display_name,
                    'file_name': field.name.split('/')[-1],
                    'file_url': field.url,
                    'file_size': field.size,
                    'uploaded_at': field.uploaded_at.strftime('%Y-%m-%d %H:%M') if hasattr(field,
                                                                                           'uploaded_at') else None,
                    'field_name': field_name
                })

        return documents

    def get_total_documents_count(self):
        """Count total uploaded documents"""
        document_fields = [
            'organization_calendar_activities', 'organization_logo', 'organization_adviser_cv',
            'organization_cog', 'organization_group_picture', 'organization_cbl',
            'organization_list_members', 'organization_acceptance_letter', 'organization_ar',
            'organization_previous_calendar', 'organization_financial_report', 'organization_coa',
            'organization_member_biodata', 'organization_good_moral'
        ]

        count = 0
        for field_name in document_fields:
            if getattr(self.object, field_name):
                count += 1
        return count


class OrganizationEditView(LoginRequiredMixin, UpdateView):
    model = Organization
    form_class = OrganizationEditForm
    template_name = 'osas/modals/edit-organization.html'

    def get_queryset(self):
        return Organization.objects.select_related('organization_approved_by', 'user_account')

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            organization_data = {
                'id': self.object.id,
                'username': self.object.username,
                'email': self.object.email,
                'organization_name': self.object.organization_name,
                'organization_acronym': self.object.organization_acronym,
                'organization_type': self.object.organization_type,
                'organization_email': self.object.organization_email,
                'organization_description': self.object.organization_description,
                'organization_mission': self.object.organization_mission,
                'organization_vision': self.object.organization_vision,
                'organization_adviser_name': self.object.organization_adviser_name,
                'organization_adviser_department': self.object.organization_adviser_department,
                'organization_adviser_email': self.object.organization_adviser_email,
                'organization_adviser_phone': self.object.organization_adviser_phone,
                'organization_coadviser_name': self.object.organization_coadviser_name,
                'organization_coadviser_department': self.object.organization_coadviser_department,
                'organization_coadviser_email': self.object.organization_coadviser_email,
                'organization_coadviser_phone': self.object.organization_coadviser_phone,
                'organization_valid_from': self.object.organization_valid_from.strftime(
                    '%Y-%m-%d') if self.object.organization_valid_from else None,
                'organization_valid_until': self.object.organization_valid_until.strftime(
                    '%Y-%m-%d') if self.object.organization_valid_until else None,
                'organization_members': self.object.organization_members or [],
                'documents': self.get_documents_data(),
            }

            return JsonResponse({
                'success': True,
                'organization': organization_data
            })
        return super().get(request, *args, **kwargs)

    def get_documents_data(self):
        documents = []

        # Define all document fields
        document_fields = [
            ('organization_calendar_activities', 'Calendar of Activities'),
            ('organization_logo', 'Organization Logo'),
            ('organization_adviser_cv', 'Adviser CV'),
            ('organization_cog', 'Certificate of Grades'),
            ('organization_group_picture', 'Group Picture'),
            ('organization_cbl', 'Constitution and By-Laws'),
            ('organization_list_members', 'List of Members'),
            ('organization_acceptance_letter', 'Acceptance Letter'),
            ('organization_ar', 'Accomplishment Report'),
            ('organization_previous_calendar', 'Previous Calendar'),
            ('organization_financial_report', 'Financial Report'),
            ('organization_coa', 'Certificate of Assessment'),
            ('organization_member_biodata', 'Member Biodata'),
            ('organization_good_moral', 'Good Moral Certificate'),
        ]

        for field_name, display_name in document_fields:
            field = getattr(self.object, field_name)
            if field:
                documents.append({
                    'name': display_name,
                    'file_name': field.name.split('/')[-1],
                    'file_url': field.url,
                    'file_size': field.size,
                    'uploaded_at': field.uploaded_at.strftime('%Y-%m-%d %H:%M') if hasattr(field,
                                                                                           'uploaded_at') else None,
                    'field_name': field_name
                })

        return documents

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            if form.is_valid():
                return self.form_valid(form)
            else:
                return self.form_invalid(form)
        else:
            return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        try:
            organization = form.save(commit=False)

            # Get the raw password from the form if provided
            raw_password = form.cleaned_data.get('password')

            # Update the linked user account
            if organization.user_account:
                user = organization.user_account

                # Store old values for logging
                old_username = user.username
                old_email = user.email

                # Update user fields from organization
                user.username = organization.username
                user.email = organization.email
                user.first_name = organization.organization_name

                # Update password if provided
                if raw_password:
                    print(f"DEBUG: Updating password for user {user.username}")
                    user.set_password(raw_password)

                user.save()
                print(f"DEBUG: Updated user account: {old_username} -> {user.username}")
            else:
                print(f"DEBUG: Creating missing user account for organization")
                user = CustomUser.objects.create_user(
                    username=organization.username,
                    email=organization.email,
                    password=raw_password if raw_password else 'default_password_123',
                    user_type=15,
                    first_name=organization.organization_name,
                    is_active=True,
                    is_verified=True,
                    is_staff=True
                )
                organization.user_account = user
                print(f"DEBUG: Created user account: {user.username}")

            # Handle organization members from the hidden JSON field
            members_json = self.request.POST.get('organization_members_json')
            print(f"DEBUG: Received members JSON: {members_json}")

            if members_json:
                try:
                    members_data = json.loads(members_json)
                    print(f"DEBUG: Parsed members data: {members_data}")

                    # Validate and update organization members
                    if self.validate_members_data(members_data):
                        organization.organization_members = members_data
                        print(f"DEBUG: Successfully updated organization members")
                    else:
                        return JsonResponse({
                            'success': False,
                            'message': 'Invalid members data. Please check all member fields.'
                        }, status=400)

                except json.JSONDecodeError as e:
                    print(f"DEBUG: JSON decode error: {e}")
                    return JsonResponse({
                        'success': False,
                        'message': 'Invalid members data format.'
                    }, status=400)

            # Handle file removals
            removed_files = []
            for field_name in self.request.POST:
                if field_name.startswith('remove_') and self.request.POST[field_name] == 'true':
                    actual_field_name = field_name.replace('remove_', '')
                    setattr(organization, actual_field_name, None)
                    removed_files.append(actual_field_name)

            # Handle file uploads
            uploaded_files = []
            for field_name, file_obj in self.request.FILES.items():
                if hasattr(organization, field_name):
                    setattr(organization, field_name, file_obj)
                    uploaded_files.append(field_name)

            # Save the organization
            organization.save()
            print(f"DEBUG: Organization saved successfully with {len(organization.organization_members or [])} members")

            # Log the edit action to UserActivityLog
            activity_details = []
            if form.changed_data:
                activity_details.append(f"Updated fields: {', '.join(form.changed_data)}")
            if removed_files:
                activity_details.append(f"Removed files: {', '.join(removed_files)}")
            if uploaded_files:
                activity_details.append(f"Uploaded files: {', '.join(uploaded_files)}")
            if members_json:
                activity_details.append(f"Updated {len(organization.organization_members or [])} members")
            if raw_password:
                activity_details.append("Password updated")

            activity_message = f"Edited organization: {organization.organization_name} (ID: {organization.id})"
            if activity_details:
                activity_message += f" - {'; '.join(activity_details)}"

            UserActivityLog.objects.create(
                user=self.request.user,
                activity=activity_message
            )

            return JsonResponse({
                'success': True,
                'message': 'Organization updated successfully!',
                'organization_id': organization.id
            })

        except Exception as e:
            import traceback
            print(f"Error updating organization: {str(e)}")
            print(traceback.format_exc())
            return JsonResponse({
                'success': False,
                'message': f'Error updating organization: {str(e)}'
            }, status=400)

    def validate_members_data(self, members_data):
        if not isinstance(members_data, list):
            return False

        # ONLY REQUIRE AT LEAST 3 MEMBERS
        if len(members_data) < 3:
            return False

        for member in members_data:
            if not member.get('first_name') or not member.get('last_name'):
                return False
            if not member.get('position'):
                return False

        return True

    def form_invalid(self, form):
        errors = {}
        for field, error_list in form.errors.items():
            errors[field] = [str(error) for error in error_list]

        return JsonResponse({
            'success': False,
            'message': 'Please correct the errors below.',
            'errors': errors
        }, status=400)


class OrganizationArchiveView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.is_superuser or self.request.user.user_type in [10] or self.request.user.has_perm(
            'osas.delete_organization')

    def handle_no_permission(self):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to cancel organizations.'
            }, status=403)
        return super().handle_no_permission()

    def get_object(self):
        return get_object_or_404(Organization, pk=self.kwargs['pk'])

    def post(self, request, *args, **kwargs):
        organization = self.get_object()

        # Check if organization is already cancelled
        if organization._organization_status == 'cancelled':
            return JsonResponse({
                'success': False,
                'error': f'Organization "{organization.organization_name}" is already cancelled.'
            }, status=400)

        # Get cancellation reason from form
        archive_reason = request.POST.get('archive_reason', '').strip()

        if not archive_reason:
            return JsonResponse({
                'success': False,
                'error': 'Cancellation reason is required.'
            }, status=400)

        try:
            # Archive the organization
            organization._organization_status = 'cancelled'
            organization.is_active = False
            organization.is_archived = True
            organization.archived_at = timezone.now()
            organization.archived_by = request.user
            organization.archive_reason = archive_reason
            organization.save()

            # Archive the linked CustomUser account if it exists
            if organization.user_account:
                user_account = organization.user_account
                user_account.is_archived = True
                user_account.is_active = True
                user_account.archived_at = timezone.now()
                user_account.archived_by = request.user
                user_account.save()

                # Log both organization and user archival
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"Cancelled/archived organization: {organization.organization_name} (ID: {organization.id}) and its user account - Reason: {archive_reason}"
                )

                message = f'Organization "{organization.organization_name}" and its user account have been cancelled successfully.'
            else:
                # Log only organization archival
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"Cancelled/archived organization: {organization.organization_name} (ID: {organization.id}) - Reason: {archive_reason}"
                )

                message = f'Organization "{organization.organization_name}" has been cancelled successfully.'

            return JsonResponse({
                'success': True,
                'message': message,
                'organization_id': organization.id,
                'user_account_archived': bool(organization.user_account)
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Failed to cancel organization: {str(e)}'
            }, status=500)


class OrganizationReactivateView(LoginRequiredMixin, UpdateView):
    model = Organization
    fields = []
    template_name = 'osas/modals/reactive-organization.html'

    def get_queryset(self):
        return Organization.objects.filter(
            _organization_status='cancelled',
            is_archived=False
        )

    def form_valid(self, form):
        organization = self.object

        # Change status to pending for re-approval
        organization._organization_status = 'pending'
        organization.organization_rejection_reason = ''
        organization.organization_approved_by = None
        organization.organization_approved_at = None

        today = timezone.now().date()
        if organization.organization_valid_until < today:
            # Set new validity period (1 year from today)
            organization.organization_valid_from = today
            organization.organization_valid_until = today + timedelta(days=365)

        organization.save()

        # UserActivityLog
        UserActivityLog.objects.create(
            user=self.request.user,
            activity=f"Reactivated organization: {organization.organization_name} (ID: {organization.id}) - Status changed from cancelled to pending"
        )

        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'"{organization.organization_name}" has been reactivated and is now pending approval.'
            })

        messages.success(self.request,
                         f'"{organization.organization_name}" has been reactivated and is now pending approval.')
        return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': 'Failed to reactivate organization. Please try again.'
            })

        messages.error(self.request, 'Failed to reactivate organization. Please try again.')
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('organization-management')


@require_POST
def approve_organization(request, organization_id):
    try:
        organization = Organization.objects.get(id=organization_id)

        # Check permission
        if not (request.user.is_superuser or request.user.user_type == 10):
            return JsonResponse({
                'success': False,
                'error': 'Permission denied'
            }, status=403)

        # Check if can be approved
        if not organization.can_be_approved:
            return JsonResponse({
                'success': False,
                'error': 'Organization cannot be approved'
            })

        # Get certificate options
        generate_certificate = request.POST.get('generate_certificate') == 'true'
        certificate_date = request.POST.get('certificate_date')
        certificate_venue = request.POST.get('certificate_venue')

        # Approve organization
        organization.organization_approve_registration(request.user)

        # Grant Permission
        if organization.user_account:
            from django.contrib.auth.models import Permission, Group
            from django.contrib.contenttypes.models import ContentType

            # Get content types for the models
            organization_content_type = ContentType.objects.get_for_model(Organization)
            announcement_content_type = ContentType.objects.get_for_model(Announcement)
            certificate_content_type = ContentType.objects.get_for_model(Certificate)
            accomplishment_record_content_type = ContentType.objects.get_for_model(AccomplishmentRecord)
            supporting_file_content_type = ContentType.objects.get_for_model(SupportingFile)

            # Define the permissions to grant (all CRUD operations)
            permissions_to_grant = [
                # Organization permissions
                Permission.objects.get(content_type=organization_content_type, codename='add_organization'),
                Permission.objects.get(content_type=organization_content_type, codename='change_organization'),
                Permission.objects.get(content_type=organization_content_type, codename='delete_organization'),
                Permission.objects.get(content_type=organization_content_type, codename='view_organization'),

                # Announcement permissions
                Permission.objects.get(content_type=announcement_content_type, codename='add_announcement'),
                Permission.objects.get(content_type=announcement_content_type, codename='change_announcement'),
                Permission.objects.get(content_type=announcement_content_type, codename='delete_announcement'),
                Permission.objects.get(content_type=announcement_content_type, codename='view_announcement'),

                # Certificate permissions
                Permission.objects.get(content_type=certificate_content_type, codename='add_certificate'),
                Permission.objects.get(content_type=certificate_content_type, codename='change_certificate'),
                Permission.objects.get(content_type=certificate_content_type, codename='delete_certificate'),
                Permission.objects.get(content_type=certificate_content_type, codename='view_certificate'),

                # Accomplishment Record permissions
                Permission.objects.get(content_type=accomplishment_record_content_type, codename='add_accomplishmentrecord'),
                Permission.objects.get(content_type=accomplishment_record_content_type, codename='change_accomplishmentrecord'),
                Permission.objects.get(content_type=accomplishment_record_content_type, codename='delete_accomplishmentrecord'),
                Permission.objects.get(content_type=accomplishment_record_content_type, codename='view_accomplishmentrecord'),

                # Supporting File permissions
                Permission.objects.get(content_type=supporting_file_content_type, codename='add_supportingfile'),
                Permission.objects.get(content_type=supporting_file_content_type, codename='change_supportingfile'),
                Permission.objects.get(content_type=supporting_file_content_type, codename='delete_supportingfile'),
                Permission.objects.get(content_type=supporting_file_content_type, codename='view_supportingfile'),
            ]

            # Grant all permissions to the organization user
            organization.user_account.user_permissions.add(*permissions_to_grant)
            organization.user_account.save()

            # Also update user type to organization user if not already set
            if organization.user_account.user_type != 15:
                organization.user_account.user_type = 15
                organization.user_account.save()

        certificate_url = None
        certificate_instance = None

        # Generate certificate if requested
        if generate_certificate:
            try:
                certificate_instance = generate_certificate_png(
                    organization,
                    certificate_date=certificate_date,
                    venue=certificate_venue,
                    generated_by=request.user
                )
                if certificate_instance and certificate_instance.certificate_file:
                    certificate_url = certificate_instance.certificate_file.url
            except Exception as e:
                print(f"Certificate generation failed: {e}")
                # Continue without certificate

        # UserActivityLog
        UserActivityLog.objects.create(
            user=request.user,
            activity=f"Approved organization: {organization.organization_name} (ID: {organization.id}) - Certificate generated: {generate_certificate} - Full permissions granted including Accomplishment Reports"
        )

        return JsonResponse({
            'success': True,
            'message': f'"{organization.organization_name}" has been approved and granted full access permissions including Accomplishment Reports!',
            'certificate_url': certificate_url,
            'organization_name': organization.organization_name,
            'certificate_id': certificate_instance.id if certificate_instance else None,
            'permissions_granted': [
                'Organization Management',
                'Announcement Management',
                'Certificate Management',
                'Accomplishment Report Management',
                'Supporting File Management'
            ]
        })

    except Organization.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Organization not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@require_POST
@login_required
def renew_organization(request, organization_id):
    try:
        organization = Organization.objects.get(id=organization_id)

        # Check if user has permission
        if not (request.user.is_superuser or request.user.user_type == 10 or request.user.has_perm(
                'osas.change_organization')):
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to renew organizations'
            }, status=403)

        # Validate that organization needs renewal
        if not organization.organization_needs_renewal and organization.organization_status != 'expired':
            return JsonResponse({
                'success': False,
                'error': 'Organization does not need renewal yet'
            }, status=400)

        # Get form data
        organization_description = request.POST.get('organization_description')
        organization_mission = request.POST.get('organization_mission')
        organization_vision = request.POST.get('organization_vision')
        organization_valid_from = request.POST.get('organization_valid_from')
        organization_valid_until = request.POST.get('organization_valid_until')
        renew_count = int(request.POST.get('renew_count', organization.renew_count + 1))

        # Update organization information if provided
        if organization_description:
            organization.organization_description = organization_description
        if organization_mission:
            organization.organization_mission = organization_mission
        if organization_vision:
            organization.organization_vision = organization_vision

        # Update validity period
        if organization_valid_from:
            organization.organization_valid_from = organization_valid_from
        if organization_valid_until:
            organization.organization_valid_until = organization_valid_until

        # Update renew count
        organization.renew_count = renew_count

        # Set status to pending for re-approval
        organization.organization_status = 'pending'

        # Update created_at to current time
        organization.created_at = timezone.now()

        # Handle file uploads - ALL DOCUMENTS ARE OPTIONAL DURING RENEWAL
        all_file_fields = [
            'organization_calendar_activities',
            'organization_cbl',
            'organization_ar',
            'organization_cog',
            'organization_adviser_cv',
            'organization_group_picture',
            'organization_list_members',
            'organization_acceptance_letter',
            'organization_previous_calendar',
            'organization_good_moral',
            'organization_member_biodata',
            'organization_financial_report',
            'organization_coa'
        ]

        for field in all_file_fields:
            file = request.FILES.get(field)
            if file:
                # Validate file size based on file type
                max_size = 10 * 1024 * 1024  # 10MB default for documents
                if field in ['organization_logo', 'organization_group_picture']:
                    max_size = 5 * 1024 * 1024  # 5MB for images

                if file.size > max_size:
                    size_text = "5MB" if max_size == 5 * 1024 * 1024 else "10MB"
                    return JsonResponse({
                        'success': False,
                        'error': f'File size too large for {field.replace("organization_", "").replace("_", " ").title()}. Maximum size is {size_text}.'
                    }, status=400)

                # Validate file extensions
                if field in ['organization_logo', 'organization_group_picture']:
                    if not file.name.lower().endswith(('.jpg', '.jpeg', '.png')):
                        return JsonResponse({
                            'success': False,
                            'error': f'Invalid file type for {field.replace("organization_", "").replace("_", " ").title()}. Only JPG, JPEG, and PNG files are allowed.'
                        }, status=400)
                else:
                    if not file.name.lower().endswith(('.pdf', '.doc', '.docx')):
                        return JsonResponse({
                            'success': False,
                            'error': f'Invalid file type for {field.replace("organization_", "").replace("_", " ").title()}. Only PDF, DOC, and DOCX files are allowed.'
                        }, status=400)

                setattr(organization, field, file)

        # Save the organization
        organization.save()

        # UserActivityLog
        UserActivityLog.objects.create(
            user=request.user,
            activity=f"Renewed organization: {organization.organization_name} (ID: {organization.id}) - Renewal count: {renew_count}"
        )

        # Also keep the original print log for backup
        print(f"Organization {organization.organization_name} renewed by {request.user.username}")

        return JsonResponse({
            'success': True,
            'message': f'Organization "{organization.organization_name}" renewal submitted successfully! Status set to pending for re-approval.'
        })

    except Organization.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Organization not found'
        }, status=404)
    except Exception as e:
        print(f"Error renewing organization: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


def certificate_details(request, certificate_id):
    try:
        certificate = get_object_or_404(Certificate, id=certificate_id)

        # Check permissions
        if not (request.user.is_superuser or
                request.user.user_type in [1, 10] or
                certificate.organization.username == request.user.username):
            return JsonResponse({
                'success': False,
                'message': 'Permission denied'
            }, status=403)

        data = {
            'success': True,
            'certificate': {
                'id': certificate.id,
                'organization_name': certificate.organization.organization_name,
                'organization_type': certificate.organization.organization_type,
                'organization_type_display': certificate.organization.get_organization_type_display(),
                'issue_date': certificate.issue_date.isoformat(),
                'venue': certificate.venue,
                'generated_by': certificate.generated_by.get_full_name() if certificate.generated_by else 'System',
                'created_at': certificate.created_at.isoformat(),
                'certificate_url': certificate.certificate_file.url if certificate.certificate_file else None,
            }
        }
        return JsonResponse(data)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


def download_certificate(request, certificate_id):
    from django.http import FileResponse
    from django.shortcuts import get_object_or_404

    certificate = get_object_or_404(Certificate, id=certificate_id)

    # Check permissions
    if not (request.user.is_superuser or
            request.user.user_type in [1, 10] or
            certificate.organization.username == request.user.username):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Permission denied")

    if certificate.certificate_file:
        response = FileResponse(certificate.certificate_file.open(),
                                as_attachment=True,
                                filename=f"certificate_{certificate.organization.organization_acronym}_{certificate.issue_date}.{certificate.certificate_file.name.split('.')[-1]}")
        return response
    else:
        from django.http import HttpResponseNotFound
        return HttpResponseNotFound("Certificate file not found")


class AccomplishmentCreateView(LoginRequiredMixin, CreateView):
    model = AccomplishmentRecord
    form_class = AccomplishmentRecordForm
    success_url = reverse_lazy('dashboard')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_organization(self):
        if self.request.user.user_type == 15:  # Only organization users
            try:
                if hasattr(self.request.user, 'organization_account'):
                    return self.request.user.organization_account
            except Organization.DoesNotExist:
                pass
        return None

    def form_valid(self, form):
        organization = self.get_organization()

        # Set organization only for organization users
        if organization:
            form.instance.organization = organization
        else:
            form.instance.organization = None

        form.instance.submitted_by = self.request.user

        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                accomplishment = form.save()
                print(f"Accomplishment saved with ID: {accomplishment.id}")
                print(f"Accomplishment organization: {accomplishment.organization}")

                # Handle multiple supporting files
                supporting_files = self.request.FILES.getlist('supporting_files')
                for file in supporting_files:
                    if file:  # Only process if file exists
                        # Validate file size
                        if file.size > 10 * 1024 * 1024:
                            continue  # Skip files that are too large

                        # Validate file extension
                        valid_extensions = ['pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png', 'xls', 'xlsx', 'mp4', 'avi',
                                            'mov']
                        file_extension = file.name.split('.')[-1].lower()
                        if file_extension not in valid_extensions:
                            continue  # Skip unsupported file types

                        SupportingFile.objects.create(
                            accomplishment_record=accomplishment,
                            file=file,
                            description=f"Supporting file: {file.name}"
                        )

                # Log the activity
                UserActivityLog.objects.create(
                    user=self.request.user,
                    activity=f"{self.request.user.first_name} submitted an accomplishment report: {accomplishment.title}"
                )

                response_data = {
                    'success': True,
                    'accomplishment': {
                        'id': accomplishment.id,
                        'title': accomplishment.title,
                        'record_type': accomplishment.get_record_type_display(),
                        'date_conducted': accomplishment.date_conducted.strftime('%Y-%m-%d'),
                        'semester': accomplishment.get_semester_display(),
                        'school_year': accomplishment.school_year,
                    }
                }

                # Add organization info if available
                if accomplishment.organization:
                    response_data['accomplishment']['organization'] = accomplishment.organization.organization_name
                else:
                    response_data['accomplishment']['organization'] = "OSAS Staff Report"

                return JsonResponse(response_data)

            except Exception as e:
                print(f"Error saving accomplishment: {str(e)}")
                return JsonResponse({
                    'success': False,
                    'message': f'Error saving accomplishment report: {str(e)}'
                }, status=400)

        return super().form_valid(form)

    def form_invalid(self, form):
        print("Form invalid with errors:", form.errors)
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = {field: [str(error) for error in errors] for field, errors in form.errors.items()}
            return JsonResponse({'success': False, 'errors': errors}, status=400)
        return super().form_invalid(form)


class AccomplishmentRecordDetailView(LoginRequiredMixin, DetailView):
    model = AccomplishmentRecord

    def get_queryset(self):
        return AccomplishmentRecord.objects.select_related(
            'submitted_by',
            'organization'
        ).prefetch_related('supporting_files')

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Get supporting files
            supporting_files = self.object.supporting_files.all()
            supporting_files_data = []
            for file in supporting_files:
                supporting_files_data.append({
                    'id': file.id,
                    'file_name': file.filename,
                    'file_url': file.file.url,
                    'file_size': file.file.size,
                    'description': file.description,
                    'uploaded_at': file.uploaded_at.strftime('%Y-%m-%d %H:%M')
                })

            # Prepare main report data
            main_report_data = {
                'file_name': self.object.main_report.name.split('/')[-1],
                'file_url': self.object.main_report.url,
                'file_size': self.object.main_report.size
            }

            # Debug school year
            print(f"DEBUG - School Year from object: '{self.object.school_year}'")
            print(f"DEBUG - School Year exists: {bool(self.object.school_year)}")

            fallback_school_year = self._get_school_year_from_date(self.object.date_conducted)
            print(f"DEBUG - Fallback school year: '{fallback_school_year}'")
            school_year_to_use = self.object.school_year if self.object.school_year else fallback_school_year

            # Prepare report data
            report_data = {
                'id': self.object.id,
                'title': self.object.title,
                'record_type': self.object.record_type,
                'record_type_display': self.object.get_record_type_display(),
                'semester': self.object.semester,
                'semester_display': self.object.get_semester_display(),
                'date_conducted': self.object.date_conducted.strftime('%Y-%m-%d'),
                'date_conducted_year': self.object.date_conducted.year,
                'school_year': school_year_to_use,
                'venue': self.object.venue,
                'objectives': self.object.objectives,
                'outcomes': self.object.outcomes,
                'number_of_participants': self.object.number_of_participants,
                'duration_hours': str(self.object.duration_hours),
                'budget_utilized': str(self.object.budget_utilized) if self.object.budget_utilized else None,
                'submitted_by_name': self.object.submitted_by_name,
                'organization_name': self.object.organization_name,
                'organization_acronym': self.object.organization_acronym,
                'created_at': self.object.created_at.strftime('%Y-%m-%d %H:%M'),
                'main_report': main_report_data,
                'supporting_files': supporting_files_data,
                'has_supporting_files': self.object.has_supporting_files,
                'display_period': self.object.display_period
            }

            print(f"DEBUG - Final school year in response: '{report_data['school_year']}'")

            return JsonResponse({
                'success': True,
                'report': report_data
            })
        return super().get(request, *args, **kwargs)

    def _get_school_year_from_date(self, date_conducted):
        year = date_conducted.year
        if date_conducted.month <= 6:  # Jan-June
            return f"{year-1}-{year}"
        else:  # July-Dec
            return f"{year}-{year+1}"


class AccomplishmentRecordUpdateView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        # Allow report owners or admins to edit
        report = get_object_or_404(AccomplishmentRecord, pk=self.kwargs['pk'])
        return (self.request.user == report.submitted_by or
                self.request.user.is_superuser or
                self.request.user.user_type in [1, 10])  # OSAS Admin or Superuser

    def handle_no_permission(self):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to edit this report.'
            }, status=403)
        return super().handle_no_permission()

    def get(self, request, *args, **kwargs):
        try:
            report = get_object_or_404(AccomplishmentRecord, pk=kwargs['pk'])

            # Get supporting files
            supporting_files = report.supporting_files.all()
            supporting_files_data = []
            for file in supporting_files:
                supporting_files_data.append({
                    'id': file.id,
                    'file_name': file.filename,
                    'file_url': file.file.url,
                    'file_size': file.file.size,
                    'description': file.description,
                    'uploaded_at': file.uploaded_at.strftime('%Y-%m-%d %H:%M')
                })

            # Prepare main report data
            main_report_data = {
                'id': report.id,
                'file_name': report.main_report.name.split('/')[-1],
                'file_url': report.main_report.url,
                'file_size': report.main_report.size
            }

            report_data = {
                'id': report.id,
                'title': report.title or '',
                'record_type': report.record_type,
                'record_type_display': report.get_record_type_display(),
                'semester': report.semester,
                'semester_display': report.get_semester_display(),
                'date_conducted': report.date_conducted.strftime('%Y-%m-%d'),
                'school_year': report.school_year,
                'venue': report.venue or '',
                'objectives': report.objectives or '',
                'outcomes': report.outcomes or '',
                'number_of_participants': report.number_of_participants,
                'duration_hours': str(report.duration_hours),
                'budget_utilized': str(report.budget_utilized) if report.budget_utilized else '0.00',
                'submitted_by_name': report.submitted_by_name,
                'organization_name': report.organization_name,
                'organization_id': report.organization.id if report.organization else None,
                'created_at': report.created_at.strftime('%Y-%m-%d %H:%M'),
                'updated_at': report.updated_at.strftime('%Y-%m-%d %H:%M'),
                'main_report': main_report_data,
                'supporting_files': supporting_files_data,
                'has_supporting_files': report.has_supporting_files,
            }

            return JsonResponse({
                'success': True,
                'report': report_data
            })
        except Exception as e:
            logger.error(f"Error loading accomplishment report for editing: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'Failed to load report details: {str(e)}'
            }, status=500)

    def post(self, request, *args, **kwargs):
        try:
            report = get_object_or_404(AccomplishmentRecord, pk=kwargs['pk'])

            # Create a mutable copy of the POST data
            data = request.POST.copy()

            # Handle file fields separately
            files_data = request.FILES.copy()

            # If main report is not being updated, keep the existing one
            if 'main_report' not in files_data:
                files_data['main_report'] = report.main_report

            form = AccomplishmentRecordForm(
                data,
                files_data,
                instance=report,
                user=request.user
            )

            if form.is_valid():
                updated_report = form.save()

                # Handle new supporting files
                new_supporting_files = request.FILES.getlist('supporting_files')
                for supporting_file in new_supporting_files:
                    SupportingFile.objects.create(
                        accomplishment_record=updated_report,
                        file=supporting_file
                    )

                # Log the activity
                UserActivityLog.objects.create(
                    user=request.user,
                    activity=f"{request.user.get_full_name()} updated accomplishment report: {updated_report.title}"
                )

                return JsonResponse({
                    'success': True,
                    'message': 'Accomplishment report updated successfully!'
                })

            errors = {field: [str(error) for error in errors] for field, errors in form.errors.items()}
            return JsonResponse({'success': False, 'errors': errors}, status=400)

        except Exception as e:
            logger.error(f"Error updating accomplishment report: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': 'An unexpected error occurred while updating the report'
            }, status=500)


class SupportingFileDeleteView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        supporting_file = get_object_or_404(SupportingFile, pk=self.kwargs['pk'])
        return (self.request.user == supporting_file.accomplishment_record.submitted_by or
                self.request.user.is_superuser or
                self.request.user.user_type in [1, 10])

    def post(self, request, *args, **kwargs):
        try:
            supporting_file = get_object_or_404(SupportingFile, pk=kwargs['pk'])
            supporting_file.delete()

            return JsonResponse({
                'success': True,
                'message': 'Supporting file deleted successfully!'
            })
        except Exception as e:
            logger.error(f"Error deleting supporting file: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': 'Failed to delete supporting file'
            }, status=500)


class AccomplishmentReportArchiveView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.is_superuser or self.request.user.user_type in [1, 10, 15]

    def handle_no_permission(self):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to archive accomplishment reports.'
            }, status=403)
        return super().handle_no_permission()

    def get_object(self):
        return get_object_or_404(AccomplishmentRecord, pk=self.kwargs['pk'])

    def get(self, request, *args, **kwargs):
        report = self.get_object()

        # Check if user has permission to view this report
        if not (request.user.is_superuser or
                request.user.user_type in [1, 10] or
                report.submitted_by == request.user or
                (request.user.user_type == 15 and report.organization == request.user.organization_account)):
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to view this accomplishment report.'
            }, status=403)

        report_data = {
            'id': report.id,
            'title': report.title,
            'record_type': report.record_type,
            'record_type_display': report.get_record_type_display(),
            'date_conducted': report.date_conducted.isoformat() if report.date_conducted else None,
            'venue': report.venue,
            'semester': report.semester,
            'semester_display': report.get_semester_display(),
            'school_year': report.school_year,
            'objectives': report.objectives,
            'outcomes': report.outcomes,
            'number_of_participants': report.number_of_participants,
            'duration_hours': str(report.duration_hours) if report.duration_hours else '0.00',
            'budget_utilized': str(report.budget_utilized) if report.budget_utilized else '0.00',
            'submitted_by_name': report.submitted_by_name,
            'organization_name': report.organization_name,
            'organization_id': report.organization.id if report.organization else None,
            'created_at': report.created_at.isoformat() if report.created_at else None,
            'updated_at': report.updated_at.isoformat() if report.updated_at else None,
            'is_archived': report.is_archived,
            'supporting_files_count': report.supporting_files.count(),
            'main_report': {
                'file_name': report.main_report.name.split('/')[-1] if report.main_report else None,
                'file_url': report.main_report.url if report.main_report else None,
                'file_size': report.main_report.size if report.main_report else 0
            } if report.main_report else None
        }

        return JsonResponse({
            'success': True,
            'report': report_data
        })

    def post(self, request, *args, **kwargs):
        report = self.get_object()

        # Check if report is already archived
        if report.is_archived:
            return JsonResponse({
                'success': False,
                'error': f'Accomplishment report "{report.title}" is already archived.'
            }, status=400)

        # Check if user has permission to archive this report
        if not (request.user.is_superuser or
                request.user.user_type in [1, 10] or
                report.submitted_by == request.user or
                (request.user.user_type == 15 and report.organization == request.user.organization_account)):
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to archive this accomplishment report.'
            }, status=403)

        try:
            # Get archive reason from request
            archive_reason = request.POST.get('archive_reason', '').strip()

            # Archive the report using the model method
            report.archive(
                user=request.user,
                reason=archive_reason
            )

            # Log the activity
            UserActivityLog.objects.create(
                user=request.user,
                activity=f"{request.user.get_full_name()} archived accomplishment report: {report.title}" +
                         (f" - Reason: {archive_reason}" if archive_reason else "")
            )

            message = f'Accomplishment report "{report.title}" has been archived successfully.'

            return JsonResponse({
                'success': True,
                'message': message,
                'report_id': report.id
            })

        except Exception as e:
            logger.error(f"Error archiving accomplishment report {report.id}: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'Failed to archive accomplishment report: {str(e)}'
            }, status=500)